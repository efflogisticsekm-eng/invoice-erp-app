import os
import sys
import time
import smtplib
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.encoders import encode_base64
import pandas as pd
import numpy as np
import requests
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 1. Load Configurations from Env
ERP_USERNAME = os.getenv("ERP_USERNAME")
ERP_PASSWORD = os.getenv("ERP_PASSWORD")
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
SENDER_EMAIL = os.getenv("SENDER_EMAIL")
SENDER_PASSWORD = os.getenv("SENDER_PASSWORD")
RECEIVER_EMAIL = os.getenv("RECEIVER_EMAIL")

# Default settings if no env (e.g. for local testing, fallback to local .env values if present)
if not ERP_USERNAME or not ERP_PASSWORD:
    # Try loading from local .env files
    from dotenv import load_dotenv
    from pathlib import Path
    env_path = Path(__file__).parent.parent / "INDIA" / "consignee-app" / ".env"
    load_dotenv(dotenv_path=env_path)
    ERP_USERNAME = os.getenv("ERP_USERNAME")
    ERP_PASSWORD = os.getenv("ERP_PASSWORD")
    SUPABASE_URL = os.getenv("SUPABASE_URL")
    SUPABASE_KEY = os.getenv("SUPABASE_KEY")
    SENDER_EMAIL = os.getenv("SENDER_EMAIL")
    SENDER_PASSWORD = os.getenv("SENDER_PASSWORD")
    RECEIVER_EMAIL = os.getenv("RECEIVER_EMAIL")

# Strip surrounding quotes from env variables (handles user paste errors in GitHub Secrets)
if ERP_USERNAME: ERP_USERNAME = ERP_USERNAME.strip("'\"")
if ERP_PASSWORD: ERP_PASSWORD = ERP_PASSWORD.strip("'\"")
if SUPABASE_URL: SUPABASE_URL = SUPABASE_URL.strip("'\"")
if SUPABASE_KEY: SUPABASE_KEY = SUPABASE_KEY.strip("'\"")
if SENDER_EMAIL: SENDER_EMAIL = SENDER_EMAIL.strip("'\"")
if SENDER_PASSWORD: SENDER_PASSWORD = SENDER_PASSWORD.strip("'\"")
if RECEIVER_EMAIL: RECEIVER_EMAIL = RECEIVER_EMAIL.strip("'\"")

DOWNLOAD_DIR = os.path.expanduser("~/Downloads/erp_temp_downloads")
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

# 2. Date Parsing Helper
def parse_date(val):
    if pd.isna(val) or not val:
        return None
    val_str = str(val).strip().split()[0]
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%d.%m.%Y'):
        try:
            return datetime.strptime(val_str, fmt)
        except ValueError:
            continue
    try:
        return pd.to_datetime(val_str).to_pydatetime()
    except Exception:
        return None

# 3. Delay Calculator
def calculate_delay(start_str, end_str, consignor, holidays, exclude_sundays):
    if pd.isna(end_str) or not str(end_str).strip():
        return None
    
    start = parse_date(start_str)
    end = parse_date(end_str)
    
    if not start or not end:
        return None
    
    diff_days = (end.date() - start.date()).days
    calculated_delay = diff_days
    
    # Exclude Sundays & Holidays
    temp = start
    normalized_holidays = []
    for h in holidays:
        pd_h = parse_date(h)
        if pd_h:
            normalized_holidays.append(pd_h.date())
            
    while temp.date() < end.date():
        temp_date = temp.date()
        is_sunday = temp.weekday() == 6 # Sunday in python is 6
        
        if temp_date in normalized_holidays or (exclude_sundays and is_sunday and temp_date not in normalized_holidays):
            calculated_delay -= 1
        temp += pd.Timedelta(days=1)
        
    return max(0, calculated_delay)

# 4. Fetch supervisor mappings from Supabase
def fetch_supervisor_mappings():
    print("Fetching supervisor/branch mappings from Supabase...")
    try:
        url = f"{SUPABASE_URL}/rest/v1/supervisor_branch_mapping?select=*"
        headers = {
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}"
        }
        res = requests.get(url, headers=headers)
        res.raise_for_status()
        data = res.json()
        print(f"Loaded {len(data)} supervisor mappings.")
        return {item['supervisor_name'].strip().upper(): item['branch'].strip().upper() for item in data}
    except Exception as e:
        print(f"Error fetching supervisor mappings: {e}. Defaulting to empty mapping.")
        return {}

# 5. Playwright ERP Download
def download_erp_reports():
    print("Starting Playwright ERP download flow...")
    despatch_url = "https://eff.aadhocc.in/eff_2021/main/effdespatch"
    lr_url = "https://eff.aadhocc.in/eff_2021/main/lr/"
    
    despatch_file_path = os.path.join(DOWNLOAD_DIR, "despatch_raw.xlsx")
    lr_file_path = os.path.join(DOWNLOAD_DIR, "lr_raw.xlsx")
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()
        
        # Listen to browser console and page errors
        page.on("console", lambda msg: print(f"Browser Console: {msg.text}"))
        page.on("pageerror", lambda err: print(f"Browser Page Error: {err}"))
        
        try:
            # Step 1: Navigating to main login page directly
            main_login_url = "https://eff.aadhocc.in/eff_2021/login"
            print(f"Navigating to login page: {main_login_url}...")
            page.goto(main_login_url)
            page.wait_for_load_state("load")
            
            # Check for login inputs
            if page.locator("#login_user_id").count() > 0 or "login" in page.url.lower():
                print("Performing ERP Login...")
                if page.locator("#login_user_id").count() > 0:
                    page.fill("#login_user_id", ERP_USERNAME)
                else:
                    page.fill("input[type='text']", ERP_USERNAME)
                    
                if page.locator("#login_password").count() > 0:
                    page.fill("#login_password", ERP_PASSWORD)
                else:
                    page.fill("input[type='password']", ERP_PASSWORD)
                    
                # Take screenshot to verify fields are filled
                page.screenshot(path=os.path.join(DOWNLOAD_DIR, "filled.png"), timeout=5000)
                print("Filled credentials screenshot saved.")
                
                # Submit login form
                submit_button = page.locator("form#login_form button[type='submit'], button[type='submit']")
                if submit_button.count() > 0:
                    submit_button.first.click()
                else:
                    page.keyboard.press("Enter")
                    
                # Take screenshot immediately after click attempt
                page.wait_for_timeout(2000) # wait 2s to allow page load/update
                page.screenshot(path=os.path.join(DOWNLOAD_DIR, "clicked.png"), timeout=5000)
                print("Clicked submit screenshot saved.")
                
                # Wait for navigation to complete (expecting to leave the login page)
                try:
                    page.wait_for_url("**/login", exclude=True, timeout=10000)
                    page.wait_for_load_state("load")
                    print("Login complete. Current URL:", page.url)
                except Exception as nav_err:
                    print("Navigation timeout or did not leave login page. Current URL:", page.url)
                    # Check if error message is displayed
                    if page.locator("#auth_msg").count() > 0:
                        error_text = page.locator("#auth_msg").inner_text()
                        print(f"ERP Login Error Message: {error_text.strip()}")
            
            # Navigate to Despatch page
            print(f"Navigating to Despatch Listing: {despatch_url}...")
            page.goto(despatch_url)
            page.wait_for_load_state("load")
            
            # Read fromDate and toDate input values pre-filled on the page
            page.locator("#fromDate").wait_for(state="visible", timeout=15000)
            from_date = page.locator("#fromDate").get_attribute("value")
            to_date = page.locator("#toDate").get_attribute("value")
            print(f"Found pre-filled dates: fromDate={from_date}, toDate={to_date}")
            
            # Construct the correct download URL and set it as the href of the export link
            correct_href = f"https://eff.aadhocc.in/eff_2021/main/effdespatch/exportDespatchExcel?despatch_number=&location_id=&lr_number=&from_date={from_date}&to_date={to_date}&delivery_staff_search="
            print(f"Setting export link href to: {correct_href}")
            page.evaluate(f"document.querySelector('a.exportDespatchExcel').href = '{correct_href}'")
            
            # Download Despatch Report by clicking the export button
            print("Downloading Despatch raw report...")
            despatch_btn = page.locator("a.exportDespatchExcel").first
            despatch_btn.wait_for(state="visible", timeout=15000)
            
            with page.expect_download(timeout=60000) as download_info:
                despatch_btn.click(no_wait_after=True)
            download = download_info.value
            download.save_as(despatch_file_path)
            print("Despatch report saved to:", despatch_file_path)
            
            # Navigate to LR page
            print(f"Navigating to LR Report page: {lr_url}...")
            page.goto(lr_url)
            page.wait_for_load_state("load")
            
            # Download LR Report by clicking the export button
            print("Downloading LR raw report...")
            lr_btn = page.locator("a.export_lr_excel, button#excelExport1, #excelExport1").first
            lr_btn.wait_for(state="visible", timeout=15000)
            
            with page.expect_download(timeout=60000) as download_info_lr:
                lr_btn.click(no_wait_after=True)
            download_lr = download_info_lr.value
            download_lr.save_as(lr_file_path)
            print("LR raw report saved to:", lr_file_path)
            
            return lr_file_path, despatch_file_path
            
        except Exception as e:
            print("Error downloading from ERP:", e)
            screenshot_path = os.path.join(DOWNLOAD_DIR, "error_screenshot.png")
            try:
                page.screenshot(path=screenshot_path, timeout=5000)
                print("Screenshot saved to:", screenshot_path)
            except Exception as ss_err:
                print("Could not save error screenshot:", ss_err)
            browser.close()
            raise e
        finally:
            browser.close()

# 6. Apply openpyxl styles
def apply_styles(ws, row_count, col_count, sheet_type="default"):
    # Styles Definition
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # slate-800
    header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
    
    data_font = Font(name="Arial", size=10)
    data_border = Border(
        bottom=Side(style="thin", color="E2E8F0"), # slate-200
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0")
    )
    
    # Header styles
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = Border(
            top=Side(style="thin", color="000000"),
            bottom=Side(style="medium", color="000000"),
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000")
        )
        
    # Data row styles
    for row in range(2, row_count + 1):
        # Zebra striping for even rows
        row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") if row % 2 == 0 else PatternFill(fill_type=None)
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = data_border
            if row % 2 == 0:
                cell.fill = row_fill
            
            # Alignments
            if sheet_type == "summary":
                cell.alignment = Alignment(vertical="center", horizontal="left")
            elif col == 1:
                cell.alignment = Alignment(vertical="center", horizontal="left")
            else:
                cell.alignment = Alignment(vertical="center", horizontal="center")

# 7. Generate Excel file
def generate_excel_report(lr_file, despatch_file, supervisor_map):
    print("Generating Interactive_Delivery_Report.xlsx...")
    
    # Read raw sheets
    def load_df(file_path):
        with open(file_path, "rb") as f:
            head = f.read(4)
        if head == b"PK\x03\x04" or head == b"\xd0\xcf\x11\xe0":
            return pd.read_excel(file_path)
        else:
            for enc in ["utf-8", "latin1", "utf-8-sig"]:
                try:
                    return pd.read_csv(file_path, encoding=enc)
                except Exception:
                    continue
            return pd.read_csv(file_path)
            
    df_lr = load_df(lr_file)
    df_despatch = load_df(despatch_file)
    
    # Exclude Sundays (True by default)
    exclude_sundays = True
    custom_holidays = [] # Can be populated if needed, or fetched from database
    
    # Clean LR column names
    df_lr.columns = [str(c).strip().upper() for c in df_lr.columns]
    df_despatch.columns = [str(c).strip().upper() for c in df_despatch.columns]
    
    # Map LRs to despatch info for faster lookup
    # Key fields in despatch report: DESPATCH DATE, DISPATCH DATE, DESPATCH NO, DISPATCH NO, LR NO, LRNO
    lr_col_desp = next((c for c in df_despatch.columns if c in ['LR NO', 'LRNO', 'LR_NO']), None)
    desp_date_col = next((c for c in df_despatch.columns if c in ['DESPATCH DATE', 'DISPATCH DATE', 'DESPATCH_DATE']), None)
    desp_no_col = next((c for c in df_despatch.columns if c in ['DESPATCH NO', 'DISPATCH NO', 'DESPATCH_NO', 'DISPATCH_NO']), None)
    driver_col = next((c for c in df_despatch.columns if c in ['DELIVERY DRIVER', 'DRIVER', 'DRIVER_NAME']), None)
    
    despatch_map = {}
    if lr_col_desp and desp_date_col:
        for _, r in df_despatch.iterrows():
            lr = str(r[lr_col_desp]).strip()
            despatch_map[lr] = {
                "despatch_date": str(r[desp_date_col]).strip() if pd.notna(r[desp_date_col]) else "",
                "despatch_no": str(r[desp_no_col]).strip() if desp_no_col and pd.notna(r[desp_no_col]) else "",
                "driver": str(r[driver_col]).strip() if driver_col and pd.notna(r[driver_col]) else ""
            }
            
    # Process LR rows
    lr_col = next((c for c in df_lr.columns if c in ['LR NO', 'LRNO', 'LR_NUMBER', 'LR_NO']), None)
    date_col = next((c for c in df_lr.columns if c in ['DATE', 'LR DATE', 'LR_DATE']), None)
    del_time_col = next((c for c in df_lr.columns if c in ['DELIVERY TIME', 'DELIVERY_TIME', 'DELIVERED_DATE']), None)
    consignor_col = next((c for c in df_lr.columns if c in ['CONSIGNOR', 'CONSIGNOR_NAME']), None)
    consignee_col = next((c for c in df_lr.columns if c in ['CONSIGNEE', 'CONSIGNEE_NAME']), None)
    dest_col = next((c for c in df_lr.columns if c in ['DESTINATION', 'PLACE', 'AREA']), None)
    status_col = next((c for c in df_lr.columns if c in ['LR STATUS', 'LRSTATUS', 'STATUS']), None)
    sup_col = next((c for c in df_lr.columns if c in ['LD SUPERVISOR', 'SUPERVISOR', 'LD_SUPERVISOR']), None)
    box_col = next((c for c in df_lr.columns if c in ['BOX QTY', 'BOXQTY', 'BOXES', 'QUANTITY']), None)
    
    processed_lrs = []
    cancelled_lrs = []
    
    consignor_counts = {}
    dest_counts = {}
    branch_counts = {}
    
    raw_total = len(df_lr)
    cancelled_count = 0
    total_excluded_consignors = 0
    delivered_count = 0
    open_count = 0
    despatched_count = 0
    
    delay_counts = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 'more': 0, 'invalid': 0}
    
    branches_list = set(supervisor_map.values())
    
    # Leaderboard trackers
    branch_stats = {}
    driver_stats = {}
    
    for _, r in df_lr.iterrows():
        lr_no = str(r[lr_col]).strip() if lr_col else ""
        raw_status = str(r[status_col]).strip() if status_col else "Open"
        consignor = str(r[consignor_col]).strip() if consignor_col else "UNKNOWN"
        consignee = str(r[consignee_col]).strip() if consignee_col else "UNKNOWN"
        area = str(r[dest_col]).strip() if dest_col else "UNKNOWN"
        lr_date = str(r[date_col]).strip() if date_col else ""
        del_time = str(r[del_time_col]).strip() if del_time_col else ""
        box_qty = int(r[box_col]) if box_col and pd.notna(r[box_col]) else 0
        
        # Mapped Status
        status_map = {
            'Despatched': 'On transit',
            'Open': 'Not Despatched',
            'Delivered': 'Delivery Process completed.',
            'Despatched from Branch': 'Cancelled LR'
        }
        mapped_status = status_map.get(raw_status, "Not Despatched")
        if "cancelled" in raw_status.lower() or mapped_status == 'Cancelled LR':
            mapped_status = 'Cancelled LR'
            
        # Exclude EFF consignors
        if consignor.upper().startswith('EFF'):
            total_excluded_consignors += 1
            continue
            
        # Supervisor and Branch Mapping
        sup_val = str(r[sup_col]).strip().upper() if sup_col and pd.notna(r[sup_col]) else ""
        if not sup_val and lr_no in despatch_map:
            # Try getting driver name or supervisor mapping if available
            pass
        branch = supervisor_map.get(sup_val, "N/A")
        
        item = {
            "lrNo": lr_no,
            "area": area,
            "branch": branch,
            "supervisor": sup_val if sup_val else "N/A",
            "consignor": consignor,
            "consignee": consignee,
            "date": lr_date,
            "deliveryTime": del_time,
            "boxQty": box_qty,
            "status": mapped_status,
            "delay": None
        }
        
        # Check Cancellation
        if mapped_status == 'Cancelled LR':
            cancelled_count += 1
            cancelled_lrs.append(item)
            continue
            
        # Delivery delays
        if mapped_status == 'Delivery Process completed.':
            delivered_count += 1
            despatch_info = despatch_map.get(lr_no, {})
            despatch_date = despatch_info.get("despatch_date", lr_date) # Fallback to LR date if no despatch date
            
            delay = calculate_delay(despatch_date, del_time, consignor, holidays=custom_holidays, exclude_sundays=exclude_sundays)
            item["delay"] = delay
            
            if delay is not None:
                if delay in delay_counts:
                    delay_counts[delay] += 1
                else:
                    delay_counts['more'] += 1
            else:
                delay_counts['invalid'] += 1
        elif mapped_status == 'On transit':
            despatched_count += 1
        else:
            open_count += 1
            
        processed_lrs.append(item)
        
        # Accumulate Leaderboard Stats for Branch
        b_name = branch
        if b_name not in branch_stats:
            branch_stats[b_name] = {
                "name": b_name, "totalLrs": 0, "deliveredLrs": 0, "totalBoxes": 0,
                "deliveryPoints": set(), "totalDelayDays": 0, "delaysCount": 0
            }
        bs = branch_stats[b_name]
        bs["totalLrs"] += 1
        bs["totalBoxes"] += box_qty
        p_key = f"{consignee.lower()}||{area.lower()}"
        if consignee or area:
            bs["deliveryPoints"].add(p_key)
        if mapped_status == 'Delivery Process completed.' and item["delay"] is not None:
            bs["deliveredLrs"] += 1
            bs["totalDelayDays"] += item["delay"]
            bs["delaysCount"] += 1
            
        # Accumulate Leaderboard Stats for Driver
        driver_name = despatch_map.get(lr_no, {}).get("driver", "").strip()
        if driver_name and driver_name != '-':
            split_drivers = [d.strip() for d in driver_name.split(',') if d.strip()]
            for d in split_drivers:
                if d not in driver_stats:
                    driver_stats[d] = {
                        "name": d, "totalLrs": 0, "deliveredLrs": 0, "totalBoxes": 0,
                        "deliveryPoints": set(), "totalDelayDays": 0, "delaysCount": 0
                    }
                ds = driver_stats[d]
                ds["totalLrs"] += 1
                ds["totalBoxes"] += box_qty
                if consignee or area:
                    ds["deliveryPoints"].add(p_key)
                if mapped_status == 'Delivery Process completed.' and item["delay"] is not None:
                    ds["deliveredLrs"] += 1
                    ds["totalDelayDays"] += item["delay"]
                    ds["delaysCount"] += 1

    active_total = delivered_count + open_count + despatched_count
    
    # Build Excel Workbook
    processed_file = os.path.join(DOWNLOAD_DIR, "Interactive_Delivery_Report.xlsx")
    writer = pd.ExcelWriter(processed_file, engine='openpyxl')
    
    # Helper to convert Set to count
    def get_leaderboard_row(item_stats, is_branch=True):
        rows_data = []
        for name, s in item_stats.items():
            avg_delay = (s["totalDelayDays"] / s["delaysCount"]) if s["delaysCount"] > 0 else 0
            pts = len(s["deliveryPoints"])
            # Score formula: (Delivered LRs * 10) + (Total Boxes * 1) + (Points * 20) - (Average Delay * 30)
            score = (s["deliveredLrs"] * 10) + (s["totalBoxes"] * 1) + (pts * 20) - (avg_delay * 30)
            score = max(0, int(round(score)))
            rows_data.append({
                "Name": name,
                "Type": "Branch" if is_branch else "Driver",
                "Performance Score": score,
                "Delivered LRs": s["deliveredLrs"],
                "Delivery Points": pts,
                "Delivered Boxes": s["totalBoxes"],
                "Average Delay (Days)": round(avg_delay, 1) if s["delaysCount"] > 0 else '-'
            })
        df = pd.DataFrame(rows_data)
        if not df.empty:
            df = df.sort_values(by="Performance Score", ascending=False)
            df.insert(0, "Rank", range(1, len(df) + 1))
        return df

    # Sheet 1: Overall Summary
    def get_progress_bar(val, max_val=100):
        if not max_val or max_val <= 0:
            return '░░░░░░░░░░ 0%'
        bars = 10
        filled = min(bars, max(0, int(round((val / max_val) * bars))))
        empty = bars - filled
        pct = int(round((val / max_val) * 100))
        return '█' * filled + '░' * empty + f" {pct}%"
        
    pct_delivered = f"{round((delivered_count/active_total)*100, 1)}%" if active_total > 0 else "0%"
    pct_open = f"{round((open_count/active_total)*100, 1)}%" if active_total > 0 else "0%"
    pct_transit = f"{round((despatched_count/active_total)*100, 1)}%" if active_total > 0 else "0%"
    
    summary_data = [
        ["DELIVERY DELAY REPORT - SUMMARY", "", "", ""],
        ["", "", "", ""],
        ["OVERALL STATS", "COUNT", "% (of Active)", "VISUAL CHART"],
        ["Total LRs in File", raw_total, "-", ""],
        ["Cancelled LRs", cancelled_count, "-", ""],
        ["Excluded EFF LRs", total_excluded_consignors, "-", ""],
        ["", "", "", ""],
        ["A) Total LR Count (Except Cancelled & EFF)", active_total, "100%", get_progress_bar(active_total, active_total)],
        ["Delivered", delivered_count, pct_delivered, get_progress_bar(delivered_count, active_total)],
        ["Not Despatched (Open)", open_count, pct_open, get_progress_bar(open_count, active_total)],
        ["On Transit (Despatched)", despatched_count, pct_transit, get_progress_bar(despatched_count, active_total)],
        ["", "", "", ""],
        ["DELAY BREAKDOWN (Delivered Only)", "COUNT", "% (of Active)", "VISUAL CHART"],
        ["Same Day (0)", delay_counts[0], f"{round((delay_counts[0]/active_total)*100, 1)}%" if active_total > 0 else "0%", get_progress_bar(delay_counts[0], active_total)],
        ["Next Day (1)", delay_counts[1], f"{round((delay_counts[1]/active_total)*100, 1)}%" if active_total > 0 else "0%", get_progress_bar(delay_counts[1], active_total)],
        ["2nd Day", delay_counts[2], f"{round((delay_counts[2]/active_total)*100, 1)}%" if active_total > 0 else "0%", get_progress_bar(delay_counts[2], active_total)],
        ["3rd Day", delay_counts[3], f"{round((delay_counts[3]/active_total)*100, 1)}%" if active_total > 0 else "0%", get_progress_bar(delay_counts[3], active_total)],
        ["4th Day", delay_counts[4], f"{round((delay_counts[4]/active_total)*100, 1)}%" if active_total > 0 else "0%", get_progress_bar(delay_counts[4], active_total)],
        ["5th Day", delay_counts[5], f"{round((delay_counts[5]/active_total)*100, 1)}%" if active_total > 0 else "0%", get_progress_bar(delay_counts[5], active_total)],
        ["6th Day", delay_counts[6], f"{round((delay_counts[6]/active_total)*100, 1)}%" if active_total > 0 else "0%", get_progress_bar(delay_counts[6], active_total)],
        ["7th Day", delay_counts[7], f"{round((delay_counts[7]/active_total)*100, 1)}%" if active_total > 0 else "0%", get_progress_bar(delay_counts[7], active_total)],
        ["> 7 Days", delay_counts['more'], f"{round((delay_counts['more']/active_total)*100, 1)}%" if active_total > 0 else "0%", get_progress_bar(delay_counts['more'], active_total)]
    ]
    
    if delay_counts['invalid'] > 0:
        summary_data.append(["No Date Info / Invalid", delay_counts['invalid'], f"{round((delay_counts['invalid']/active_total)*100, 1)}%" if active_total > 0 else "0%", get_progress_bar(delay_counts['invalid'], active_total)])
        
    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name="1. Overall Summary", index=False, header=False)
    
    # Sheet 2: Destination Breakdown
    dest_breakdown = []
    for dest, counts in dest_counts.items():
        pass # To be implemented via dynamic pivot
        
    # Standard pivot-like breakdowns for destinations, consignors, and branches
    def create_breakdown_df(lrs_list, key_field):
        breakdown_map = {}
        for x in lrs_list:
            val = x[key_field]
            if val not in breakdown_map:
                breakdown_map[val] = {c: 0 for c in [0,1,2,3,4,5,6,7,'more','invalid','open','transit','cancelled']}
                breakdown_map[val]['total'] = 0
            
            # Map categories
            if x["status"] == 'Cancelled LR':
                cat = 'cancelled'
            elif x["status"] == 'Not Despatched':
                cat = 'open'
            elif x["status"] == 'On transit':
                cat = 'transit'
            else:
                d = x["delay"]
                if d is None:
                    cat = 'invalid'
                elif d > 7:
                    cat = 'more'
                else:
                    cat = d
            
            breakdown_map[val][cat] += 1
            breakdown_map[val]['total'] += 1
            
        rows_data = []
        for name, cats in breakdown_map.items():
            row = {
                key_field.capitalize(): name,
                "Same Day (0)": cats[0], "Next Day (1)": cats[1],
                "2nd Day": cats[2], "3rd Day": cats[3], "4th Day": cats[4],
                "5th Day": cats[5], "6th Day": cats[6], "7th Day": cats[7],
                "> 7 Days": cats['more'], "No Date": cats['invalid'],
                "Not Despatched": cats['open'], "On Transit": cats['transit'],
                "Cancelled": cats['cancelled'], "TOTAL": cats['total']
            }
            rows_data.append(row)
        df = pd.DataFrame(rows_data)
        if not df.empty:
            df = df.sort_values(by="TOTAL", ascending=False)
        return df

    df_dest = create_breakdown_df(processed_lrs + cancelled_lrs, "area")
    df_dest.rename(columns={"Area": "Destination"}, inplace=True)
    df_dest.to_excel(writer, sheet_name="2. Destination Breakdown", index=False)
    
    df_consignor = create_breakdown_df(processed_lrs + cancelled_lrs, "consignor")
    df_consignor.rename(columns={"Consignor": "Consignor Name"}, inplace=True)
    df_consignor.to_excel(writer, sheet_name="3. Consignor Breakdown", index=False)
    
    # Sheet 4: Active LRs
    df_active = pd.DataFrame([{
        "LR NO": x["lrNo"], "AREA": x["area"], "BRANCH": x["branch"], "SUPERVISOR": x["supervisor"],
        "CONSIGNOR": x["consignor"], "CONSIGNEE": x["consignee"], "DATE": x["date"],
        "DELIVERY TIME": x["deliveryTime"], "DELAY (DAYS)": x["delay"] if x["delay"] is not None else '-',
        "STATUS": x["status"]
    } for x in processed_lrs])
    df_active.to_excel(writer, sheet_name="4. Active LRs", index=False)
    
    # Sheet 5: Cancelled LRs
    df_cancelled = pd.DataFrame([{
        "LR NO": x["lrNo"], "AREA": x["area"], "BRANCH": x["branch"], "SUPERVISOR": x["supervisor"],
        "CONSIGNOR": x["consignor"], "CONSIGNEE": x["consignee"], "DATE": x["date"],
        "DELIVERY TIME": x["deliveryTime"], "STATUS": x["status"]
    } for x in cancelled_lrs])
    df_cancelled.to_excel(writer, sheet_name="5. Cancelled LRs", index=False)
    
    # Sheet 6: Branch Breakdown
    df_branch = create_breakdown_df(processed_lrs + cancelled_lrs, "branch")
    df_branch.to_excel(writer, sheet_name="6. Branch Breakdown", index=False)
    
    # Sheet 7: Despatch Summary Report
    # We aggregate despatch raw report data
    despatch_rows = []
    # Key fields in despatch report: DESPATCH DATE, DISPATCH DATE, DESPATCH NO, DISPATCH NO, LR NO, LRNO
    if not df_despatch.empty:
        # We group df_despatch by branch, despatch date, despatch no, and delivery driver
        desp_no_field = next((c for c in df_despatch.columns if c in ['DESPATCH NO', 'DISPATCH NO', 'DESPATCH_NO', 'DISPATCH_NO']), None)
        desp_date_field = next((c for c in df_despatch.columns if c in ['DESPATCH DATE', 'DISPATCH DATE', 'DESPATCH_DATE']), None)
        driver_field = next((c for c in df_despatch.columns if c in ['DELIVERY DRIVER', 'DRIVER', 'DRIVER_NAME']), None)
        box_field = next((c for c in df_despatch.columns if c in ['BOX QTY', 'BOXQTY', 'BOXES', 'QUANTITY']), None)
        dest_field = next((c for c in df_despatch.columns if c in ['DESTINATION', 'PLACE', 'AREA']), None)
        
        # Build unique list of despatches
        df_desp_grouped = df_despatch.groupby([desp_no_field, desp_date_field, driver_field] if desp_no_field and desp_date_field and driver_field else [df_despatch.columns[0]])
        for keys, group in df_desp_grouped:
            desp_no = keys[0] if isinstance(keys, tuple) else keys
            desp_date = keys[1] if isinstance(keys, tuple) else ""
            driver = keys[2] if isinstance(keys, tuple) else ""
            
            # Find branch from supervisors of LRs in this group
            gp_lrs = group[lr_col_desp].dropna().astype(str).tolist() if lr_col_desp in group.columns else []
            gp_supervisors = df_lr[df_lr[lr_col].astype(str).isin(gp_lrs)][sup_col].dropna().unique() if lr_col in df_lr.columns and sup_col in df_lr.columns else []
            gp_branches = [supervisor_map.get(str(s).strip().upper(), "N/A") for s in gp_supervisors]
            branch = gp_branches[0] if gp_branches else "N/A"
            
            box_qty_tot = group[box_field].sum() if box_field in group.columns else 0
            del_points = group[dest_field].nunique() if dest_field in group.columns else 0
            
            # Find delivery times in LR report
            lr_del_times = df_lr[df_lr[lr_col].astype(str).isin(gp_lrs)][del_time_col].dropna().tolist() if lr_col in df_lr.columns and del_time_col in df_lr.columns else []
            lr_del_parsed = [parse_date(t) for t in lr_del_times if parse_date(t)]
            first_del = min(lr_del_parsed).strftime("%d/%m/%Y %I:%M %p") if lr_del_parsed else "-"
            last_del = max(lr_del_parsed).strftime("%d/%m/%Y %I:%M %p") if lr_del_parsed else "-"
            
            despatch_rows.append({
                "Branch": branch,
                "Despatch No": desp_no,
                "Despatch Date": desp_date,
                "Total LR count": len(group),
                "Total Delivery Point": del_points,
                "Delivery Driver": driver,
                "Box Qty": box_qty_tot,
                "First Delivery Time": first_del,
                "Last Delivery Time": last_del
            })
            
    df_desp_report = pd.DataFrame(despatch_rows)
    df_desp_report.to_excel(writer, sheet_name="7. Despatch Summary Report", index=False)
    
    # Sheet 8: Performance Leaderboard
    df_branch_leader = get_leaderboard_row(branch_stats, is_branch=True)
    df_driver_leader = get_leaderboard_row(driver_stats, is_branch=False)
    
    # Combine Branch & Driver Leaderboards with empty spacer row
    leader_rows = []
    if not df_branch_leader.empty:
        leader_rows.append(["BRANCH PERFORMANCE LEADERBOARD", "", "", "", "", "", "", ""])
        leader_rows.append(list(df_branch_leader.columns))
        leader_rows.extend(df_branch_leader.values.tolist())
        leader_rows.append([]) # spacer
        
    if not df_driver_leader.empty:
        leader_rows.append(["DRIVER PERFORMANCE LEADERBOARD", "", "", "", "", "", "", ""])
        leader_rows.append(list(df_driver_leader.columns))
        leader_rows.extend(df_driver_leader.values.tolist())
        
    df_leaderboard = pd.DataFrame(leader_rows)
    df_leaderboard.to_excel(writer, sheet_name="8. Performance Leaderboard", index=False, header=False)
    
    # Sheet 9: Supervisor Mapping
    df_sup_mapping = pd.DataFrame([{"Supervisor": k, "Branch": v} for k, v in supervisor_map.items()])
    df_sup_mapping.to_excel(writer, sheet_name="9. Supervisor Mapping", index=False)
    
    # Save Workbook
    writer.close()
    print("Workbook generated successfully!")
    
    # Apply Openpyxl design aesthetics
    wb = openpyxl.load_workbook(processed_file)
    
    # Style 1. Overall Summary
    apply_styles(wb["1. Overall Summary"], len(df_summary), 4, "summary")
    
    # Style 2. Destination Breakdown
    apply_styles(wb["2. Destination Breakdown"], len(df_dest) + 1, len(df_dest.columns))
    
    # Style 3. Consignor Breakdown
    apply_styles(wb["3. Consignor Breakdown"], len(df_consignor) + 1, len(df_consignor.columns))
    
    # Style 4. Active LRs
    apply_styles(wb["4. Active LRs"], len(df_active) + 1, len(df_active.columns))
    
    # Style 5. Cancelled LRs
    apply_styles(wb["5. Cancelled LRs"], len(df_cancelled) + 1, len(df_cancelled.columns))
    
    # Style 6. Branch Breakdown
    apply_styles(wb["6. Branch Breakdown"], len(df_branch) + 1, len(df_branch.columns))
    
    # Style 7. Despatch Summary Report
    apply_styles(wb["7. Despatch Summary Report"], len(df_desp_report) + 1, len(df_desp_report.columns))
    
    # Style 8. Performance Leaderboard
    ws_leader = wb["8. Performance Leaderboard"]
    apply_styles(ws_leader, len(df_leaderboard), 8, "leaderboard")
    
    # Style 9. Supervisor Mapping
    apply_styles(wb["9. Supervisor Mapping"], len(df_sup_mapping) + 1, len(df_sup_mapping.columns))
    
    wb.save(processed_file)
    print("Workbook styled and saved!")
    return processed_file

# 8. Email function
def email_report(processed_file_path, raw_lr_path, raw_despatch_path):
    print("Sending daily report email...")
    
    # Set up email message
    msg = MIMEMultipart()
    msg['From'] = SENDER_EMAIL
    msg['To'] = RECEIVER_EMAIL
    today_str = datetime.now().strftime("%d-%m-%Y")
    msg['Subject'] = f"Daily ERP Dispatch & Delivery Performance Report - {today_str} 📊"
    
    body = f"""Dear User,

Please find attached the daily ERP Dispatch & Delivery delay processing report sheets for {today_str}.

Included sheets in Interactive_Delivery_Report.xlsx:
1. Overall Summary (Delivered/Open/Transit ratios & delay charts)
2. Destination Breakdown
3. Consignor Breakdown
4. Active LRs
5. Cancelled LRs
6. Branch Breakdown
7. Despatch Summary Report
8. Performance Leaderboards (Branches & Drivers)
9. Supervisor Master Mapping Table

Also attached are the raw downloaded ERP sheets for your reference.

Best Regards,
ERP Daily Automation Engine ⚡
"""
    msg.attach(MIMEText(body, 'plain'))
    
    # Attach files
    files_to_attach = [
        (processed_file_path, f"Interactive_Delivery_Report_{today_str}.xlsx"),
        (raw_lr_path, f"raw_lr_data_{today_str}.xlsx"),
        (raw_despatch_path, f"raw_despatch_data_{today_str}.xlsx")
    ]
    
    for file_path, attachment_name in files_to_attach:
        if os.path.exists(file_path):
            with open(file_path, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
                encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename= {attachment_name}",
                )
                msg.attach(part)
                print(f"Attached: {attachment_name}")
                
    # Connect to Gmail SMTP
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.send_message(msg)
        server.quit()
        print("🎉 Daily report email sent successfully!")
    except Exception as e:
        print("❌ Error sending email:", e)
        raise e

# 9. Main orchestrator
def main():
    print(f"[{datetime.now()}] Starting daily report process...")
    try:
        # Download files
        lr_file, despatch_file = download_erp_reports()
        
        # Load mappings
        supervisor_map = fetch_supervisor_mappings()
        
        # Process and generate formatted workbook
        processed_file = generate_excel_report(lr_file, despatch_file, supervisor_map)
        
        # Email report
        email_report(processed_file, lr_file, despatch_file)
        
        print(f"[{datetime.now()}] All tasks completed successfully! Have a great day!")
    except Exception as e:
        print(f"[{datetime.now()}] Critical failure in daily automation run: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
