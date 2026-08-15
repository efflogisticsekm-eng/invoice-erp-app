#!/usr/bin/env python3
import os
import sys
import re
import gspread
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, timezone
from google.oauth2.service_account import Credentials
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.encoders import encode_base64

# Define directories and credentials path
BASE_DIR = "/Users/anwar/Desktop/Antigravity-Related"
CREDENTIALS_PATH = os.path.join(BASE_DIR, "ERP nxt Data collection", "Invoice_Extractor_Tool", "credentials.json")
DOWNLOAD_DIR = os.path.expanduser("~/Downloads/erp_temp_downloads")

# Load environment variables for email
SENDER_EMAIL = os.getenv("SENDER_EMAIL")
SENDER_PASSWORD = os.getenv("SENDER_PASSWORD")
RECEIVER_EMAIL = os.getenv("RECEIVER_EMAIL")

if not SENDER_EMAIL or not SENDER_PASSWORD:
    # Try loading from local .env files
    from dotenv import load_dotenv
    from pathlib import Path
    env_path = Path(BASE_DIR) / "INDIA" / "consignee-app" / ".env"
    load_dotenv(dotenv_path=env_path)
    SENDER_EMAIL = os.getenv("SENDER_EMAIL")
    SENDER_PASSWORD = os.getenv("SENDER_PASSWORD")
    RECEIVER_EMAIL = os.getenv("RECEIVER_EMAIL")

BRANCH_PREFIX_MAP = {
    "CALICUT PETTY CASH IBB": "CLT NEW Petty Cash",
    "MALAPPURAM PETTY CASH": "MLPM Petty Cash",
    "KANHANGAD PETTY CASH": "KSD NEW Petty Cash",
    "CASH - OFFICE PETTY CASH": "NEW Petty Cash - EDATHALA",
    "ANAND PETTY CASH KANNUR": "KNR NEW Petty Cash",
    "KOLLAM PETTY CASH -IBB Saifudeen": "KLM NEW Petty Cash",
    "KOTTAYAM PETTY CASH": "KOTTAYAM Petty Cash -"
}

MONTHS_MAP = {
    'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
    'JUL': 7, 'JULY': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12
}

def clean_val(val, default=""):
    if pd.isna(val) or val is None or val is pd.NaT:
        return default
    s = str(val).strip()
    if s.lower() in ('nan', 'nat', 'none', 'null', '-'):
        return default
    return s

def parse_date(val_str):
    if not val_str:
        return None
    val_str = str(val_str).strip()
    val_str = val_str.replace('/', '-').replace('.', '-')
    
    # Check for text dates like "7April"
    m_text = re.match(r'(\d+)\s*([A-Za-z]+)', val_str)
    if m_text:
        day = int(m_text.group(1))
        month_str = m_text.group(2).upper()[:3]
        month = MONTHS_MAP.get(month_str, 8) # default to Aug if not matched
        year = datetime.now().year
        return datetime(year, month, day)

    formats = (
        '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %I:%M:%S %p', '%Y-%m-%d %H:%M', '%Y-%m-%d %I:%M %p', '%Y-%m-%d',
        '%d-%m-%Y %H:%M:%S', '%d-%m-%Y %I:%M:%S %p', '%d-%m-%Y %H:%M', '%d-%m-%Y %I:%M %p', '%d-%m-%Y',
        '%m-%d-%Y %H:%M:%S', '%m-%d-%Y %I:%M:%S %p', '%m-%d-%Y %H:%M', '%m-%d-%Y %I:%M %p', '%m-%d-%Y',
    )
    for fmt in formats:
        try:
            return datetime.strptime(val_str, fmt)
        except ValueError:
            continue
            
    try:
        return pd.to_datetime(val_str, dayfirst=True).to_pydatetime()
    except Exception:
        return None

def parse_date_range_from_title(title):
    pattern = r'(\d{1,2})\s*([A-Z]{3,4})\s*-\s*(\d{1,2})\s*([A-Z]{3,4})'
    match = re.search(pattern, title.upper())
    if match:
        start_day = int(match.group(1))
        start_month_str = match.group(2)
        end_day = int(match.group(3))
        end_month_str = match.group(4)
        
        start_month = MONTHS_MAP.get(start_month_str)
        if not start_month:
            for k, v in MONTHS_MAP.items():
                if k in start_month_str:
                    start_month = v
                    break
        
        end_month = MONTHS_MAP.get(end_month_str)
        if not end_month:
            for k, v in MONTHS_MAP.items():
                if k in end_month_str:
                    end_month = v
                    break
                    
        if start_month and end_month:
            return start_day, start_month, end_day, end_month
    return None

def date_falls_in_range(tx_date, start_day, start_month, end_day, end_month):
    year = tx_date.year
    start_date = datetime(year, start_month, start_day)
    if end_month < start_month:  # Crossover (Dec -> Jan)
        if tx_date.month >= start_month:
            end_date = datetime(year + 1, end_month, end_day)
        else:
            start_date = datetime(year - 1, start_month, start_day)
            end_date = datetime(year, end_month, end_day)
    else:
        end_date = datetime(year, end_month, end_day)
        
    tx_date_clean = datetime(tx_date.year, tx_date.month, tx_date.day)
    return start_date <= tx_date_clean <= end_date

def load_df(file_path):
    with open(file_path, "rb") as f:
        head = f.read(4)
    if head == b"PK\x03\x04" or head == b"\xd0\xcf\x11\xe0":
        return pd.read_excel(file_path)
    else:
        import csv
        import io
        cleaned_rows = []
        for enc in ["utf-8", "latin1", "utf-8-sig"]:
            try:
                with open(file_path, "r", encoding=enc) as f_csv:
                    reader = csv.reader(f_csv)
                    header = next(reader)
                    num_cols = len(header)
                    cleaned_rows.append(header)
                    
                    for row in reader:
                        if len(row) > num_cols:
                            last_col_val = ",".join(row[num_cols-1:])
                            cleaned_row = row[:num_cols-1] + [last_col_val]
                            cleaned_rows.append(cleaned_row)
                        else:
                            cleaned_row = row + [""] * (num_cols - len(row))
                            cleaned_rows.append(cleaned_row)
                
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerows(cleaned_rows)
                output.seek(0)
                return pd.read_csv(output)
            except Exception:
                continue
        return pd.read_csv(file_path)

def calculate_branch_stats(pc_rows):
    if not pc_rows or len(pc_rows) <= 1:
        return 0.0, 0.0, 0.0
        
    headers = [h.strip() for h in pc_rows[0]]
    payment_idx = headers.index("Payment") if "Payment" in headers else 6
    balance_idx = headers.index("Balance") if "Balance" in headers else 7
    
    closing_balance = 0.0
    total_payments = 0.0
    
    for row in reversed(pc_rows[1:]):
        if len(row) > balance_idx and row[balance_idx].strip():
            try:
                bal_clean = row[balance_idx].replace(',', '').replace('(', '-').replace(')', '').strip()
                closing_balance = float(bal_clean)
                break
            except ValueError:
                continue
                
    for row in pc_rows[1:]:
        if len(row) > payment_idx and row[payment_idx].strip():
            try:
                pay_val = float(row[payment_idx].replace(',', ''))
                total_payments += pay_val
            except ValueError:
                continue
                
    recommended_topup = abs(closing_balance) if closing_balance < 0 else 0.0
    return closing_balance, total_payments, recommended_topup

def main():
    print("Initializing Petty Cash Auditing & Reconciliation System...", flush=True)
    
    # 1. Authorize Google Sheets API
    if not os.path.exists(CREDENTIALS_PATH):
        print(f"Error: Google Service Account credentials not found at {CREDENTIALS_PATH}", flush=True)
        sys.exit(1)
        
    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=scopes)
    client = gspread.authorize(creds)
    print("Authorized Google API successfully.", flush=True)

    # 2. Get list of all spreadsheets
    print("Discovering shared spreadsheets on Google Drive...", flush=True)
    all_spreadsheets = []
    for attempt in range(5):
        try:
            all_spreadsheets = client.openall()
            break
        except Exception as api_err:
            if attempt == 4:
                raise api_err
            backoff_secs = (attempt + 1) * 5
            print(f"Google API Error: {api_err}. Retrying in {backoff_secs}s...", flush=True)
            time.sleep(backoff_secs)
    print(f"Found {len(all_spreadsheets)} accessible spreadsheets.", flush=True)

    # Locate Purchase Register
    pr_sheet = next((s for s in all_spreadsheets if s.title.strip().upper() == "PURCHASE REGISTER"), None)
    if not pr_sheet:
        print("Error: 'PURCHASE REGISTER' spreadsheet not shared or not found.", flush=True)
        sys.exit(1)
    print(f"Connected to Purchase Register (ID: {pr_sheet.id})", flush=True)

    # 3. Read Purchase Register 'MAIN' and 'PETTY CASH MASTER'
    try:
        main_ws = pr_sheet.worksheet("MAIN")
        master_ws = pr_sheet.worksheet("PETTY CASH MASTER")
    except Exception as e:
        print(f"Error accessing worksheets in Purchase Register: {e}", flush=True)
        sys.exit(1)

    mapping_data = master_ws.get_all_values()
    if len(mapping_data) > 1:
        branch_map = {}
        for r in mapping_data[1:]:
            if len(r) >= 2 and r[0].strip() and r[1].strip():
                branch_map[r[0].strip()] = r[1].strip()
    else:
        branch_map = BRANCH_PREFIX_MAP
    print("Branch Name Mapping:", branch_map, flush=True)

    pr_rows = main_ws.get_all_values()
    pr_data_rows = pr_rows[4:]
    
    party_col_idx = 4  
    net_pay_col_idx = 11 
    date_col_idx = 12 
    
    transfers = []
    for r_idx, row in enumerate(pr_data_rows):
        if len(row) <= max(party_col_idx, net_pay_col_idx, date_col_idx):
            continue
        party_name = row[party_col_idx].strip()
        net_pay_str = row[net_pay_col_idx].strip()
        date_str = row[date_col_idx].strip()
        
        matched_branch_pr = None
        for pr_name in branch_map.keys():
            if pr_name.upper() in party_name.upper():
                matched_branch_pr = pr_name
                break
                
        if matched_branch_pr:
            try:
                clean_pay = net_pay_str.replace(',', '')
                amount = float(clean_pay) if clean_pay else 0.0
            except ValueError:
                amount = 0.0
                
            tx_date = parse_date(date_str)
            if tx_date and amount > 0:
                transfers.append({
                    "row_number": r_idx + 5, 
                    "branch_pr_name": matched_branch_pr,
                    "branch_prefix": branch_map[matched_branch_pr],
                    "party_entered": party_name,
                    "amount": amount,
                    "date": tx_date,
                    "date_str": date_str
                })
                
    print(f"Found {len(transfers)} petty cash transfers in Purchase Register.", flush=True)

    # 4. Load ERP reports for Paid LRs and GDMs
    lr_file = os.path.join(DOWNLOAD_DIR, "lr_raw.xlsx")
    despatch_file = os.path.join(DOWNLOAD_DIR, "despatch_raw.xlsx")
    
    if not os.path.exists(lr_file) or not os.path.exists(despatch_file):
        print("Raw ERP downloads not found in ~/Downloads. Checking scratch fallbacks...", flush=True)
        fallback_dir = os.path.join(BASE_DIR, "scratch", "manual_run_july18_tables", "debug-artifacts")
        lr_file = os.path.join(fallback_dir, "lr_raw.xlsx")
        despatch_file = os.path.join(fallback_dir, "despatch_raw.xlsx")
        
    print(f"Using ERP LR report: {lr_file}", flush=True)
    print(f"Using ERP Despatch report: {despatch_file}", flush=True)

    try:
        df_lr = load_df(lr_file)
        df_lr.columns = [str(c).strip().upper() for c in df_lr.columns]
        
        df_desp = load_df(despatch_file)
        df_desp.columns = [str(c).strip().upper() for c in df_desp.columns]
        print(f"Loaded {len(df_lr)} actual LRs and {len(df_desp)} actual dispatches from ERP.", flush=True)
    except Exception as e:
        print(f"Error loading ERP files: {e}. Checking without ERP lookups.", flush=True)
        df_lr = pd.DataFrame()
        df_desp = pd.DataFrame()

    # Pre-index ERP data
    lr_db = {}
    if not df_lr.empty:
        lr_no_col = next((c for c in df_lr.columns if c in ['LR NO', 'LRNO', 'LR_NUMBER', 'LR_NO']), None)
        fright_col = next((c for c in df_lr.columns if c in ['TOTAL FRIGHT', 'TOTAL_FRIGHT', 'FRIGHT', 'LR AMOUNT']), None)
        status_col = next((c for c in df_lr.columns if c in ['LR STATUS', 'STATUS']), None)
        consignor_col = next((c for c in df_lr.columns if c in ['CONSIGNOR', 'CONSIGNOR_NAME']), None)
        consignee_col = next((c for c in df_lr.columns if c in ['CONSIGNEE', 'CONSIGNEE_NAME']), None)
        dest_col = next((c for c in df_lr.columns if c in ['DESTINATION', 'PLACE', 'AREA']), None)
        box_col = next((c for c in df_lr.columns if c in ['BOX QTY', 'BOXES', 'QUANTITY']), None)
        
        for _, r in df_lr.iterrows():
            lr_no = clean_val(r[lr_no_col]) if lr_no_col else ""
            if lr_no:
                try:
                    fright_val = float(clean_val(r[fright_col], "0").replace(',', ''))
                except ValueError:
                    fright_val = 0.0
                try:
                    box_val = int(float(clean_val(r[box_col], "0")))
                except ValueError:
                    box_val = 0
                lr_db[lr_no] = {
                    "total_fright": fright_val,
                    "status": clean_val(r[status_col]),
                    "consignor": clean_val(r[consignor_col]),
                    "consignee": clean_val(r[consignee_col]),
                    "destination": clean_val(r[dest_col]),
                    "box_qty": box_val
                }

    # Pre-index GDM to LR mapping
    gdm_lrs = {} 
    if not df_desp.empty:
        desp_no_col = next((c for c in df_desp.columns if c in ['DESPATCH NO', 'DISPATCH NO', 'DESPATCH_NO']), None)
        lr_no_col = next((c for c in df_desp.columns if c in ['LR NO', 'LRNO', 'LR_NUMBER', 'LR_NO']), None)
        
        for _, r in df_desp.iterrows():
            gdm = clean_val(r[desp_no_col]) if desp_no_col else ""
            lr = clean_val(r[lr_no_col]) if lr_no_col else ""
            if gdm and lr:
                if gdm not in gdm_lrs:
                    gdm_lrs[gdm] = []
                gdm_lrs[gdm].append(lr)
 
    # Load bill_clear exports
    bill_clear_db = {}
    import glob
    bc_files = glob.glob(os.path.join(DOWNLOAD_DIR, "bill_clear_*.xlsx"))
    print(f"Loading {len(bc_files)} bill clearance report files...", flush=True)
    for bf in bc_files:
        try:
            df_bc = load_df(bf)
            if df_bc.empty:
                continue
            df_bc.columns = [str(c).strip().upper() for c in df_bc.columns]
            
            lr_col = next((c for c in df_bc.columns if c in ['LR NUMBER', 'LR_NUMBER', 'LR NO', 'LRNO', 'LR']), None)
            fright_col = next((c for c in df_bc.columns if c in ['FREIGHT', 'TOTAL FRIGHT', 'FRIGHT AMOUNT']), None)
            topay_col = next((c for c in df_bc.columns if c in ['TO PAY', 'TOPAY', 'TO_PAY']), None)
            qty_col = next((c for c in df_bc.columns if c in ['QUANTITY', 'QTY', 'BOX COUNT', 'BOXES']), None)
            consignor_col = next((c for c in df_bc.columns if c in ['CONSIGNOR', 'CONSIGNOR_NAME']), None)
            consignee_col = next((c for c in df_bc.columns if c in ['CONSIGNEE', 'CONSIGNEE_NAME']), None)
            dest_col = next((c for c in df_bc.columns if c in ['DESTINATION', 'PLACE', 'AREA']), None)
            
            for _, r in df_bc.iterrows():
                lr_no = clean_val(r[lr_col]) if lr_col else ""
                if lr_no:
                    try:
                        fright_val = float(clean_val(r[fright_col], "0").replace('INR', '').replace(',', '').strip())
                    except ValueError:
                        fright_val = 0.0
                    try:
                        topay_val = float(clean_val(r[topay_col], "0").replace('INR', '').replace(',', '').strip())
                    except ValueError:
                        topay_val = 0.0
                    try:
                        qty_val = int(float(clean_val(r[qty_col], "0")))
                    except ValueError:
                        qty_val = 0
                        
                    bill_clear_db[lr_no] = {
                        "total_fright": fright_val,
                        "topay": topay_val,
                        "box_qty": qty_val,
                        "status": "PAID" if topay_val == 0.0 else "TO PAY",
                        "consignor": clean_val(r[consignor_col]) if consignor_col else "",
                        "consignee": clean_val(r[consignee_col]) if consignee_col else "",
                        "destination": clean_val(r[dest_col]) if dest_col else ""
                    }
        except Exception as bc_err:
            print(f"Error parsing bill clear file {bf}: {bc_err}", flush=True)
            
    print(f"Loaded {len(bill_clear_db)} cleared LRs from bill_clear files.", flush=True)

    gdm_scraped_db = {}
    gdm_json_path = os.path.join(DOWNLOAD_DIR, "gdm_details.json")
    if os.path.exists(gdm_json_path):
        try:
            import json
            with open(gdm_json_path, "r", encoding="utf-8") as json_f:
                gdm_scraped_db = json.load(json_f)
            print(f"Loaded {len(gdm_scraped_db)} scraped GDMs from gdm_details.json", flush=True)
        except Exception as json_err:
            print(f"Error loading gdm_details.json: {json_err}", flush=True)

    # 5. Core Reconciliation Engine
    discrepancies_funding = []
    
    # Store branch summary details: { branch_name: { closing_balance, total_payments, recommended_topup } }
    branch_topup_summary = {}
    sheet_data_cache = {}

    def get_cached_sheet_data(ss_obj):
        s_id = ss_obj.id
        if s_id in sheet_data_cache:
            return sheet_data_cache[s_id]
            
        print(f"Loading data for spreadsheet: '{ss_obj.title}' into cache...", flush=True)
        cache_entry = {
            "title": ss_obj.title,
            "Petty Cash": [],
            "PAID LR": [],
            "GDM": [],
            "Rate": [],
            "Despatch working": []
        }
        
        worksheets = ss_obj.worksheets()
        ws_titles = {w.title.strip().upper(): w for w in worksheets}
        
        for ws_name in ["Petty Cash", "PAID LR", "GDM", "Rate", "Despatch working"]:
            ws = None
            for t in [ws_name, ws_name.upper(), ws_name.lower(), ws_name.title()]:
                if t.strip().upper() in ws_titles:
                    ws = ws_titles[t.strip().upper()]
                    break
            if ws:
                try:
                    time.sleep(1.5)
                    cache_entry[ws_name] = ws.get_all_values()
                except Exception as e:
                    print(f"  [Warning] Worksheet '{ws.title}' failed to load: {e}", flush=True)
                    cache_entry[ws_name] = []
            else:
                cache_entry[ws_name] = []
                
        sheet_data_cache[s_id] = cache_entry
        return cache_entry

    # Reconcile HO CASH transfers
    for tx_idx, tx in enumerate(transfers):
        branch_pref = tx["branch_prefix"]
        tx_date = tx["date"]
        amount = tx["amount"]
        
        matched_sheet = None
        for s in all_spreadsheets:
            title = s.title.strip()
            if title.upper().startswith(branch_pref.upper()):
                date_range = parse_date_range_from_title(title)
                if date_range:
                    sd, sm, ed, em = date_range
                    if date_falls_in_range(tx_date, sd, sm, ed, em):
                        matched_sheet = s
                        break
                        
        if not matched_sheet:
            discrepancies_funding.append({
                "Date": tx["date_str"],
                "Branch": tx["branch_pr_name"],
                "Transfer Amount": amount,
                "Status": "MISSING SPREADSHEET",
                "Details": f"No Petty Cash spreadsheet found for date range containing {tx['date_str']}"
            })
            continue
 
        sheet_cache = get_cached_sheet_data(matched_sheet)
        pc_rows = sheet_cache["Petty Cash"]

        matched_booking = False
        if pc_rows:
            pc_headers = [h.strip() for h in pc_rows[0]]
            date_idx = pc_headers.index("Date") if "Date" in pc_headers else 1
            type_idx = pc_headers.index("Payment / Receipt") if "Payment / Receipt" in pc_headers else 3
            receipt_idx = pc_headers.index("Receipt") if "Receipt" in pc_headers else 5
            
            for pc_row in pc_rows[1:]:
                if len(pc_row) <= max(date_idx, type_idx, receipt_idx):
                    continue
                row_date_str = pc_row[date_idx].strip()
                row_type = pc_row[type_idx].strip()
                row_receipt_str = pc_row[receipt_idx].strip()
                
                if "HO" in row_type.upper() or "CASH" in row_type.upper():
                    r_date = parse_date(row_date_str)
                    if r_date and abs((r_date - tx_date).days) <= 2:
                        try:
                            booked_amt = float(row_receipt_str.replace(',', '')) if row_receipt_str else 0.0
                        except ValueError:
                            booked_amt = 0.0
                            
                        if abs(booked_amt - amount) < 1.0: 
                            matched_booking = True
                            break
                            
        if not matched_booking:
            discrepancies_funding.append({
                "Date": tx["date_str"],
                "Branch": tx["branch_pr_name"],
                "Transfer Amount": amount,
                "Status": "UNBOOKED FUND",
                "Details": f"Fund transfer of {amount} on {tx['date_str']} is not recorded as 'HO CASH' in sheet '{matched_sheet.title}'"
            })

    # Calculate Top-up and balance stats for each branch
    for pr_name, prefix in branch_map.items():
        branch_sheets = [s for s in all_spreadsheets if s.title.strip().upper().startswith(prefix.upper())]
        if not branch_sheets:
            branch_topup_summary[pr_name] = {
                "closing_balance": 0.0,
                "total_payments": 0.0,
                "recommended_topup": 0.0,
                "status": "Spreadsheet Missing"
            }
            continue
            
        now_ist = datetime.now(timezone(timedelta(hours=5, minutes=30)))
        active_sheet = None
        for s in branch_sheets:
            dr = parse_date_range_from_title(s.title)
            if dr and date_falls_in_range(now_ist, dr[0], dr[1], dr[2], dr[3]):
                active_sheet = s
                break
        if not active_sheet:
            active_sheet = branch_sheets[0] 
            
        sheet_cache = get_cached_sheet_data(active_sheet)
        pc_rows = sheet_cache["Petty Cash"]
        c_bal, t_pay, r_top = calculate_branch_stats(pc_rows)
        
        branch_topup_summary[pr_name] = {
            "closing_balance": c_bal,
            "total_payments": t_pay,
            "recommended_topup": r_top,
            "status": "Success",
            "sheet_title": active_sheet.title
        }

    # Dynamic Reconstruction Fallback for Local/Offline Testing
    # If bill_clear_db is empty, we reconstruct expected LR data using raw reports and PAID LR registers
    if len(bill_clear_db) == 0:
        print("Adhoc Bill Clearance Database is empty. Reconstructing it dynamically for audit...", flush=True)
        # 1. Add Paid LRs from PAID LR worksheets
        for s_id, cache in sheet_data_cache.items():
            paid_lr_rows = cache.get("PAID LR", [])
            if len(paid_lr_rows) > 1:
                # Find date, lr, and amount indices
                def get_idx(lst, names, default):
                    for n in names:
                        for idx, item in enumerate(lst):
                            if item == n:
                                return idx
                    for n in names:
                        for idx, item in enumerate(lst):
                            if n in item:
                                return idx
                    return default

                headers = [h.strip().upper() for h in paid_lr_rows[0]]
                lr_idx = get_idx(headers, ["LR NO", "LR NUMBER", "LR_NO", "LRNO"], 2)
                amt_idx = get_idx(headers, ["AMOUNT", "TOTAL", "FRIGHT", "LR AMOUNT"], 10)
                
                for r in paid_lr_rows[1:]:
                    if len(r) > max(lr_idx, amt_idx):
                        lr_no = r[lr_idx].strip()
                        if lr_no and not lr_no.upper().startswith("TOTAL"):
                            try:
                                amt_val = float(r[amt_idx].replace(',', '').strip())
                            except ValueError:
                                amt_val = 0.0
                                
                            # Lookup standard details in lr_db
                            consignor = ""
                            consignee = ""
                            box_qty = 0
                            if lr_no in lr_db:
                                consignor = lr_db[lr_no]["consignor"]
                                consignee = lr_db[lr_no]["consignee"]
                                box_qty = lr_db[lr_no]["box_qty"]
                                if amt_val == 0.0:
                                    amt_val = lr_db[lr_no]["total_fright"]
                                    
                            bill_clear_db[lr_no] = {
                                "total_fright": amt_val,
                                "topay": 0.0,
                                "box_qty": box_qty,
                                "status": "PAID",
                                "consignor": consignor,
                                "consignee": consignee
                            }
        # 2. Add Topay LRs from Despatch Working sheet
        for s_id, cache in sheet_data_cache.items():
            dw_rows = cache.get("Despatch working", [])
            dw_header_idx = None
            for idx, r in enumerate(dw_rows):
                r_upper = [c.upper() for c in r]
                if any("LR NO" in c or "DESPATCH" in c or "CONSIGNOR" in c for c in r_upper):
                    dw_header_idx = idx
                    break
            if dw_header_idx is not None:
                dw_headers = [h.strip().upper() for h in dw_rows[dw_header_idx]]
                dw_lr_idx = get_idx(dw_headers, ["LR NO", "LR NUMBER", "LR_NO", "LRNO", "LR"], 4)
                dw_topay_idx = get_idx(dw_headers, ["TOPAY", "TO PAY"], 12)

                for r in dw_rows[dw_header_idx + 1:]:
                    if len(r) > max(dw_lr_idx, dw_topay_idx):
                        lr_no = r[dw_lr_idx].strip()
                        if lr_no and not lr_no.upper().startswith("TOTAL"):
                            try:
                                topay_val = float(r[dw_topay_idx].replace(',', '').strip())
                            except ValueError:
                                topay_val = 0.0
                            if topay_val > 0.0:
                                consignor = ""
                                consignee = ""
                                box_qty = 0
                                if lr_no in lr_db:
                                    consignor = lr_db[lr_no]["consignor"]
                                    consignee = lr_db[lr_no]["consignee"]
                                    box_qty = lr_db[lr_no]["box_qty"]
                                    
                                bill_clear_db[lr_no] = {
                                    "total_fright": topay_val,
                                    "topay": topay_val,
                                    "box_qty": box_qty,
                                    "status": "TO PAY",
                                    "consignor": consignor,
                                    "consignee": consignee
                                }
        print(f"Dynamically reconstructed {len(bill_clear_db)} LRs in bill clearance database.", flush=True)

    # Lists to store aggregated results across all branch spreadsheets
    all_balance_mismatches = []
    all_missing_gdms = []
    all_unloading_variances = []
    all_other_expenses = []

    # Reconcile PAID LRs & GDMs using cached data
    for s_id, cache in sheet_data_cache.items():
        title = cache["title"]
        print(f"\nAuditing details for branch sheet: {title}...", flush=True)
        
        pc_rows = cache.get("Petty Cash", [])
        dw_rows = cache.get("Despatch working", [])
        gdm_rows = cache.get("GDM", [])
        rate_rows = cache.get("Rate", [])
        paid_lr_rows = cache.get("PAID LR", [])
        
        branch_name = title
        for b_pr, prefix in branch_map.items():
            if title.upper().startswith(prefix.upper()):
                branch_name = b_pr
                break
                
        # Parse Petty Cash Sheet
        pc_entries = []
        if len(pc_rows) > 0:
            pc_headers = [h.strip() for h in pc_rows[0]]
            date_idx = pc_headers.index("Date") if "Date" in pc_headers else 1
            gdm_idx = pc_headers.index("GDM No") if "GDM No" in pc_headers else 2
            type_idx = pc_headers.index("Payment / Receipt") if "Payment / Receipt" in pc_headers else 3
            details_idx = pc_headers.index("Details") if "Details" in pc_headers else 4
            receipt_idx = pc_headers.index("Receipt") if "Receipt" in pc_headers else 5
            payment_idx = pc_headers.index("Payment") if "Payment" in pc_headers else 6
            balance_idx = pc_headers.index("Balance") if "Balance" in pc_headers else 7
            remark_idx = pc_headers.index("Remark") if "Remark" in pc_headers else 8
            
            for r_idx, row in enumerate(pc_rows[1:]):
                if len(row) <= max(date_idx, gdm_idx, type_idx, payment_idx):
                    continue
                date_str = row[date_idx].strip()
                gdm_no_str = row[gdm_idx].strip()
                type_str = row[type_idx].strip()
                details_str = row[details_idx].strip() if len(row) > details_idx else ""
                receipt_str = row[receipt_idx].strip() if len(row) > receipt_idx else ""
                payment_str = row[payment_idx].strip() if len(row) > payment_idx else ""
                balance_str = row[balance_idx].strip() if len(row) > balance_idx else ""
                remark_str = row[remark_idx].strip() if len(row) > remark_idx else ""
                
                if not date_str or "OPENING" in details_str.upper() or "OPENING" in type_str.upper():
                    continue
                    
                try:
                    receipt_val = float(receipt_str.replace(',', '')) if receipt_str else 0.0
                except ValueError:
                    receipt_val = 0.0
                try:
                    payment_val = float(payment_str.replace(',', '')) if payment_str else 0.0
                except ValueError:
                    payment_val = 0.0
                try:
                    balance_val = float(balance_str.replace(',', '')) if balance_str else 0.0
                except ValueError:
                    balance_val = 0.0
                    
                pc_entries.append({
                    "row_no": r_idx + 2,
                    "date_str": date_str,
                    "date": parse_date(date_str),
                    "gdm_no": gdm_no_str,
                    "type": type_str,
                    "details": details_str,
                    "receipt": receipt_val,
                    "payment": payment_val,
                    "balance": balance_val,
                    "remark": remark_str
                })
        
        # Parse Despatch Working Sheet
        dw_entries = []
        dw_header_idx = None
        for idx, r in enumerate(dw_rows):
            r_upper = [c.upper() for c in r]
            if any("LR NO" in c or "DESPATCH" in c or "CONSIGNOR" in c for c in r_upper):
                dw_header_idx = idx
                break
                
        if dw_header_idx is not None:
            dw_headers = [h.strip() for h in dw_rows[dw_header_idx]]
            
            def find_col_idx(names, default):
                # First pass: exact matches
                for name in names:
                    for i, h in enumerate(dw_headers):
                        if h and name.upper() == h.strip().upper():
                            return i
                # Second pass: substring matches
                for name in names:
                    for i, h in enumerate(dw_headers):
                        if h and name.upper() in h.strip().upper():
                            return i
                return default

                
            dw_date_idx = find_col_idx(["Despatch Date"], 0)
            dw_consignor_idx = find_col_idx(["Consignor"], 1)
            dw_consignee_idx = find_col_idx(["Consignee"], 2)
            dw_destination_idx = find_col_idx(["Destination"], 3)
            dw_lr_idx = find_col_idx(["Lr no", "LR Number"], 4)
            dw_invoice_idx = find_col_idx(["Invoice No"], 5)
            dw_weight_idx = find_col_idx(["Weight"], 6)
            dw_qty_idx = find_col_idx(["Box Qty", "Quantity"], 7)
            dw_fright_idx = find_col_idx(["Total Fright", "Freight"], 8)
            dw_gdm_idx = find_col_idx(["DESPATCH", "GDM", "Despatch No"], 9)
            dw_topay_idx = find_col_idx(["Topay", "To Pay"], 12)
            
            # Unloading columns (handles KNR duplicates)
            ul_indices = [i for i, h in enumerate(dw_headers) if h and 'UNLOADING CHARGE IN MASTER' in h.upper()]
            dw_unloading_idx = ul_indices[0] if len(ul_indices) > 0 else 13
            dw_claimed_ul_idx = 14
            claimed_ul_indices = [i for i, h in enumerate(dw_headers) if h and 'CLAIMED UL' in h.upper()]
            if claimed_ul_indices:
                dw_claimed_ul_idx = claimed_ul_indices[0]
            elif len(ul_indices) > 1:
                dw_claimed_ul_idx = ul_indices[1]
                
            dw_bata_idx = find_col_idx(["Route Bata", "BATA"], 15)
            dw_toll_idx = find_col_idx(["Toll / Parking", "Toll/Parking"], 16)
            dw_bonus_idx = find_col_idx(["Bonus"], 17)
            dw_total_idx = find_col_idx(["Total"], 18)
            dw_remark_idx = find_col_idx(["REMARK"], 25)
            
            for r_idx, row in enumerate(dw_rows[dw_header_idx + 1:]):
                max_mapped = max(
                    dw_date_idx, dw_consignor_idx, dw_consignee_idx, dw_destination_idx,
                    dw_lr_idx, dw_invoice_idx, dw_qty_idx, dw_fright_idx, dw_gdm_idx,
                    dw_topay_idx, dw_unloading_idx, dw_claimed_ul_idx, dw_bata_idx,
                    dw_toll_idx, dw_bonus_idx, dw_total_idx
                )
                if len(row) <= max_mapped:
                    row = row + [""] * (max_mapped - len(row) + 1)
                    
                gdm_val = row[dw_gdm_idx].strip()
                lr_val = row[dw_lr_idx].strip()
                if not lr_val or lr_val.upper() in ("LR NO", "TOTAL", "SUB TOTAL", "CANCELLED", ""):
                    continue
                    
                try:
                    qty_val = int(float(row[dw_qty_idx].replace(',', '').strip())) if row[dw_qty_idx].strip() else 0
                except ValueError:
                    qty_val = 0
                try:
                    fright_val = float(row[dw_fright_idx].replace(',', '').strip()) if row[dw_fright_idx].strip() else 0.0
                except ValueError:
                    fright_val = 0.0
                try:
                    topay_val = float(row[dw_topay_idx].replace(',', '').strip()) if row[dw_topay_idx].strip() else 0.0
                except ValueError:
                    topay_val = 0.0
                try:
                    unloading_val = float(row[dw_unloading_idx].replace(',', '').strip()) if row[dw_unloading_idx].strip() else 0.0
                except ValueError:
                    unloading_val = 0.0
                try:
                    claimed_ul_val = float(row[dw_claimed_ul_idx].replace(',', '').strip()) if row[dw_claimed_ul_idx].strip() else 0.0
                except ValueError:
                    claimed_ul_val = 0.0
                try:
                    bata_val = float(row[dw_bata_idx].replace(',', '').strip()) if row[dw_bata_idx].strip() else 0.0
                except ValueError:
                    bata_val = 0.0
                try:
                    toll_val = float(row[dw_toll_idx].replace(',', '').strip()) if row[dw_toll_idx].strip() else 0.0
                except ValueError:
                    toll_val = 0.0
                try:
                    bonus_val = float(row[dw_bonus_idx].replace(',', '').strip()) if row[dw_bonus_idx].strip() else 0.0
                except ValueError:
                    bonus_val = 0.0
                    
                dw_entries.append({
                    "row_no": r_idx + dw_header_idx + 2,
                    "date_str": row[dw_date_idx].strip(),
                    "consignor": row[dw_consignor_idx].strip(),
                    "consignee": row[dw_consignee_idx].strip(),
                    "destination": row[dw_destination_idx].strip(),
                    "lr_no": lr_val,
                    "invoice_no": row[dw_invoice_idx].strip(),
                    "box_qty": qty_val,
                    "fright": fright_val,
                    "gdm_no": gdm_val,
                    "topay": topay_val,
                    "unloading_master": unloading_val,
                    "claimed_ul": claimed_ul_val,
                    "bata": bata_val,
                    "toll": toll_val,
                    "bonus": bonus_val,
                    "remark": row[dw_remark_idx].strip() if len(row) > dw_remark_idx else ""
                })

        # Parse GDM details from the GDM worksheet
        gdm_meta = {}
        if len(gdm_rows) > 4:
            gdm_headers = [h.strip() for h in gdm_rows[3]]
            gdm_no_idx = gdm_headers.index("DESPATCH") if "DESPATCH" in gdm_headers else 0
            driver_idx = gdm_headers.index("Driver Name") if "Driver Name" in gdm_headers else 2
            gdm_adv_idx = gdm_headers.index("GDM ADVANCE") if "GDM ADVANCE" in gdm_headers else 3
            
            for r in gdm_rows[4:]:
                if len(r) > max(gdm_no_idx, driver_idx):
                    g_no = r[gdm_no_idx].strip()
                    d_name = r[driver_idx].strip()
                    try:
                        g_adv = float(r[gdm_adv_idx].replace(',', '').strip()) if len(r) > gdm_adv_idx and r[gdm_adv_idx].strip() else 0.0
                    except ValueError:
                        g_adv = 0.0
                    if g_no and g_no.upper() not in ("DESPATCH", "TOTAL", "SUB TOTAL", ""):
                        gdm_meta[g_no] = {
                            "driver_name": d_name,
                            "gdm_advance": g_adv
                        }

        # -------------------------------------------------------------
        # AUDIT STEPS
        # -------------------------------------------------------------
        
        # Collect unique GDM numbers across sheets
        gdms_in_branch = set()
        for e in pc_entries:
            if e["gdm_no"] and e["gdm_no"].isdigit():
                gdms_in_branch.add(e["gdm_no"])
        for e in dw_entries:
            if e["gdm_no"] and e["gdm_no"].isdigit():
                gdms_in_branch.add(e["gdm_no"])
        for g_no in gdm_meta.keys():
            if g_no.isdigit():
                gdms_in_branch.add(g_no)
                
        # 1. Audit GDM Balances (Mismatch Report & Missing GDM Report)
        for gdm in sorted(list(gdms_in_branch)):
            pc_gdm_entries = [e for e in pc_entries if e["gdm_no"] == gdm]
            dw_gdm_entries = [e for e in dw_entries if e["gdm_no"] == gdm]
            
            # Expected values from Despatch Working sheet
            expected_topay = sum(e["topay"] for e in dw_gdm_entries)
            expected_unloading = sum(e["unloading_master"] for e in dw_gdm_entries)
            expected_bata = sum(e["bata"] for e in dw_gdm_entries)
            expected_toll = sum(e["toll"] for e in dw_gdm_entries)
            
            # GDM Advance: Initial GDM ADVANCE entered in Petty Cash
            route_advance = sum(e["payment"] for e in pc_gdm_entries if "GDM ADVANCE" in e["type"].upper())
            # If not in Petty Cash, fallback to GDM worksheet
            if route_advance == 0.0 and gdm in gdm_meta:
                route_advance = gdm_meta[gdm]["gdm_advance"]
                
            # Expected Balance = (Topay Amount + Route Advance) - (Unloading Charge in Master + Toll/Parking + Route Bata)
            expected_balance = (expected_topay + route_advance) - (expected_unloading + expected_toll + expected_bata)
            
            # Actual Balance in Petty Cash = GDM RECEIPT - GDM ADDITIONAL ADVANCE
            actual_receipts = sum(e["receipt"] for e in pc_gdm_entries if "GDM RECEIPT" in e["type"].upper())
            actual_add_advances = sum(e["payment"] for e in pc_gdm_entries if "GDM ADDITIONAL ADVANCE" in e["type"].upper())
            actual_balance = actual_receipts - actual_add_advances
            
            # Verify if GDM is missing from Petty Cash sheet
            if len(pc_gdm_entries) == 0:
                if expected_topay > 0.0 or route_advance > 0.0:
                    driver_name = gdm_meta.get(gdm, {}).get("driver_name", "Unknown Driver")
                    all_missing_gdms.append({
                        "Sheet": title,
                        "GDM": gdm,
                        "Driver": driver_name,
                        "Topay": expected_topay,
                        "Advance": route_advance,
                        "Remarks": "GDM with active Topay or Advance has no entries in Petty Cash worksheet."
                    })
            else:
                # GDM Balance Mismatch calculation
                variance = actual_balance - expected_balance
                remarks = []
                
                # Check Topay recording correctness per LR in Despatch Working vs Bill Clearance
                for lr_entry in dw_gdm_entries:
                    lr_no = lr_entry["lr_no"]
                    if lr_no in bill_clear_db:
                        bc_topay = bill_clear_db[lr_no]["topay"]
                        if abs(lr_entry["topay"] - bc_topay) > 1.0:
                            remarks.append(f"LR {lr_no}: Topay recorded ({lr_entry['topay']}) differs from Bill Clearance ({bc_topay})")
                
                # Check for Paid LR Missing entries
                # Map all corresponding LRs from Despatch Working and check if Paid LRs are recorded in Petty Cash
                for lr_entry in dw_gdm_entries:
                    lr_no = lr_entry["lr_no"]
                    if lr_no in bill_clear_db:
                        if bill_clear_db[lr_no]["status"] == "PAID":
                            bc_paid_amt = bill_clear_db[lr_no]["total_fright"]
                            if bc_paid_amt > 0.0:
                                # Look for direct receipt entry in Petty Cash for this specific LR number
                                pc_lr_receipts = [e for e in pc_entries if lr_no.upper() in e["details"].upper() or lr_no.upper() in e["remark"].upper() or lr_no.upper() in e["gdm_no"].upper()]
                                if len(pc_lr_receipts) == 0:
                                    # Not found in Petty Cash worksheet!
                                    remarks.append(f"LR {lr_no}: Paid amount ({bc_paid_amt}) missing from Petty Cash sheet")
                                    # Also report as an individual row in Balance Mismatch Report
                                    all_balance_mismatches.append({
                                        "Sheet": title,
                                        "GDM": f"Paid LR {lr_no}",
                                        "Expected": bc_paid_amt,
                                        "Actual": 0.0,
                                        "Variance": -bc_paid_amt,
                                        "Remarks": f"LR {lr_no} marked as Paid in Bill Clearance but has no direct entry in Petty Cash sheet."
                                    })
                                else:
                                    pc_lr_paid = sum(e["receipt"] for e in pc_lr_receipts)
                                    if abs(pc_lr_paid - bc_paid_amt) > 1.0:
                                        remarks.append(f"LR {lr_no}: Paid amount in Petty Cash ({pc_lr_paid}) differs from Bill Clearance ({bc_paid_amt})")
                                        all_balance_mismatches.append({
                                            "Sheet": title,
                                            "GDM": f"Paid LR {lr_no}",
                                            "Expected": bc_paid_amt,
                                            "Actual": pc_lr_paid,
                                            "Variance": pc_lr_paid - bc_paid_amt,
                                            "Remarks": f"LR {lr_no} receipt amount mismatch: expected {bc_paid_amt}, entered {pc_lr_paid}"
                                        })
                
                # Verify GDM balance mismatch
                if abs(variance) > 2.0 or remarks:
                    desc_remarks = "Matched" if abs(variance) <= 2.0 else f"Balance mismatch of {variance:,.2f}"
                    if remarks:
                        desc_remarks += ". " + "; ".join(remarks)
                    all_balance_mismatches.append({
                        "Sheet": title,
                        "GDM": gdm,
                        "Expected": expected_balance,
                        "Actual": actual_balance,
                        "Variance": variance,
                        "Remarks": desc_remarks
                    })

        # 2. Audit Unloading Rate Variance (Unloading Rate Variance Report)
        # For each LR in Despatch working, match against Master Rate worksheet
        if dw_header_idx is not None:
            for lr_entry in dw_entries:
                lr_no = lr_entry["lr_no"]
                consignor = lr_entry["consignor"]
                consignee = lr_entry["consignee"]
                box_qty = lr_entry["box_qty"]
                claimed_ul = lr_entry["claimed_ul"]
                
                if not lr_no or box_qty == 0:
                    continue
                    
                # Clean Names for flexible matching
                c_consignor = consignor.strip().upper()
                c_consignee = consignee.strip().upper()
                
                # Determine Box Type/Description
                box_desc = "Standard Box"
                # Look up box type description in PAID LR or LR worksheets
                for ws_rows in [paid_lr_rows, cache.get("LR", [])]:
                    if len(ws_rows) > 1:
                        headers_ws = [h.strip().upper() for h in ws_rows[0]]
                        lr_ws_idx = next((i for i, h in enumerate(headers_ws) if "LR NO" in h or "LR" in h), 2)
                        boxes_ws_idx = next((i for i, h in enumerate(headers_ws) if "BOXES" in h or "BOX COUNT" in h), 9)
                        for r_ws in ws_rows[1:]:
                            if len(r_ws) > max(lr_ws_idx, boxes_ws_idx) and r_ws[lr_ws_idx].strip().upper() == lr_no.upper():
                                box_desc = r_ws[boxes_ws_idx].strip()
                                break
                        if box_desc != "Standard Box":
                            break
                            
                # Fallback to ERP lr_db description
                if box_desc == "Standard Box" and lr_no in lr_db:
                    box_desc = lr_db[lr_no].get("status", "Standard Box")
                    
                # Match standard rate in Rate worksheet
                matched_rate = None
                rate_remarks = ""
                if len(rate_rows) > 0:
                    # Rate Worksheet structure: [Consignor, Consignee, Rate, Sack, Drum]
                    # Consignor is in column A (index 0), Consignee in column B (index 1)
                    for rate_row in rate_rows[1:]:
                        if len(rate_row) >= 3:
                            r_consignor = rate_row[0].strip().upper()
                            r_consignee = rate_row[1].strip().upper()
                            
                            # Flexible mapping match
                            if (r_consignor in c_consignor or c_consignor in r_consignor) and (r_consignee in c_consignee or c_consignee in r_consignee):
                                try:
                                    std_rate = float(rate_row[2].strip()) if rate_row[2].strip() else 0.0
                                except ValueError:
                                    std_rate = 0.0
                                try:
                                    sack_rate = float(rate_row[3].strip()) if len(rate_row) > 3 and rate_row[3].strip() else 0.0
                                except ValueError:
                                    sack_rate = 0.0
                                try:
                                    drum_rate = float(rate_row[4].strip()) if len(rate_row) > 4 and rate_row[4].strip() else 0.0
                                except ValueError:
                                    drum_rate = 0.0
                                    
                                if "SACK" in box_desc.upper():
                                    matched_rate = sack_rate if sack_rate > 0.0 else std_rate
                                    rate_remarks = f"Sack rate applied ({matched_rate})"
                                elif "DRUM" in box_desc.upper():
                                    matched_rate = drum_rate if drum_rate > 0.0 else std_rate
                                    rate_remarks = f"Drum rate applied ({matched_rate})"
                                else:
                                    matched_rate = std_rate
                                    rate_remarks = f"Standard rate applied ({matched_rate})"
                                break
                                
                if matched_rate is None:
                    # Default rate fallback
                    matched_rate = 0.0
                    rate_remarks = "No rate row found in Master Rate Sheet"
                    
                # Despatch Working rate paid per box
                despatch_rate = round(claimed_ul / box_qty, 2)
                rate_diff = despatch_rate - matched_rate
                
                # If there's a discrepancy, report it
                if abs(rate_diff) > 0.01:
                    all_unloading_variances.append({
                        "Sheet": title,
                        "Consignor": consignor,
                        "Consignee": consignee,
                        "BoxType": box_desc,
                        "MasterRate": matched_rate,
                        "DespatchRate": despatch_rate,
                        "Difference": rate_diff,
                        "Remarks": f"LR {lr_no}: Total claimed {claimed_ul} for {box_qty} boxes. {rate_remarks}."
                    })

        # 3. Daily Other Expenses Breakdown
        # Collect daily payments that are NOT GDM-related and NOT HO CASH receipts
        for pc_e in pc_entries:
            if pc_e["payment"] > 0:
                type_upper = pc_e["type"].upper()
                details_upper = pc_e["details"].upper()
                
                is_gdm = "GDM" in type_upper or "GDM" in details_upper or pc_e["gdm_no"] != ""
                is_ho = "HO" in type_upper or "CASH" in type_upper or "FUND" in details_upper
                
                if not is_gdm and not is_ho:
                    all_other_expenses.append({
                        "Sheet": title,
                        "Date": pc_e["date_str"],
                        "Category": pc_e["type"] if pc_e["type"] else "Other Payment",
                        "Details": pc_e["details"],
                        "Amount": pc_e["payment"],
                        "Remark": pc_e["remark"]
                    })

    # 6. Generate Excel Audit Report with premium styles
    report_file_name = f"Petty_Cash_Audit_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    report_path = os.path.join(BASE_DIR, report_file_name)
    
    print(f"\nGenerating upgraded Excel audit report at {report_path}...", flush=True)
    wb_out = openpyxl.Workbook()
    
    # Stylings
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=14, bold=True, color="1F497D")
    section_font = Font(name="Calibri", size=12, bold=True)
    
    mismatch_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") 
    missing_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") 
    
    border_side = Side(style='thin', color='D9D9D9')
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # 6a. Overview Dashboard Sheet
    ws_dash = wb_out.active
    ws_dash.title = "Overview Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True
    
    ws_dash.append([])
    ws_dash.append(["PETTY CASH RECONCILIATION & AUDIT SYSTEM - DAILY SUMMARY"])
    ws_dash.cell(2, 1).font = title_font
    ws_dash.append([f"Report Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"])
    ws_dash.append([])
    
    # Add Branch Float Summary
    ws_dash.append(["Branch Petty Cash Float & Top-up Summary"])
    ws_dash.cell(ws_dash.max_row, 1).font = section_font
    
    float_headers = ["Branch Name", "Active sheet", "Closing Balance", "Total Payments (Spent)", "Recommended Top-up"]
    ws_dash.append(float_headers)
    curr_row = ws_dash.max_row
    for col in range(1, len(float_headers) + 1):
        ws_dash.cell(curr_row, col).fill = header_fill
        ws_dash.cell(curr_row, col).font = header_font
        
    for b_name, stats in branch_topup_summary.items():
        active_title = stats.get("sheet_title", stats.get("status", ""))
        ws_dash.append([
            b_name,
            active_title,
            stats["closing_balance"],
            stats["total_payments"],
            stats["recommended_topup"]
        ])
        curr_r = ws_dash.max_row
        ws_dash.cell(curr_r, 3).number_format = '#,##0.00'
        ws_dash.cell(curr_r, 4).number_format = '#,##0.00'
        ws_dash.cell(curr_r, 5).number_format = '#,##0.00'
        if stats["recommended_topup"] > 0:
            ws_dash.cell(curr_r, 5).font = Font(name="Calibri", size=11, bold=True, color="C00000")
            
    ws_dash.append([])
    
    # Add Discrepancy counts
    ws_dash.append(["Summary of Audit Exceptions Found"])
    ws_dash.cell(ws_dash.max_row, 1).font = section_font
    
    summary_headers = ["Audit Checklist Category", "Exceptions Count", "Review Priority"]
    ws_dash.append(summary_headers)
    curr_row = ws_dash.max_row
    for col in range(1, 4):
        ws_dash.cell(curr_row, col).fill = header_fill
        ws_dash.cell(curr_row, col).font = header_font
        
    audit_stats = [
        ("Petty Cash Balance Mismatch Issues (Section 1)", len(all_balance_mismatches), "HIGH" if len(all_balance_mismatches) > 0 else "NORMAL"),
        ("Missing / Unrecorded GDMs (Section 2)", len(all_missing_gdms), "MEDIUM" if len(all_missing_gdms) > 0 else "NORMAL"),
        ("Unloading Rate Variances (Section 3)", len(all_unloading_variances), "LOW" if len(all_unloading_variances) > 0 else "NORMAL"),
        ("Other Payments / Expenses Logged (Section 4)", len(all_other_expenses), "INFO")
    ]
    for row_val in audit_stats:
        ws_dash.append(row_val)
        curr_r = ws_dash.max_row
        ws_dash.cell(curr_r, 1).font = data_font
        ws_dash.cell(curr_r, 2).font = data_font
        ws_dash.cell(curr_r, 3).font = Font(name="Calibri", size=11, bold=True)
        if row_val[1] > 0 and row_val[2] in ("HIGH", "MEDIUM"):
            ws_dash.cell(curr_r, 3).font = Font(name="Calibri", size=11, bold=True, color="FF0000")
            
    for col in ws_dash.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_dash.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 6b. Section 1 Sheet: Petty Cash Balance Mismatch
    ws_s1 = wb_out.create_sheet("Balance Mismatch Report")
    ws_s1.views.sheetView[0].showGridLines = True
    s1_headers = ["Branch Sheet", "GDM / LR No", "Expected Balance", "Actual Balance in PC", "Variance", "Remarks"]
    ws_s1.append(s1_headers)
    for col in range(1, len(s1_headers) + 1):
        ws_s1.cell(1, col).fill = header_fill
        ws_s1.cell(1, col).font = header_font
        
    for issue in all_balance_mismatches:
        ws_s1.append([issue["Sheet"], issue["GDM"], issue["Expected"], issue["Actual"], issue["Variance"], issue["Remarks"]])
        curr_r = ws_s1.max_row
        ws_s1.cell(curr_r, 3).number_format = '#,##0.00'
        ws_s1.cell(curr_r, 4).number_format = '#,##0.00'
        ws_s1.cell(curr_r, 5).number_format = '#,##0.00'
        
        # Highlight in orange
        fill = mismatch_fill
        for c in range(1, len(s1_headers) + 1):
            ws_s1.cell(curr_r, c).fill = fill
            ws_s1.cell(curr_r, c).border = cell_border
            
    for col in ws_s1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_s1.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 6c. Section 2 Sheet: Missing GDM Report
    ws_s2 = wb_out.create_sheet("Missing GDM Report")
    ws_s2.views.sheetView[0].showGridLines = True
    s2_headers = ["Branch Sheet", "GDM Number", "Driver Name", "Expected Topay", "Route Advance", "Remarks"]
    ws_s2.append(s2_headers)
    for col in range(1, len(s2_headers) + 1):
        ws_s2.cell(1, col).fill = header_fill
        ws_s2.cell(1, col).font = header_font
        
    for issue in all_missing_gdms:
        ws_s2.append([issue["Sheet"], issue["GDM"], issue["Driver"], issue["Topay"], issue["Advance"], issue["Remarks"]])
        curr_r = ws_s2.max_row
        ws_s2.cell(curr_r, 4).number_format = '#,##0.00'
        ws_s2.cell(curr_r, 5).number_format = '#,##0.00'
        
        # Highlight in yellow
        fill = missing_fill
        for c in range(1, len(s2_headers) + 1):
            ws_s2.cell(curr_r, c).fill = fill
            ws_s2.cell(curr_r, c).border = cell_border
            
    for col in ws_s2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_s2.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 6d. Section 3 Sheet: Unloading Rate Variance Report
    ws_s3 = wb_out.create_sheet("Unloading Rate Variance")
    ws_s3.views.sheetView[0].showGridLines = True
    s3_headers = ["Branch Sheet", "Consignor", "Consignee", "Box Type", "Master Rate", "Despatch Working Rate", "Difference", "Remarks"]
    ws_s3.append(s3_headers)
    for col in range(1, len(s3_headers) + 1):
        ws_s3.cell(1, col).fill = header_fill
        ws_s3.cell(1, col).font = header_font
        
    for issue in all_unloading_variances:
        ws_s3.append([issue["Sheet"], issue["Consignor"], issue["Consignee"], issue["BoxType"], issue["MasterRate"], issue["DespatchRate"], issue["Difference"], issue["Remarks"]])
        curr_r = ws_s3.max_row
        ws_s3.cell(curr_r, 5).number_format = '#,##0.00'
        ws_s3.cell(curr_r, 6).number_format = '#,##0.00'
        ws_s3.cell(curr_r, 7).number_format = '#,##0.00'
        
        # Color highlighting
        fill = mismatch_fill
        for c in range(1, len(s3_headers) + 1):
            ws_s3.cell(curr_r, c).fill = fill
            ws_s3.cell(curr_r, c).border = cell_border
            
    for col in ws_s3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_s3.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 6e. Section 4 Sheet: Daily Categorized Other Expenses
    ws_s4 = wb_out.create_sheet("Daily Other Expenses")
    ws_s4.views.sheetView[0].showGridLines = True
    s4_headers = ["Branch Sheet", "Date", "Category", "Details", "Amount", "Remark"]
    ws_s4.append(s4_headers)
    for col in range(1, len(s4_headers) + 1):
        ws_s4.cell(1, col).fill = header_fill
        ws_s4.cell(1, col).font = header_font
        
    # Sort expenses by date
    try:
        sorted_expenses = sorted(all_other_expenses, key=lambda x: parse_date(x["Date"]) or datetime.min)
    except Exception:
        sorted_expenses = all_other_expenses
        
    for exp in sorted_expenses:
        ws_s4.append([exp["Sheet"], exp["Date"], exp["Category"], exp["Details"], exp["Amount"], exp["Remark"]])
        curr_r = ws_s4.max_row
        ws_s4.cell(curr_r, 5).number_format = '#,##0.00'
        for c in range(1, len(s4_headers) + 1):
            ws_s4.cell(curr_r, c).font = data_font
            ws_s4.cell(curr_r, c).border = cell_border
            
    for col in ws_s4.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_s4.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save workbook
    wb_out.save(report_path)
    print(f"Upgraded audit report generated at: {report_path}", flush=True)

    # 7. Print Terminal Summary
    print("\n" + "="*50, flush=True)
    print("RECONCILE SUMMARY (UPGRADED):", flush=True)
    print(f" - Petty Cash Balance Mismatches: {len(all_balance_mismatches)}", flush=True)
    print(f" - Missing GDM Exceptions: {len(all_missing_gdms)}", flush=True)
    print(f" - Unloading Rate Variances: {len(all_unloading_variances)}", flush=True)
    print(f" - Daily Other Expenses Logged: {len(all_other_expenses)}", flush=True)
    print("="*50 + "\n", flush=True)

    # 8. Send Email Report
    print("Preparing daily email report...", flush=True)
    recipients = ["anwar@efflogistics.biz"]
    if RECEIVER_EMAIL:
        for r in RECEIVER_EMAIL.split(","):
            email_clean = r.strip()
            if email_clean and email_clean not in recipients:
                recipients.append(email_clean)
                
    if not SENDER_EMAIL or not SENDER_PASSWORD:
        print("Warning: SENDER_EMAIL or SENDER_PASSWORD not set. Skipping email dispatch.", flush=True)
        return
        
    msg = MIMEMultipart()
    msg["From"] = SENDER_EMAIL
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = f"Daily Petty Cash Audit & Exceptions Report - {datetime.now().strftime('%Y-%m-%d')}"
    
    html_body = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            h2 {{ color: #1F497D; border-bottom: 2px solid #1F497D; padding-bottom: 5px; }}
            h3 {{ color: #1F497D; margin-top: 25px; }}
            table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 10px; text-align: left; }}
            th {{ background-color: #1F497D; color: white; }}
            tr:nth-child(even) {{ background-color: #f2f2f2; }}
            .highlight-red {{ color: #C00000; font-weight: bold; }}
            .highlight-orange {{ background-color: #FCE4D6; }}
            .highlight-yellow {{ background-color: #FFF2CC; }}
            .card {{ background-color: #f9f9f9; border: 1px solid #e0e0e0; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
        </style>
    </head>
    <body>
        <h2>Daily Petty Cash Audit & Exceptions Report</h2>
        <p>Hi Anwar,</p>
        <p>Please find the automated Petty Cash Reconciliation and Audit report summary for today. The complete detailed worksheets are attached as an Excel report.</p>
        
        <h3>1. Branch Petty Cash Float & Top-up Summary</h3>
        <table>
            <thead>
                <tr>
                    <th>Branch Name</th>
                    <th>Active Sheet Name</th>
                    <th>Closing Balance (Rs)</th>
                    <th>Total Payments (Rs)</th>
                    <th>Recommended Top-up (Rs)</th>
                </tr>
            </thead>
            <tbody>
    """
    
    for b_name, stats in branch_topup_summary.items():
        active_title = stats.get("sheet_title", stats.get("status", ""))
        rec_top = stats["recommended_topup"]
        top_style = "class='highlight-red'" if rec_top > 0 else ""
        html_body += f"""
                <tr>
                    <td>{b_name}</td>
                    <td>{active_title}</td>
                    <td>{stats['closing_balance']:,.2f}</td>
                    <td>{stats['total_payments']:,.2f}</td>
                    <td {top_style}>{rec_top:,.2f}</td>
                </tr>
        """
        
    html_body += f"""
            </tbody>
        </table>
        
        <h3>2. Audit Exceptions Summary</h3>
        <div class="card">
            <ul>
                <li><strong>Petty Cash Balance Mismatch Issues (Section 1):</strong> {len(all_balance_mismatches)} issues found</li>
                <li><strong>Missing / Unrecorded GDMs in Petty Cash (Section 2):</strong> {len(all_missing_gdms)} issues found</li>
                <li><strong>Unloading Rate Variances (Section 3):</strong> {len(all_unloading_variances)} issues found</li>
                <li><strong>Daily Other Expenses Breakdown (Section 4):</strong> {len(all_other_expenses)} payments logged</li>
            </ul>
        </div>
        
        <p>Regards,<br><strong>EFF Logistics Auto-Scheduler</strong></p>
    </body>
    </html>
    """
    
    msg.attach(MIMEText(html_body, "html"))
    
    # Attach Excel file
    try:
        with open(report_path, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename= {report_file_name}",
            )
            msg.attach(part)
            print(f"Attached Excel report to email: {report_file_name}", flush=True)
    except Exception as att_err:
        print(f"Error attaching Excel file: {att_err}", flush=True)
        
    # Send via SMTP
    try:
        print(f"Connecting to SMTP to send email to {', '.join(recipients)}...", flush=True)
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.send_message(msg)
        server.quit()
        print("🎉 Daily exceptions report email sent successfully!", flush=True)
    except Exception as mail_err:
        print(f"❌ Error sending email: {mail_err}", flush=True)

if __name__ == "__main__":
    main()
