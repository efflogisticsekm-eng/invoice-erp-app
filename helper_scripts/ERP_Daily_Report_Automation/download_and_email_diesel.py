import asyncio
import os
import sys
import argparse
import pandas as pd
import numpy as np
import io
import re
import smtplib
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.encoders import encode_base64
from playwright.async_api import async_playwright

# Master Vehicle List transcribed from user screenshot
VEHICLE_MASTER = [
    {"sno": 57, "vehicle_no": "KL41T0343", "branch": "ASIAN KOLLAM", "type": "BOLERO 1.7 TON - 4 WHEEL", "rqrd_mileage": 14.0},
    {"sno": 51, "vehicle_no": "KL41T0349", "branch": "ASIAN KOLLAM", "type": "BOLERO 1.7 TON - 4 WHEEL", "rqrd_mileage": 14.0},
    {"sno": 31, "vehicle_no": "KL41V5396", "branch": "ASIAN KOLLAM", "type": "BOLERO 2 TON - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 21, "vehicle_no": "KL41W2362", "branch": "ASIAN KOLLAM", "type": "BOLERO 2 TON - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 20, "vehicle_no": "KL41W2402", "branch": "ASIAN KOLLAM", "type": "BOLERO 2 TON - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 44, "vehicle_no": "KL41T6290", "branch": "ASIAN KOLLAM", "type": "EICHER 2095 17 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 59, "vehicle_no": "KL41R1049", "branch": "ASIAN KOLLAM", "type": "TATA 407 - 4 WHEEL", "rqrd_mileage": 10.0},
    {"sno": 53, "vehicle_no": "KL41T0305", "branch": "ASIAN THRISSUR", "type": "BOLERO 1.7 TON - 4 WHEEL", "rqrd_mileage": 14.0},
    {"sno": 49, "vehicle_no": "KL41T0308", "branch": "ASIAN THRISSUR", "type": "BOLERO 1.7 TON - 4 WHEEL", "rqrd_mileage": 14.0},
    {"sno": 52, "vehicle_no": "KL41T0318", "branch": "ASIAN THRISSUR", "type": "BOLERO 1.7 TON - 4 WHEEL", "rqrd_mileage": 14.0},
    {"sno": 50, "vehicle_no": "KL41T0325", "branch": "ASIAN THRISSUR", "type": "BOLERO 1.7 TON - 4 WHEEL", "rqrd_mileage": 14.0},
    {"sno": 30, "vehicle_no": "KL41V5310", "branch": "ASIAN THRISSUR", "type": "BOLERO 2 TON - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 12, "vehicle_no": "KL41X3227", "branch": "ASIAN THRISSUR", "type": "EICHER 2075 14 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 3, "vehicle_no": "KL41X6713", "branch": "ASIAN THRISSUR", "type": "EICHER 2075 14 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 48, "vehicle_no": "KL41R0872", "branch": "ASIAN THRISSUR", "type": "TATA 407 - 4 WHEEL", "rqrd_mileage": 10.0},
    {"sno": 60, "vehicle_no": "KL41R1047", "branch": "ASIAN THRISSUR", "type": "TATA 407 - 4 WHEEL", "rqrd_mileage": 10.0},
    {"sno": 16, "vehicle_no": "KL41W2729", "branch": "CALICUT", "type": "BOLERO 2 TON - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 39, "vehicle_no": "KL41V4310", "branch": "CALICUT", "type": "EICHER 2075 14 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 40, "vehicle_no": "KL41V4346", "branch": "CALICUT", "type": "EICHER 2075 14 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 5, "vehicle_no": "KL41X5914", "branch": "CALICUT", "type": "EICHER 2075 14 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 41, "vehicle_no": "KL41V4070", "branch": "CALICUT", "type": "EICHER 2075 17 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 25, "vehicle_no": "KL41V9406", "branch": "CALICUT", "type": "LEYLAND DOST - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 55, "vehicle_no": "KL41T0365", "branch": "EDATHALA", "type": "BOLERO 1.7 TON - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 32, "vehicle_no": "KL41V5385", "branch": "EDATHALA", "type": "BOLERO 2 TON - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 22, "vehicle_no": "KL41W2358", "branch": "EDATHALA", "type": "BOLERO 2 TON - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 19, "vehicle_no": "KL41W2479", "branch": "EDATHALA", "type": "BOLERO 2 TON - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 15, "vehicle_no": "KL41W2783", "branch": "EDATHALA", "type": "BOLERO 2 TON - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 13, "vehicle_no": "KL41W2797", "branch": "EDATHALA", "type": "BOLERO 2 TON - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 2, "vehicle_no": "KL41S1218", "branch": "NOT ASSIGNED", "type": "EICHER 2059 - 4 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 1, "vehicle_no": "KL41S1324", "branch": "EDATHALA", "type": "EICHER 2059 - 4 WHEEL", "rqrd_mileage": 9.0},
    {"sno": 4, "vehicle_no": "KL41X6268", "branch": "EDATHALA", "type": "EICHER 2075 14 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 42, "vehicle_no": "KL41V4364", "branch": "EDATHALA", "type": "EICHER 2075 17 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 37, "vehicle_no": "KL41V4853", "branch": "EDATHALA", "type": "EICHER 2075 17 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 28, "vehicle_no": "KL41V7229", "branch": "EDATHALA", "type": "EICHER 2075 17 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 46, "vehicle_no": "KL41T6165", "branch": "EDATHALA", "type": "EICHER 2095 17 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 43, "vehicle_no": "KL41T6195", "branch": "EDATHALA", "type": "EICHER 2095 17 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 45, "vehicle_no": "KL41T6122", "branch": "EDATHALA", "type": "EICHER 2095 20 FT - 6 WHEEL", "rqrd_mileage": 7.5},
    {"sno": 24, "vehicle_no": "KL41T6286", "branch": "EDATHALA", "type": "EICHER 2095 20 FT - 6 WHEEL", "rqrd_mileage": 7.5},
    {"sno": 33, "vehicle_no": "KL41V5130", "branch": "EDATHALA", "type": "EICHER 2095 20 FT - 6 WHEEL", "rqrd_mileage": 7.5},
    {"sno": 34, "vehicle_no": "KL41V5174", "branch": "EDATHALA", "type": "EICHER 2095 22FT - 6 WHEEL", "rqrd_mileage": 7.5},
    {"sno": 35, "vehicle_no": "KL41V4930", "branch": "EDATHALA", "type": "EICHER 2110 24FT - 6 WHEEL", "rqrd_mileage": 7.0},
    {"sno": 36, "vehicle_no": "KL41V5019", "branch": "EDATHALA", "type": "EICHER 2110 24FT - 6 WHEEL", "rqrd_mileage": 7.0},
    {"sno": 27, "vehicle_no": "KL41V7195", "branch": "EDATHALA", "type": "EICHER 2110 24FT - 6 WHEEL", "rqrd_mileage": 7.0},
    {"sno": 26, "vehicle_no": "KL41V9571", "branch": "EDATHALA", "type": "LEYLAND DOST - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 47, "vehicle_no": "KL41T6187", "branch": "EDATHALA", "type": "TATA ULTRA - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 29, "vehicle_no": "KL41T0331", "branch": "KANNUR", "type": "BOLERO 1.7 TON - 4 WHEEL", "rqrd_mileage": 14.0},
    {"sno": 17, "vehicle_no": "KL41W3221", "branch": "KANNUR", "type": "BOLERO 2 TON - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 7, "vehicle_no": "KL41X4354", "branch": "KANNUR", "type": "EICHER 2075 14 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 38, "vehicle_no": "KL41V4560", "branch": "KANNUR", "type": "EICHER 2075 17 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 56, "vehicle_no": "KL41T0367", "branch": "KASARGOD", "type": "BOLERO 1.7 TON - 4 WHEEL", "rqrd_mileage": 14.0},
    {"sno": 9, "vehicle_no": "KL41X4064", "branch": "KASARGOD", "type": "EICHER 2075 14 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 54, "vehicle_no": "KL41T0342", "branch": "KOLLAM", "type": "BOLERO 1.7 TON - 4 WHEEL", "rqrd_mileage": 14.0},
    {"sno": 8, "vehicle_no": "KL41X4039", "branch": "KOLLAM", "type": "EICHER 2075 14 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 11, "vehicle_no": "KL41X4096", "branch": "KOLLAM", "type": "EICHER 2075 14 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 14, "vehicle_no": "KL41W2763", "branch": "KOLLAM", "type": "BOLERO 2 TON - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 58, "vehicle_no": "KL41R0874", "branch": "KOLLAM", "type": "TATA 407 - 4 WHEEL", "rqrd_mileage": 10.0},
    {"sno": 18, "vehicle_no": "KL41W2559", "branch": "MALAPPURAM", "type": "BOLERO 2 TON - 4 WHEEL", "rqrd_mileage": 13.0},
    {"sno": 10, "vehicle_no": "KL41X4082", "branch": "MALAPPURAM", "type": "EICHER 2075 14 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 6, "vehicle_no": "KL41X5674", "branch": "MALAPPURAM", "type": "EICHER 2075 14 FT - 6 WHEEL", "rqrd_mileage": 8.0},
    {"sno": 23, "vehicle_no": "KL65M3564", "branch": "HO", "type": "CAR", "rqrd_mileage": None}
]

# Helper lookup dictionaries
VEHICLE_INFO_LOOKUP = {v["vehicle_no"].replace(" ", "").upper(): v for v in VEHICLE_MASTER}

def clean_val(val):
    if pd.isna(val):
        return ""
    val_str = str(val).strip()
    if val_str.startswith("'"):
        val_str = val_str[1:]
    return val_str.strip()

def parse_date(date_str):
    date_str = clean_val(date_str)
    if not date_str or date_str == "-":
        return None
    if date_str.endswith("'"):
        date_str = date_str[:-1]
    
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None

def normalize_vehicle(v_no):
    return re.sub(r'\s+', '', clean_val(v_no)).upper()

async def download_transactions(username, password, temp_dir, target_date=None):
    if target_date is None:
        target_date = datetime.now() - timedelta(days=1)
    start_date = target_date - timedelta(days=30)
    start_str = start_date.strftime("%d-%m-%Y")
    end_str = target_date.strftime("%d-%m-%Y")

    async with async_playwright() as p:
        print("Starting Playwright Chromium browser...")
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(viewport={"width": 1280, "height": 800})
        page = await context.new_page()
        
        login_url = "https://www.iocxtrapower.com/account/login?returnUrl=%2F"
        print(f"Navigating to login page: {login_url}...")
        await page.goto(login_url, timeout=60000)
        await page.wait_for_timeout(3000)
        
        await page.fill("#email", username)
        await page.fill("input[type='password']", password)
        print("Credentials entered. Submitting login...")
        await page.locator("button", has_text="Log In").click()
        
        print("Waiting for page redirection...")
        await page.wait_for_timeout(8000)
        
        # Bypass overlays
        try:
            print("Dismissing guided tour if visible...")
            skip_btn = page.locator("button:has-text('Skip')")
            await skip_btn.wait_for(state="visible", timeout=4000)
            await skip_btn.click()
            await page.wait_for_timeout(1000)
        except Exception:
            pass
            
        try:
            print("Dismissing independence/seasonal banner if visible...")
            close_btn = page.locator("button.btn-close")
            await close_btn.wait_for(state="visible", timeout=4000)
            await close_btn.click()
            await page.wait_for_timeout(1000)
        except Exception:
            pass
            
        print("Navigating to Financials/Transaction Details...")
        financials_link = page.locator("a[href='/Transactions/BalanceInfo']").first
        await financials_link.wait_for(state="visible", timeout=5000)
        await financials_link.click()
        await page.wait_for_timeout(4000)
        
        tx_details_link = page.locator("a[href='/Transactions/TransactionDetails']").first
        await tx_details_link.wait_for(state="visible", timeout=5000)
        await tx_details_link.click()
        await page.wait_for_timeout(8000)
        
        print(f"Setting date filter range: {start_str} to {end_str}...")
        start_input = page.locator("input[formcontrolname='startDate']")
        await start_input.wait_for(state="visible", timeout=5000)
        await start_input.click()
        await page.keyboard.press("Control+A")
        await page.keyboard.press("Backspace")
        await start_input.fill(start_str)
        await page.wait_for_timeout(500)
        
        end_input = page.locator("input[formcontrolname='endDate']")
        await end_input.wait_for(state="visible", timeout=5000)
        await end_input.click()
        await page.keyboard.press("Control+A")
        await page.keyboard.press("Backspace")
        await end_input.fill(end_str)
        await page.wait_for_timeout(500)
        
        print("Clicking Search to apply date filter...")
        await page.locator("button", has_text="Search").first.click()
        await page.wait_for_timeout(6000)
        
        print("Locating Excel export icon...")
        excel_btn = page.locator("img[src*='xls.png']").first
        await excel_btn.wait_for(state="visible", timeout=5000)
        
        print("Triggering download...")
        async with page.expect_download() as download_info:
            await excel_btn.click()
        download = await download_info.value
        
        raw_download_path = os.path.join(temp_dir, "raw_transactions_download.xlsx")
        await download.save_as(raw_download_path)
        print(f"File successfully saved to: {raw_download_path}")
        await browser.close()
        return raw_download_path

def process_data(raw_csv_path, target_date):
    print(f"Processing raw transactions from: {raw_csv_path}...")
    
    header_line_idx = -1
    lines = []
    with open(raw_csv_path, "r", encoding="utf-8") as f:
        for idx, line in enumerate(f):
            lines.append(line)
            if "Card PAN" in line and "Txn ID" in line:
                header_line_idx = idx
                
    if header_line_idx == -1:
        raise ValueError("Could not find the transaction table header row in downloaded file!")
        
    csv_data = "".join(lines[header_line_idx:])
    df = pd.read_csv(io.StringIO(csv_data), index_col=False)
    
    # Clean columns & values
    df.columns = [c.strip() for c in df.columns]
    for col in df.columns:
        df[col] = df[col].apply(clean_val)
        
    # Filter for DIESEL and Sales only
    df = df[df["Product"].str.upper().str.contains("DIESEL")]
    df = df[df["Txn Type"].str.upper() == "SALE"]
    
    df["Parsed Date"] = df["Txn Date"].apply(parse_date)
    df = df.dropna(subset=["Parsed Date"])
    
    df["Norm Vehicle"] = df["Vehicle No. (Card)"].apply(normalize_vehicle)
    
    # Sort chronologically for odometer calculations
    df = df.sort_values(by=["Norm Vehicle", "Parsed Date"])
    
    # Calculate mileage & amount per record
    records = []
    
    for vehicle, group in df.groupby("Norm Vehicle"):
        group = group.sort_values(by="Parsed Date")
        prev_odo = None
        for idx, row in group.iterrows():
            curr_odo_str = row["Odometer (User Entry)"]
            try:
                curr_odo = float(curr_odo_str) if curr_odo_str and curr_odo_str != "-" else None
            except ValueError:
                curr_odo = None
                
            qty_str = row["Quantity"]
            try:
                qty = float(qty_str) if qty_str and qty_str != "-" else 0.0
            except ValueError:
                qty = 0.0
                
            amt_str = row["Amount"]
            try:
                amt = float(amt_str) if amt_str and amt_str != "-" else 0.0
            except ValueError:
                amt = 0.0
                
            rsp_str = row["RSP"]
            try:
                rsp = float(rsp_str) if rsp_str and rsp_str != "-" else 0.0
            except ValueError:
                rsp = 0.0
            
            distance = np.nan
            mileage = np.nan
            if curr_odo is not None and prev_odo is not None:
                distance = curr_odo - prev_odo
                if distance >= 0 and qty > 0:
                    mileage = round(distance / qty, 2)
            
            # Look up branch and type
            m_v_info = VEHICLE_INFO_LOOKUP.get(vehicle)
            branch = m_v_info["branch"] if m_v_info else "OTHERS"
            v_type = m_v_info["type"] if m_v_info else "UNKNOWN"
            sno = m_v_info["sno"] if m_v_info else 999
            
            records.append({
                "SNO.": sno,
                "Vehicle No.": row["Vehicle No. (Card)"],
                "Norm Vehicle": vehicle,
                "Branch": branch,
                "Vehicle Type": v_type,
                "Txn Date": row["Txn Date"],
                "Merchant Name": row["Merchant Name"],
                "Location": row["Location"],
                "Current Odometer": curr_odo if curr_odo is not None else np.nan,
                "Previous Odometer": prev_odo if prev_odo is not None else np.nan,
                "Distance Run (km)": distance if not pd.isna(distance) else np.nan,
                "Fuel Qty (Litres)": qty,
                "RSP (Rs)": rsp,
                "Amount (Rs)": amt,
                "Mileage (km/L)": mileage if not pd.isna(mileage) else np.nan,
                "Parsed Date": row["Parsed Date"]
            })
            
            if curr_odo is not None:
                prev_odo = curr_odo
                
    processed_df = pd.DataFrame(records)
    return processed_df

def generate_excel_report(processed_df, target_date, output_path):
    print(f"Generating monthly grid Excel report: {output_path}...")
    
    year = target_date.year
    month = target_date.month
    
    import calendar
    last_day = calendar.monthrange(year, month)[1]
    
    # List of days formatted as DD/MM/YY
    month_days = [datetime(year, month, d).strftime("%d/%m/%y") for d in range(1, last_day + 1)]
    
    # Group processing for grid format
    mileage_dict = {}
    amount_dict = {}
    
    if not processed_df.empty:
        for idx, row in processed_df.iterrows():
            v_norm = normalize_vehicle(row["Vehicle No."])
            day_str = row["Parsed Date"].strftime("%d/%m/%y")
            key = (v_norm, day_str)
            
            if not pd.isna(row["Mileage (km/L)"]):
                mileage_dict[key] = row["Mileage (km/L)"]
            
            if not pd.isna(row["Amount (Rs)"]) and row["Amount (Rs)"] > 0:
                amount_dict[key] = amount_dict.get(key, 0.0) + row["Amount (Rs)"]
                
    mileage_rows = []
    amount_rows = []
    
    seen_master = set(VEHICLE_INFO_LOOKUP.keys())
    extra_vehicles = []
    
    if not processed_df.empty:
        for norm_v in processed_df["Norm Vehicle"].unique():
            if norm_v not in seen_master:
                v_row = processed_df[processed_df["Norm Vehicle"] == norm_v].iloc[0]
                extra_vehicles.append({
                    "sno": 999,
                    "vehicle_no": v_row["Vehicle No."],
                    "branch": v_row["Branch"],
                    "type": v_row["Vehicle Type"]
                })
                
    report_vehicles = VEHICLE_MASTER + extra_vehicles
    
    for v in report_vehicles:
        v_norm = normalize_vehicle(v["vehicle_no"])
        
        m_row = {
            "SNO.": v["sno"] if v["sno"] != 999 else "",
            "VEHICLENO.": v["vehicle_no"],
            "BRANCH": v["branch"],
            "VEHICLE TYPE": v["type"],
            "RQRD MILEAGE": v.get("rqrd_mileage") if v.get("rqrd_mileage") is not None else ""
        }
        
        a_row = {
            "SNO.": v["sno"] if v["sno"] != 999 else "",
            "VEHICLENO.": v["vehicle_no"],
            "BRANCH": v["branch"],
            "VEHICLE TYPE": v["type"]
        }
        
        m_vals = []
        a_vals = []
        
        for d_str in month_days:
            key = (v_norm, d_str)
            m_val = mileage_dict.get(key, np.nan)
            a_val = amount_dict.get(key, np.nan)
            
            m_row[d_str] = m_val if not pd.isna(m_val) else ""
            a_row[d_str] = a_val if not pd.isna(a_val) else ""
            
            if not pd.isna(m_val):
                m_vals.append(m_val)
            if not pd.isna(a_val):
                a_vals.append(a_val)
                
        m_row["AVG"] = round(np.mean(m_vals), 2) if m_vals else ""
        a_row["TOTAL"] = round(np.sum(a_vals), 2) if a_vals else ""
        
        mileage_rows.append(m_row)
        amount_rows.append(a_row)
        
    avg_m_row = {
        "SNO.": "",
        "VEHICLENO.": "AVERAGE",
        "BRANCH": "",
        "VEHICLE TYPE": "",
        "RQRD MILEAGE": ""
    }
    
    total_a_row = {
        "SNO.": "",
        "VEHICLENO.": "TOTAL",
        "BRANCH": "",
        "VEHICLE TYPE": ""
    }
    
    for d_str in month_days:
        col_m_vals = []
        col_a_vals = []
        for v in report_vehicles:
            v_norm = normalize_vehicle(v["vehicle_no"])
            key = (v_norm, d_str)
            m_val = mileage_dict.get(key, np.nan)
            a_val = amount_dict.get(key, np.nan)
            if not pd.isna(m_val):
                col_m_vals.append(m_val)
            if not pd.isna(a_val):
                col_a_vals.append(a_val)
                
        avg_m_row[d_str] = round(np.mean(col_m_vals), 2) if col_m_vals else ""
        total_a_row[d_str] = round(np.sum(col_a_vals), 2) if col_a_vals else ""
        
    all_m_avgs = [r["AVG"] for r in mileage_rows if r["AVG"] != ""]
    all_a_totals = [r["TOTAL"] for r in amount_rows if r["TOTAL"] != ""]
    
    avg_m_row["AVG"] = round(np.mean(all_m_avgs), 2) if all_m_avgs else ""
    total_a_row["TOTAL"] = round(np.sum(all_a_totals), 2) if all_a_totals else ""
    
    mileage_rows.append(avg_m_row)
    amount_rows.append(total_a_row)
    
    mileage_grid_df = pd.DataFrame(mileage_rows)
    amount_grid_df = pd.DataFrame(amount_rows)
    
    writer = pd.ExcelWriter(output_path, engine="openpyxl")
    mileage_grid_df.to_excel(writer, sheet_name="Daily Mileage Summary", index=False)
    amount_grid_df.to_excel(writer, sheet_name="Daily Amount Summary", index=False)
    
    workbook = writer.book
    
    # Apply cell highlights for mileage shortage
    from openpyxl.styles import PatternFill, Font
    pink_fill = PatternFill(start_color="FFD2E2", end_color="FFD2E2", fill_type="solid")
    red_font = Font(color="9C0006", bold=True)
    
    if "Daily Mileage Summary" in workbook.sheetnames:
        ws = workbook["Daily Mileage Summary"]
        headers = [cell.value for cell in ws[1]]
        try:
            veh_col_idx = headers.index("VEHICLENO.") + 1
            avg_col_idx = headers.index("AVG") + 1
        except ValueError:
            veh_col_idx = 2
            avg_col_idx = ws.max_column
            
        for r_idx in range(2, ws.max_row):
            veh_no = ws.cell(row=r_idx, column=veh_col_idx).value
            if not veh_no or str(veh_no).upper() in ("AVERAGE", "TOTAL"):
                continue
                
            v_norm = normalize_vehicle(str(veh_no))
            v_info = VEHICLE_INFO_LOOKUP.get(v_norm)
            if v_info and v_info.get("rqrd_mileage") is not None:
                rqrd = v_info["rqrd_mileage"]
                
                # Date columns (cols 6 to avg_col_idx - 1)
                for c_idx in range(6, avg_col_idx):
                    val = ws.cell(row=r_idx, column=c_idx).value
                    try:
                        if val != "" and val is not None:
                            val_float = float(val)
                            if val_float < rqrd:
                                ws.cell(row=r_idx, column=c_idx).fill = pink_fill
                                ws.cell(row=r_idx, column=c_idx).font = red_font
                    except ValueError:
                        pass
                
                # AVG column
                val = ws.cell(row=r_idx, column=avg_col_idx).value
                try:
                    if val != "" and val is not None:
                        val_float = float(val)
                        if val_float < rqrd:
                            ws.cell(row=r_idx, column=avg_col_idx).fill = pink_fill
                            ws.cell(row=r_idx, column=avg_col_idx).font = red_font
                except ValueError:
                    pass
                    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 11)
            
    writer.close()
    print("Excel report successfully created.")

def generate_email_body_html(processed_df, target_date):
    print("Generating Email Body HTML grouped by Branch...")
    
    target_date_str = target_date.strftime("%d/%m/%Y")
    
    if processed_df.empty:
        return f"""
        <html>
        <body style="font-family: Arial, sans-serif; color: #333;">
            <h2 style="color: #1F497D;">Daily IOC Xtrapower Diesel Report ({target_date_str})</h2>
            <p>No transactions found in the raw downloaded file.</p>
        </body>
        </html>
        """
        
    day_df = processed_df[processed_df["Txn Date"].str.contains(target_date_str)].copy()
    
    if day_df.empty:
        return f"""
        <html>
        <body style="font-family: Arial, sans-serif; color: #333;">
            <h2 style="color: #1F497D;">Daily IOC Xtrapower Diesel Report ({target_date_str})</h2>
            <p>No transactions recorded on <b>{target_date_str}</b>.</p>
            <p>Please refer to the attached Excel file for the full month's logs.</p>
        </body>
        </html>
        """
        
    total_amount = day_df["Amount (Rs)"].sum()
    total_qty = day_df["Fuel Qty (Litres)"].sum()
    valid_mileages = day_df["Mileage (km/L)"].dropna()
    avg_mileage = valid_mileages.mean() if not valid_mileages.empty else np.nan
    
    summary_section = f"""
    <div style="font-family: Arial, sans-serif; margin-bottom: 25px;">
        <h2 style="color: #1F497D; margin-bottom: 5px;">Daily IOC Xtrapower Diesel Report</h2>
        <p style="color: #666; margin-top: 0; font-size: 14px;">Report Date: <b>{target_date_str}</b> (12:00:00 AM to 11:59:59 PM)</p>
        
        <table style="border-collapse: collapse; width: 100%; max-width: 500px; margin-top: 15px; margin-bottom: 20px; font-size: 14px;">
            <tr style="background-color: #f2f5f9;">
                <td style="padding: 10px; border: 1px solid #ddd; font-weight: bold; width: 60%;">Total Spend Today</td>
                <td style="padding: 10px; border: 1px solid #ddd; color: #1F497D; font-weight: bold; font-size: 16px;">Rs. {total_amount:,.2f}</td>
            </tr>
            <tr>
                <td style="padding: 10px; border: 1px solid #ddd; font-weight: bold;">Total Fuel Filled</td>
                <td style="padding: 10px; border: 1px solid #ddd;">{total_qty:,.2f} Litres</td>
            </tr>
            <tr style="background-color: #f2f5f9;">
                <td style="padding: 10px; border: 1px solid #ddd; font-weight: bold;">Fleet Average Mileage</td>
                <td style="padding: 10px; border: 1px solid #ddd; font-weight: bold; color: #2E6B34;">
                    {"{:.2f} km/L".format(avg_mileage) if not pd.isna(avg_mileage) else "-"}
                </td>
            </tr>
        </table>
    </div>
    """
    
    branch_tables_html = ""
    
    for branch, b_group in day_df.groupby("Branch"):
        b_amount = b_group["Amount (Rs)"].sum()
        b_qty = b_group["Fuel Qty (Litres)"].sum()
        b_valid_m = b_group["Mileage (km/L)"].dropna()
        b_avg_m = b_valid_m.mean() if not b_valid_m.empty else np.nan
        
        rows_html = ""
        for idx, row in b_group.iterrows():
            m_val = "{:.2f}".format(row["Mileage (km/L)"]) if not pd.isna(row["Mileage (km/L)"]) else "-"
            dist_val = "{:.1f}".format(row["Distance Run (km)"]) if not pd.isna(row["Distance Run (km)"]) else "-"
            odo_curr = "{:,.0f}".format(row["Current Odometer"]) if not pd.isna(row["Current Odometer"]) else "-"
            odo_prev = "{:,.0f}".format(row["Previous Odometer"]) if not pd.isna(row["Previous Odometer"]) else "-"
            
            txn_time = row["Txn Date"]
            match = re.search(r'\d{2}:\d{2}:\d{2}', txn_time)
            if match:
                txn_time = match.group(0)
            else:
                txn_time = txn_time.split(" ")[-1] if " " in txn_time else txn_time
            
            v_norm = normalize_vehicle(row["Vehicle No."])
            v_info = VEHICLE_INFO_LOOKUP.get(v_norm)
            rqrd_mileage = v_info.get("rqrd_mileage") if v_info else None
            
            mileage_style = "color: #2E6B34; background-color: #f7faf7;"
            if rqrd_mileage is not None and not pd.isna(row["Mileage (km/L)"]):
                if row["Mileage (km/L)"] < rqrd_mileage:
                    mileage_style = "color: #9C0006; background-color: #FFD2E2;"
                    
            rqrd_val = "{:.2f}".format(rqrd_mileage) if rqrd_mileage is not None else "-"
            rows_html += f"""
            <tr>
                <td style="padding: 8px; border: 1px solid #ddd;">{row["Vehicle No."]}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{txn_time}</td>
                <td style="padding: 8px; border: 1px solid #ddd; font-size: 12px;">{row["Merchant Name"]} ({row["Location"]})</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{odo_curr}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{odo_prev}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: right; font-weight: bold; background-color: #fafafa;">{dist_val}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{row["Fuel Qty (Litres)"]:.2f}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: right; font-weight: bold; color: #1F497D;">Rs. {row["Amount (Rs)"]:,.2f}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: right; font-weight: bold; color: #555;">{rqrd_val}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: right; font-weight: bold; {mileage_style}">{m_val}</td>
            </tr>
            """
            
        b_avg_m_str = "{:.2f} km/L".format(b_avg_m) if not pd.isna(b_avg_m) else "-"
        
        branch_tables_html += f"""
        <div style="font-family: Arial, sans-serif; margin-bottom: 30px;">
            <h3 style="color: #1F497D; border-bottom: 2px solid #1F497D; padding-bottom: 5px; margin-bottom: 10px;">
                Branch: {branch}
            </h3>
            <table style="width: 100%; border-collapse: collapse; font-size: 13px; border: 1px solid #ddd;">
                <thead>
                    <tr style="background-color: #1F497D; color: white; font-weight: bold; text-align: left;">
                        <th style="padding: 10px; border: 1px solid #ddd;">Vehicle No.</th>
                        <th style="padding: 10px; border: 1px solid #ddd; text-align: center;">Time</th>
                        <th style="padding: 10px; border: 1px solid #ddd;">Merchant Name & Location</th>
                        <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">Current Odo</th>
                        <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">Previous Odo</th>
                        <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">Run (km)</th>
                        <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">Qty (L)</th>
                        <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">Amount (Rs)</th>
                        <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">Rqrd Mileage</th>
                        <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">Mileage (km/L)</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                </tbody>
            </table>
            <div style="margin-top: 5px; font-size: 13px; text-align: right; color: #555;">
                Total Spend for {branch}: <b>Rs. {b_amount:,.2f}</b> | Average Mileage: <b>{b_avg_m_str}</b>
            </div>
        </div>
        """
        
    html_body = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; color: #333; }}
        </style>
    </head>
    <body>
        {summary_section}
        {branch_tables_html}
        <hr style="border: 0; border-top: 1px solid #eee; margin: 30px 0;">
        <p style="font-size: 12px; color: #777;">
            <i>This is an automated report generated from the IOC Xtrapower portal. Please refer to the attached Excel file for the complete month-to-date grid sheet.</i>
        </p>
    </body>
    </html>
    """
    return html_body

def send_report_email(html_body, excel_path, target_date):
    sender_email = os.getenv("SENDER_EMAIL")
    sender_password = os.getenv("SENDER_PASSWORD")
    receiver_email = os.getenv("RECEIVER_EMAIL")
    
    if not sender_email or not sender_password or not receiver_email:
        print("⚠️ Email credentials (SENDER_EMAIL, SENDER_PASSWORD, RECEIVER_EMAIL) are missing. Skipping email sending.")
        return
        
    date_str = target_date.strftime("%d/%m/%Y")
    subject = f"Daily IOC Xtrapower Diesel Mileage & Amount Report - {date_str}"
    
    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = receiver_email
    msg["Subject"] = subject
    
    msg.attach(MIMEText(html_body, "html"))
    
    if os.path.exists(excel_path):
        attachment_name = f"IOC_Diesel_Report_{target_date.strftime('%Y_%m_%d')}.xlsx"
        with open(excel_path, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename={attachment_name}",
            )
            msg.attach(part)
            print(f"Attached Excel report: {attachment_name}")
            
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(sender_email, sender_password)
        
        recipients = [r.strip() for r in receiver_email.split(",") if r.strip()]
        server.send_message(msg, to_addrs=recipients)
        server.quit()
        print("🎉 Diesel mileage report email sent successfully!")
    except Exception as e:
        print("❌ Error occurred during SMTP dispatch:", e)
        raise e

async def main_orchestrator():
    parser = argparse.ArgumentParser(description="IOC Xtrapower Diesel Mileage Daily Automation Report")
    parser.add_argument("--date", help="Target date in YYYY-MM-DD format (default: yesterday)")
    parser.add_argument("--local-file", help="Path to local raw CSV file (bypass portal scraping for local testing)")
    args = parser.parse_args()
    
    if args.date:
        target_date = datetime.strptime(args.date, "%Y-%m-%d")
    else:
        target_date = datetime.now() - timedelta(days=1)
        
    print(f"Target execution date: {target_date.strftime('%Y-%m-%d')}")
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    workspace_dir = os.path.abspath(os.path.join(script_dir, "..", ".."))
    temp_dir = os.path.join(workspace_dir, "temp_reports")
    os.makedirs(temp_dir, exist_ok=True)
    
    raw_csv_path = None
    if args.local_file:
        print(f"Using local file instead of scraping: {args.local_file}")
        raw_csv_path = args.local_file
    else:
        username = os.getenv("IOC_USERNAME", "EFFLOGISKRL")
        password = os.getenv("IOC_PASSWORD", "Eff@2026Logis")
        
        try:
            raw_csv_path = await download_transactions(username, password, temp_dir, target_date=target_date)
        except Exception as e:
            print(f"❌ Portal scraping failed: {e}")
            sys.exit(1)
            
    try:
        processed_df = process_data(raw_csv_path, target_date)
    except Exception as e:
        print(f"❌ Data processing failed: {e}")
        sys.exit(1)
        
    excel_report_path = os.path.join(temp_dir, f"IOC_Diesel_Monthly_Grid_{target_date.strftime('%Y_%m')}.xlsx")
    try:
        generate_excel_report(processed_df, target_date, excel_report_path)
    except Exception as e:
        print(f"❌ Excel generation failed: {e}")
        sys.exit(1)
        
    try:
        email_body_html = generate_email_body_html(processed_df, target_date)
    except Exception as e:
        print(f"❌ Email HTML generation failed: {e}")
        sys.exit(1)
        
    try:
        send_report_email(email_body_html, excel_report_path, target_date)
    except Exception as e:
        print(f"❌ SMTP dispatch failed: {e}")
        sys.exit(1)
        
    archive_dir = os.path.join(workspace_dir, "helper_scripts", "ERP_Daily_Report_Automation", "archive")
    os.makedirs(archive_dir, exist_ok=True)
    archive_file_path = os.path.join(archive_dir, f"ioc_raw_transactions_{target_date.strftime('%Y_%m_%d')}.xlsx")
    try:
        if raw_csv_path and os.path.exists(raw_csv_path):
            import shutil
            shutil.move(raw_csv_path, archive_file_path)
            print(f"Raw data file archived at: {archive_file_path}")
    except Exception as e:
        print(f"Archiving raw data failed: {e}")

if __name__ == "__main__":
    asyncio.run(main_orchestrator())
