import os
import sys
sys.path.append("/Users/anwar/Library/Python/3.9/lib/python/site-packages")
import time
import smtplib
import argparse
from datetime import datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.encoders import encode_base64
import pandas as pd
import numpy as np
import requests
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image, ImageDraw, ImageFont
import gspread
from google.oauth2.service_account import Credentials

# 1. Load Configurations from Env
ERP_USERNAME = os.getenv("ERP_USERNAME")
ERP_PASSWORD = os.getenv("ERP_PASSWORD")
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
SENDER_EMAIL = os.getenv("SENDER_EMAIL")
SENDER_PASSWORD = os.getenv("SENDER_PASSWORD")
RECEIVER_EMAIL = os.getenv("RECEIVER_EMAIL")

# WhatsApp Business API Config
WHATSAPP_TOKEN = os.getenv("WHATSAPP_TOKEN")
WHATSAPP_PHONE_NUMBER_ID = os.getenv("WHATSAPP_PHONE_NUMBER_ID")
RECIPIENT_PHONE_NUMBER = os.getenv("RECIPIENT_PHONE_NUMBER")

# Default settings if no env (e.g. for local testing, fallback to local .env values if present)
# Try loading from local .env files
from dotenv import load_dotenv
from pathlib import Path
env_path = Path("/Users/anwar/Antigravity-Related/ERP nxt Data collection/Invoice_Extractor_Tool/.env")
if env_path.exists():
    load_dotenv(dotenv_path=env_path)

# Assign variables from environment
ERP_USERNAME = os.getenv("ERP_USERNAME")
ERP_PASSWORD = os.getenv("ERP_PASSWORD")
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
SENDER_EMAIL = os.getenv("SENDER_EMAIL")
SENDER_PASSWORD = os.getenv("SENDER_PASSWORD")
RECEIVER_EMAIL = os.getenv("RECEIVER_EMAIL")
WHATSAPP_TOKEN = os.getenv("WHATSAPP_TOKEN")
WHATSAPP_PHONE_NUMBER_ID = os.getenv("WHATSAPP_PHONE_NUMBER_ID")
RECIPIENT_PHONE_NUMBER = os.getenv("RECIPIENT_PHONE_NUMBER")

# Strip surrounding quotes from env variables (handles user paste errors in GitHub Secrets)
def clean_env_var(val):
    return val.strip("'\"") if val else ""

ERP_USERNAME = clean_env_var(ERP_USERNAME)
ERP_PASSWORD = clean_env_var(ERP_PASSWORD)
SUPABASE_URL = clean_env_var(SUPABASE_URL)
SUPABASE_KEY = clean_env_var(SUPABASE_KEY)
SENDER_EMAIL = clean_env_var(SENDER_EMAIL)
SENDER_PASSWORD = clean_env_var(SENDER_PASSWORD)
RECEIVER_EMAIL = clean_env_var(RECEIVER_EMAIL)
# Support comma-separated list of receivers and ensure salim@efflogistics.biz and shajahan@efflogistics.biz are added
receivers_list = [r.strip() for r in RECEIVER_EMAIL.split(",") if r.strip()]
if "salim@efflogistics.biz" not in receivers_list:
    receivers_list.append("salim@efflogistics.biz")
if "shajahan@efflogistics.biz" not in receivers_list:
    receivers_list.append("shajahan@efflogistics.biz")
RECEIVER_EMAIL = ", ".join(receivers_list)
WHATSAPP_TOKEN = clean_env_var(WHATSAPP_TOKEN)
WHATSAPP_PHONE_NUMBER_ID = clean_env_var(WHATSAPP_PHONE_NUMBER_ID)
RECIPIENT_PHONE_NUMBER = clean_env_var(RECIPIENT_PHONE_NUMBER)
if RECIPIENT_PHONE_NUMBER:
    RECIPIENT_PHONE_NUMBER = "".join(c for c in RECIPIENT_PHONE_NUMBER if c.isdigit())

DOWNLOAD_DIR = os.path.expanduser("~/Downloads/erp_temp_downloads")
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

# Helper to load pandas dataframe from excel or csv
def load_df(file_path):
    with open(file_path, "rb") as f:
        head = f.read(4)
    if head == b"PK\x03\x04" or head == b"\xd0\xcf\x11\xe0":
        return pd.read_excel(file_path)
    else:
        import csv
        import io
        cleaned_rows = []
        for enc in ["utf-8", "latin1", "utf-8-sig"]:
            try:
                with open(file_path, "r", encoding=enc) as f_csv:
                    reader = csv.reader(f_csv)
                    header = next(reader)
                    num_cols = len(header)
                    cleaned_rows.append(header)
                    
                    for row in reader:
                        if len(row) > num_cols:
                            last_col_val = ",".join(row[num_cols-1:])
                            cleaned_row = row[:num_cols-1] + [last_col_val]
                            cleaned_rows.append(cleaned_row)
                        else:
                            cleaned_row = row + [""] * (num_cols - len(row))
                            cleaned_rows.append(cleaned_row)
                
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerows(cleaned_rows)
                output.seek(0)
                return pd.read_csv(output)
            except Exception:
                continue
        return pd.read_csv(file_path)

def discover_consignors_and_gdms():
    print("Discovering active consignors and GDMs from Google Sheets...", flush=True)
    consignors = set()
    gdms = set()
    
    creds_path = "ERP nxt Data collection/Invoice_Extractor_Tool/credentials.json"
    if not os.path.exists(creds_path):
        print(f"Google credentials not found at {creds_path}. Skipping sheet discovery.", flush=True)
        return consignors, gdms
        
    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
        client = gspread.authorize(creds)
        # Wrapped in a retry loop to handle transient 503 errors
        all_sheets = []
        for attempt in range(5):
            try:
                all_sheets = client.openall()
                break
            except Exception as api_err:
                if attempt == 4:
                    raise api_err
                backoff_secs = (attempt + 1) * 5
                print(f"Google API Error (attempt {attempt+1}/5): {api_err}. Retrying in {backoff_secs}s...", flush=True)
                time.sleep(backoff_secs)
        
        branch_prefixes = [
            "CLT NEW Petty Cash",
            "MLPM Petty Cash",
            "KSD NEW Petty Cash",
            "NEW Petty Cash - EDATHALA",
            "KNR NEW Petty Cash",
            "KLM NEW Petty Cash",
            "KOTTAYAM Petty Cash -"
        ]
        
        for s in all_sheets:
            title = s.title.strip().upper()
            is_branch = False
            for prefix in branch_prefixes:
                if title.startswith(prefix.upper()):
                    is_branch = True
                    break
            if not is_branch:
                continue
                
            print(f"Reading active consignors and GDMs from spreadsheet: {s.title}...", flush=True)
            try:
                worksheets = s.worksheets()
                # 1. PAID LR sheet for active consignors
                paid_lr_ws = next((w for w in worksheets if w.title.strip().upper() == "PAID LR"), None)
                if paid_lr_ws:
                    data = paid_lr_ws.get_all_values()
                    if len(data) > 1:
                        # Extract from index 4 (consignor column in PAID LR)
                        for row in data[1:]:
                            if len(row) > 4 and row[4].strip():
                                c_name = row[4].strip()
                                if "," in c_name:
                                    c_name = c_name.split(",")[0].strip()
                                if c_name.upper() not in ("CONSIGNOR", ""):
                                    consignors.add(c_name)
                    time.sleep(1.5)
                
                # 2. GDM sheet for GDM numbers
                gdm_ws = next((w for w in worksheets if w.title.strip().upper() == "GDM"), None)
                if gdm_ws:
                    data = gdm_ws.get_all_values()
                    if len(data) > 4: # Data starts at row 5 (index 4)
                        headers_upper = [h.strip().upper() for h in data[3]] # Headers at row 4
                        gdm_idx = next((i for i, h in enumerate(headers_upper) if "GDM" in h or "DESPATCH NO" in h or "DESPATCH_NO" in h or "DESPATCH" in h), None)
                        if gdm_idx is not None:
                            for row in data[4:]:
                                if len(row) > gdm_idx and row[gdm_idx].strip():
                                    val = row[gdm_idx].strip()
                                    if val.upper() not in ("GDM NO", "TOTAL", "SUB TOTAL", ""):
                                        gdms.add(val)
                    time.sleep(1.5)
            except Exception as sheet_err:
                print(f"Error reading worksheets of {s.title}: {sheet_err}", flush=True)
                time.sleep(1.5)
    except Exception as e:
        print(f"Error in Google Sheets discovery: {e}", flush=True)
        
    print(f"Discovery complete. Found {len(consignors)} consignors, {len(gdms)} GDMs.", flush=True)
    return consignors, gdms

# 2. Date and String Helper Utilities
def clean_val(val, default=""):
    if pd.isna(val) or val is None or val is pd.NaT:
        return default
    s = str(val).strip()
    if s.lower() in ('nan', 'nat', 'none', 'null', '-'):
        return default
    return s

def normalize_name(val):
    if not val or pd.isna(val):
        return ""
    import re
    s = str(val).strip().upper()
    s = re.sub(r'\s*\.\s*', '.', s)
    return " ".join(s.split())

def is_other_godown_delivery(val):
    if not val or pd.isna(val):
        return False
    s = str(val).strip().upper()
    return "OTHER GODOWN" in s or "OTHER_GODOWN" in s or "GODOWN" in s

def parse_date(val):
    val_str = clean_val(val)
    if not val_str:
        return None
    val_str = val_str.strip()
    
    # Fix invalid ERP times like '22:55 PM' or '14:30 AM'
    import re
    m = re.search(r'(\d{1,2}:\d{2}(?::\d{2})?)\s*([AP]M)', val_str, re.IGNORECASE)
    if m:
        time_part = m.group(1)
        ampm = m.group(2)
        hour = int(time_part.split(':')[0])
        if hour > 12:
            # It's a 24-hour time with an AM/PM suffix. Strip the suffix.
            val_str = val_str.replace(ampm, '').strip()

    formats = (
        '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %I:%M:%S %p', '%Y-%m-%d %H:%M', '%Y-%m-%d %I:%M %p', '%Y-%m-%d',
        '%d-%m-%Y %H:%M:%S', '%d-%m-%Y %I:%M:%S %p', '%d-%m-%Y %H:%M', '%d-%m-%Y %I:%M %p', '%d-%m-%Y',
        '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %I:%M:%S %p', '%d/%m/%Y %H:%M', '%d/%m/%Y %I:%M %p', '%d/%m/%Y',
        '%d.%m.%Y %H:%M:%S', '%d.%m.%Y %I:%M:%S %p', '%d.%m.%Y %H:%M', '%d.%m.%Y %I:%M %p', '%d.%m.%Y',
        '%m/%d/%Y %H:%M:%S', '%m/%d/%Y %I:%M:%S %p', '%m/%d/%Y %H:%M', '%m/%d/%Y %I:%M %p', '%m/%d/%Y',
    )
    for fmt in formats:
        try:
            return datetime.strptime(val_str, fmt)
        except ValueError:
            continue
            
    try:
        pdt = pd.to_datetime(val_str, dayfirst=True)
        if pd.isna(pdt) or pdt is pd.NaT:
            return None
        return pdt.to_pydatetime()
    except Exception:
        try:
            # Fallback for "17:00 PM" issue
            val_str = val_str.replace(" PM", "").replace(" AM", "")
            pdt = pd.to_datetime(val_str, dayfirst=True)
            if pd.isna(pdt) or pdt is pd.NaT:
                return None
            return pdt.to_pydatetime()
        except Exception:
            return None

# Fetch supervisor mappings from Supabase via REST
def fetch_supervisor_mappings():
    print("Fetching supervisor/branch mappings from Supabase...")
    try:
        url = f"{SUPABASE_URL}/rest/v1/supervisor_branch_mapping?select=*"
        headers = {
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}"
        }
        # Include a timeout and retry mechanism for network resilience
        for attempt in range(5):
            try:
                res = requests.get(url, headers=headers, timeout=15)
                res.raise_for_status()
                data = res.json()
                print(f"Loaded {len(data)} supervisor mappings.")
                break
            except Exception as retry_err:
                if attempt == 4:
                    raise retry_err
                backoff = (attempt + 1) * 5
                print(f"Supervisor mapping fetch failed (attempt {attempt+1}/5): {retry_err}. Retrying in {backoff}s...", flush=True)
                time.sleep(backoff)
        mapping = {}
        for item in data:
            name = item.get('supervisor_name')
            branch = item.get('branch')
            if name and branch:
                mapping[normalize_name(name)] = clean_val(branch).upper()
        return mapping
    except Exception as e:
        print(f"Error fetching supervisor mappings: {e}. Defaulting to empty mapping.")
        return {}

# Fetch holidays from Supabase
def fetch_holidays():
    print("Fetching holidays from Supabase...")
    try:
        url = f"{SUPABASE_URL}/rest/v1/holidays?select=*"
        headers = {
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}"
        }
        # Include timeout and retries
        for attempt in range(5):
            try:
                res = requests.get(url, headers=headers, timeout=15)
                res.raise_for_status()
                data = res.json()
                print(f"Loaded {len(data)} holidays.")
                break
            except Exception as retry_err:
                if attempt == 4:
                    raise retry_err
                backoff = (attempt + 1) * 5
                print(f"Holiday fetch failed (attempt {attempt+1}/5): {retry_err}. Retrying in {backoff}s...", flush=True)
                time.sleep(backoff)
        holidays_list = []
        for item in data:
            d_str = item.get('date')
            if d_str:
                holidays_list.append(clean_val(d_str))
        return holidays_list
    except Exception as e:
        print(f"Error fetching holidays: {e}. Defaulting to empty list.")
        return []

# Dynamic branch resolver (guesses branch based on destination/area if supervisor is missing)
def resolve_branch_name(area, supervisor, supervisor_map):
    if supervisor:
        norm_sup = normalize_name(supervisor)
        if norm_sup in supervisor_map:
            return supervisor_map[norm_sup]
            
    if area:
        norm_area = area.strip().upper()
        for branch_name in set(supervisor_map.values()):
            if branch_name in norm_area or norm_area in branch_name:
                return branch_name
                
    return "N/A"

# 5. Playwright ERP Download
def download_erp_reports(mode="morning", from_override=None, to_override=None):
    print(f"Starting Playwright ERP download flow for mode: {mode}...")
    
    # Determine dates in IST
    ist_tz = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(ist_tz)
    
    # Resolve the target date for reporting.
    # If running before 12:00 PM (noon) IST, treat it as a delayed run for the previous calendar day.
    if mode in ("daily_evening_report", "evening", "afternoon_open_lrs"):
        if now_ist.hour < 12:
            target_date = now_ist - timedelta(days=1)
        else:
            target_date = now_ist
    else:
        target_date = now_ist
        
    if mode == "evening":
        from_date_str = from_override if from_override else target_date.strftime("%Y-%m-%d")
        to_date_str = to_override if to_override else target_date.strftime("%Y-%m-%d")
    elif mode in ("daily_evening_report", "afternoon_open_lrs"):
        # Resolve dynamic lookback to find the last working day (non-Sunday, non-holiday)
        holidays_list = fetch_holidays()
        lookback_date = target_date - timedelta(days=1)
        while lookback_date.weekday() == 6 or lookback_date.strftime("%Y-%m-%d") in holidays_list:
            lookback_date -= timedelta(days=1)
        from_date_str = from_override if from_override else lookback_date.strftime("%Y-%m-%d")
        to_date_str = to_override if to_override else target_date.strftime("%Y-%m-%d")
    elif mode == "reconcile":
        holidays_list = fetch_holidays()
        yesterday = target_date - timedelta(days=1)
        # 52-Day Rolling Window: Previous Month 21st -> Current Month -> Next Month 10th
        rolling_start = target_date - timedelta(days=52)
        from_date_str = from_override if from_override else rolling_start.strftime("%Y-%m-%d")
        to_date_str = to_override if to_override else yesterday.strftime("%Y-%m-%d")
        
    print(f"Date range resolved: fromDate={from_date_str}, toDate={to_date_str}")
    
    despatch_file_path = os.path.join(DOWNLOAD_DIR, "despatch_raw.xlsx")
    lr_file_path = os.path.join(DOWNLOAD_DIR, "lr_raw.xlsx")
    
    # Clean old files to prevent reading stale reports on failures
    for f_path in [despatch_file_path, lr_file_path]:
        if os.path.exists(f_path):
            os.remove(f_path)
            
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()
        # Set a default timeout for all Playwright actions (30 seconds)
        page.set_default_timeout(30000)
        # Optional: set default navigation timeout
        page.set_default_navigation_timeout(30000)
        
        page.on("console", lambda msg: print(f"Browser Console: {msg.text}"))
        page.on("pageerror", lambda err: print(f"Browser Page Error: {err}"))
        
        try:
            main_login_url = "https://eff.aadhocc.in/eff_2021/login"
            print(f"Navigating to login page: {main_login_url}...")
            page.goto(main_login_url)
            page.wait_for_load_state("load")
            
            if page.locator("#login_user_id").count() > 0 or "login" in page.url.lower():
                print("Performing ERP Login...")
                if page.locator("#login_user_id").count() > 0:
                    page.fill("#login_user_id", ERP_USERNAME)
                else:
                    page.fill("input[type='text']", ERP_USERNAME)
                    
                if page.locator("#login_password").count() > 0:
                    page.fill("#login_password", ERP_PASSWORD)
                else:
                    page.fill("input[type='password']", ERP_PASSWORD)
                    
                submit_button = page.locator("form#login_form button[type='submit'], button[type='submit']")
                if submit_button.count() > 0:
                    submit_button.first.click()
                else:
                    page.keyboard.press("Enter")
                    
                page.wait_for_timeout(2000)
                # Wait for navigation to complete (expecting to leave the login page)
                try:
                    page.wait_for_url(lambda u: "/login" not in u, timeout=15000)
                    page.wait_for_load_state("load")
                    print("Login complete. Current URL:", page.url)
                except Exception as nav_err:
                    print("Navigation timeout or did not leave login page. Current URL:", page.url)
            
            # 1. Download Despatch Report
            despatch_url = "https://eff.aadhocc.in/eff_2021/main/effdespatch"            
            # For afternoon_open_lrs, we ONLY need the LR report, not the despatch report
            if mode != "afternoon_open_lrs":
                print("Downloading Despatch raw report...")
                page.goto(despatch_url)
                page.wait_for_load_state("load")
                
                correct_href = f"https://eff.aadhocc.in/eff_2021/main/effdespatch/exportDespatchExcel?despatch_number=&location_id=&lr_number=&from_date={from_date_str}&to_date={to_date_str}&delivery_staff_search="
                print(f"Setting export link href to: {correct_href}")
                page.locator("a.exportDespatchExcel").wait_for(state="visible", timeout=15000)
                page.evaluate(f"document.querySelector('a.exportDespatchExcel').href = '{correct_href}'")
                
                print("Downloading Despatch raw report...")
                despatch_btn = page.locator("a.exportDespatchExcel").first
                with page.expect_download(timeout=60000) as download_info:
                    despatch_btn.click(no_wait_after=True)
                download = download_info.value
                download.save_as(despatch_file_path)
                print("Despatch raw report saved to:", despatch_file_path)
                
                print("Extracting Despatch Times from Web UI...")
                try:
                    ui_times = {}
                    
                    try:
                        ui_url = f"https://eff.aadhocc.in/eff_2021/main/effdespatch?despatch_number=&location_id=&lr_number=&from_date={from_date_str}&to_date={to_date_str}&delivery_staff_search="
                        print(f"Navigating to: {ui_url}")
                        page.goto(ui_url, timeout=30000)
                        page.wait_for_timeout(5000)  # Wait for table to load via AJAX
                        page.wait_for_selector("table tbody tr td", timeout=15000)
                    except Exception as e:
                        print(f"Failed to navigate and load table: {e}")
                        
                    try:
                        page.evaluate('''() => {
                            let select = document.querySelector('select[name$="_length"]');
                            if (select) {
                                select.value = '100';
                                select.dispatchEvent(new Event('change'));
                            }
                        }''')
                        page.wait_for_timeout(5000)  # Wait for table redraw via AJAX
                    except Exception:
                        pass
                        
                    try:
                        page.wait_for_selector("table tbody tr td", timeout=15000)
                    except Exception as e:
                        print(f"Table didn't load in time: {e}")
                        
                    try:
                        page.evaluate('''() => {
                            if (typeof $ !== 'undefined' && $.fn.dataTable) {
                                let dt = $('table').DataTable();
                                if (dt) {
                                    dt.page.len(-1).draw();
                                }
                            } else {
                                let select = document.querySelector('select[name$="_length"]');
                                if (select) {
                                    let opts = Array.from(select.options).map(o => o.value);
                                    let val = opts.includes("-1") ? "-1" : (opts.includes("100") ? "100" : opts[opts.length - 1]);
                                    select.value = val;
                                    if (typeof $ !== "undefined") $(select).val(val).trigger('change');
                                    else select.dispatchEvent(new Event("change"));
                                }
                            }
                        }''')
                        page.wait_for_timeout(5000)  # Wait for full table draw via AJAX
                    except Exception as e:
                        print(f"Could not change page length: {e}")
                        
                    # SAVE SCREENSHOT FOR DEBUGGING
                    try:
                        os.makedirs(DOWNLOAD_DIR, exist_ok=True)
                        page.screenshot(path=os.path.join(DOWNLOAD_DIR, "debug_table_page_before_loop.png"), full_page=True)
                        print(f"Saved debug screenshot to {os.path.join(DOWNLOAD_DIR, 'debug_table_page_before_loop.png')}")
                    except Exception as e:
                        print(f"Failed to save screenshot: {e}")
                        
                    max_pages = 50
                    for i in range(max_pages):
                        data = None
                        for retry in range(10):
                            try:
                                data = page.evaluate('''() => {
                                    const tables = Array.from(document.querySelectorAll("table"));
                                    const table = tables.find(t => t.innerText.toLowerCase().includes("dp no"));
                                    if (!table) return {error: "No table with 'dp no' found"};
                                    const headers = Array.from(table.querySelectorAll("th")).map(th => th.innerText.trim().toLowerCase());
                                    const dpIdx = headers.findIndex(h => h.includes("dp no"));
                                    const timeIdx = headers.findIndex(h => h.includes("dp time"));
                                    const typeIdx = headers.findIndex(h => h.includes("delivery type") || (h.includes("delivery") && h.includes("type")));
                                    const branchIdx = headers.findIndex(h => h.includes("branch"));
                                    if (dpIdx === -1) return {error: "DP No header not found", headers: headers};
                                    
                                    const rows = Array.from(table.querySelectorAll("tbody tr"));
                                    if (rows.length === 0) return {error: "Table is empty or loading"};
                                    
                                    const result = {};
                                    rows.forEach(tr => {
                                        const cells = Array.from(tr.querySelectorAll("td"));
                                        const maxIdx = Math.max(dpIdx, timeIdx, typeIdx, branchIdx);
                                        if (cells.length > maxIdx) {
                                            let dp = cells[dpIdx].innerText.trim();
                                            dp = dp.replace(/\s+/g, ""); // Remove all whitespace/newlines
                                            const time = timeIdx !== -1 ? cells[timeIdx].innerText.trim() : "";
                                            const delType = typeIdx !== -1 ? cells[typeIdx].innerText.trim() : "";
                                            const branch = branchIdx !== -1 ? cells[branchIdx].innerText.trim() : "";
                                            if (dp) {
                                                result[dp] = {
                                                    "time": time,
                                                    "delivery_type": delType,
                                                    "despatch_branch": branch
                                                };
                                            }
                                        }
                                    });
                                    return result;
                                }''')
                                if data and "error" not in data:
                                    break
                            except Exception as eval_err:
                                print(f"Retry {retry} evaluating page: {eval_err}")
                            page.wait_for_timeout(1000)
                            
                        if data and "error" in data:
                            print(f"UI extraction error on page {i}: {data['error']}")
                            if "headers" in data:
                                print(f"Found headers: {data['headers']}")
                            break
                        elif data:
                            ui_times.update(data)
                            
                        has_next = page.evaluate('''() => {
                            let links = document.querySelectorAll('a');
                            for (let a of links) {
                                if (a.innerText.trim() === '>' || a.innerText.trim() === 'Next') {
                                    // Check if it's disabled or active
                                    let li = a.closest('li');
                                    if (li && (li.classList.contains('disabled') || li.classList.contains('active'))) {
                                        return false;
                                    }
                                    a.click();
                                    return true;
                                }
                            }
                            return false;
                        }''')
                        
                        if has_next:
                            try:
                                page.wait_for_load_state("load", timeout=5000)
                            except Exception:
                                pass
                            page.wait_for_timeout(3000)
                        else:
                            break
                    import json
                    import tempfile
                    ui_times_file = os.path.join(tempfile.gettempdir(), "ui_despatch_times.json")
                    with open(ui_times_file, "w") as f:
                        json.dump(ui_times, f)
                    print(f"Extracted {len(ui_times)} despatch times from UI.")
                except Exception as e:
                    print("Error extracting UI times:", e)
            
            # 2. Download LR Report
            if mode in ("morning", "daily_evening_report", "afternoon_open_lrs", "reconcile"):
                try:
                    print("Opening a fresh page for LR report download to avoid context lockups...")
                    page.close()
                except Exception:
                    pass
                page = context.new_page()
                page.set_default_timeout(30000)
                page.set_default_navigation_timeout(30000)
                page.on("console", lambda msg: print(f"Browser Console: {msg.text}"))
                page.on("pageerror", lambda err: print(f"Browser Page Error: {err}"))
                
                lr_url = "https://eff.aadhocc.in/eff_2021/main/lr/"
                print(f"Navigating to LR page: {lr_url}...")
                page.goto(lr_url)
                page.wait_for_load_state("load")
                
                # Check if session is lost on the new page, and perform login if redirected to login page
                if page.locator("#login_user_id").count() > 0 or "login" in page.url.lower():
                    print("Session lost on new page, performing login again...")
                    if page.locator("#login_user_id").count() > 0:
                        page.fill("#login_user_id", ERP_USERNAME)
                    else:
                        page.fill("input[type='text']", ERP_USERNAME)
                        
                    if page.locator("#login_password").count() > 0:
                        page.fill("#login_password", ERP_PASSWORD)
                    else:
                        page.fill("input[type='password']", ERP_PASSWORD)
                        
                    submit_button = page.locator("form#login_form button[type='submit'], button[type='submit']")
                    if submit_button.count() > 0:
                        submit_button.first.click()
                    else:
                        page.keyboard.press("Enter")
                        
                    page.wait_for_timeout(2000)
                    try:
                        page.wait_for_url(lambda u: "/login" not in u, timeout=15000)
                        page.wait_for_load_state("load")
                        print("Re-login complete. Navigating back to LR page...")
                        page.goto(lr_url)
                        page.wait_for_load_state("load")
                    except Exception as nav_err:
                        print("Re-login navigation timeout. Current URL:", page.url)
                
                # Convert dates to DD-MM-YYYY for input fields
                if mode in ("daily_evening_report", "afternoon_open_lrs", "reconcile"):
                    from_dt = datetime.strptime(from_date_str, "%Y-%m-%d") - timedelta(days=60)
                else:
                    from_dt = datetime.strptime(from_date_str, "%Y-%m-%d")
                to_dt = datetime.strptime(to_date_str, "%Y-%m-%d")
                from_date_lr = from_dt.strftime("%d-%m-%Y")
                to_date_lr = to_dt.strftime("%d-%m-%Y")
                
                print(f"Entering dates on LR search form: Date From={from_date_lr}, Date To={to_date_lr}")
                page.fill("#search_date", from_date_lr)
                page.fill("#search_date_to", to_date_lr)
                
                print("Selecting 'All' in LR Current Status filter...")
                page.select_option("#lr_current_status", "-1")
                page.wait_for_timeout(1000)
                
                print("Downloading LR raw report...")
                lr_btn = page.locator("a.export_lr_excel, button#excelExport1, #excelExport1").first
                lr_btn.wait_for(state="visible", timeout=15000)
                with page.expect_download(timeout=60000) as download_info_lr:
                    lr_btn.click(no_wait_after=True)
                download_lr = download_info_lr.value
                download_lr.save_as(lr_file_path)
                print("LR raw report saved to:", lr_file_path)
                 
                # --- NEW: Discover Active Consignors and GDMs from Google Sheets ---
                consignors, gdms = discover_consignors_and_gdms()
                 
                # --- NEW: Download Bill Clear Excel Reports for Active Consignors ---
                if consignors:
                    print("Navigating to Bill Clear page for exports...", flush=True)
                    bill_clear_url = "https://eff.aadhocc.in/eff_2021/main/bill_clear/"
                    checkpoint_file = os.path.join(DOWNLOAD_DIR, "bill_clear_checkpoint.json")
                    processed_consignors = []
                    if os.path.exists(checkpoint_file):
                        with open(checkpoint_file, "r") as f:
                            processed_consignors = json.load(f)
                    
                    try:
                        page.goto(bill_clear_url)
                        page.wait_for_load_state("load")
                        page.wait_for_timeout(3000)
                        
                        # Check session
                        if page.locator("#login_user_id").count() > 0 or "login" in page.url.lower():
                            print("Session lost on Bill Clear navigation, re-logging in...")
                            page.goto("https://eff.aadhocc.in/eff_2021/login")
                            page.wait_for_load_state("load")
                            page.fill("#login_user_id", ERP_USERNAME)
                            page.fill("#login_password", ERP_PASSWORD)
                            page.locator("button[type='submit']").click()
                            page.wait_for_timeout(3000)
                            page.goto(bill_clear_url)
                            page.wait_for_load_state("load")
                            page.wait_for_timeout(3000)
                        
                        # Map dropdown option text to value
                        consignor_id_map = {}
                        options = page.locator("#consignor_id option").all()
                        for opt in options:
                            val = opt.get_attribute("value") or ""
                            text = opt.inner_text().strip().upper()
                            if val:
                                consignor_id_map[text] = val
                        
                        print(f"Mapped {len(consignor_id_map)} consignors from ERP dropdown.", flush=True)
                        
                        # Calculate date range for bill clear (last 60 days to match any active sheet)
                        from_date_bc = (target_date - timedelta(days=60)).strftime("%d-%m-%Y")
                        to_date_bc = target_date.strftime("%d-%m-%Y")
                        
                        for cons_name in sorted(list(consignors)):
                            if cons_name in processed_consignors:
                                continue
                                
                            cons_name_upper = cons_name.strip().upper()
                            opt_val = consignor_id_map.get(cons_name_upper)
                            if not opt_val:
                                # Try partial matching
                                for map_text, map_val in consignor_id_map.items():
                                    if cons_name_upper in map_text or map_text in cons_name_upper:
                                        opt_val = map_val
                                        break
                            
                            if not opt_val:
                                print(f"  [Warning] Consignor '{cons_name}' not found in ERP dropdown map. Skipping.", flush=True)
                                continue
                                
                            print(f"  Downloading Bill Clear for '{cons_name}' (ID: {opt_val})...", flush=True)
                            try:
                                page.goto(bill_clear_url)
                                page.wait_for_load_state("load")
                                page.wait_for_timeout(1000)
                                
                                page.evaluate("""(val) => {
                                     const selectEl = document.querySelector("#consignor_id");
                                     if (selectEl) {
                                         selectEl.value = val;
                                         selectEl.dispatchEvent(new Event('change'));
                                     }
                                     if (typeof window.jQuery !== 'undefined') {
                                         window.jQuery("#consignor_id").trigger("chosen:updated").trigger("change");
                                     }
                                 }""", opt_val)
                                page.fill("#fromDate", from_date_bc)
                                page.fill("#toDate", to_date_bc)
                                page.wait_for_timeout(500)
                                
                                search_btn = page.locator("input[type='submit'][name='search']").first
                                search_btn.click()
                                page.wait_for_timeout(3000)
                                
                                excel_btn = page.locator("#excel_new")
                                with page.expect_download(timeout=30000) as bc_download_info:
                                    excel_btn.click(no_wait_after=True)
                                bc_download = bc_download_info.value
                                bc_file_path = os.path.join(DOWNLOAD_DIR, f"bill_clear_{opt_val}.xlsx")
                                bc_download.save_as(bc_file_path)
                                print(f"    Saved Bill Clear to: {bc_file_path}", flush=True)
                                processed_consignors.append(cons_name)
                                with open(checkpoint_file, "w") as f:
                                    json.dump(processed_consignors, f)
                            except Exception as bc_err:
                                print(f"    Failed to download Bill Clear for '{cons_name}': {bc_err}", flush=True)
                    except Exception as bc_page_err:
                        print(f"  Failed to load Bill Clear page: {bc_page_err}", flush=True)
                 
                # --- NEW: Scrape GDM Print Layouts for Active GDMs ---
                if gdms:
                    print(f"Scraping print layouts for {len(gdms)} active GDMs using Playwright...", flush=True)
                    gdm_details = {}
                    gdm_checkpoint = os.path.join(DOWNLOAD_DIR, "gdm_checkpoint.json")
                    if os.path.exists(gdm_checkpoint):
                        with open(gdm_checkpoint, "r") as f:
                            gdm_details = json.load(f)
                    
                    for gdm_no in sorted(list(gdms)):
                        if gdm_no in gdm_details:
                            continue
                            
                        gdm_url = f"https://eff.aadhocc.in/eff_2021/main/effdespatch/view/{gdm_no}"
                        print(f"  Scraping GDM {gdm_no} print layout: {gdm_url}...", flush=True)
                        try:
                            page.goto(gdm_url)
                            page.wait_for_load_state("load")
                            html_content = page.content()
                            
                            from bs4 import BeautifulSoup
                            soup = BeautifulSoup(html_content, "html.parser")
                            table = soup.find("table")
                            if table:
                                rows = table.find_all("tr")
                                lr_entries = []
                                t_headers = []
                                if len(rows) > 0:
                                    t_headers = [th.get_text(strip=True).upper() for th in rows[0].find_all(["td", "th"])]
                                
                                lr_no_idx = 1
                                consignor_idx = 2
                                consignee_idx = 3
                                dest_idx = 4
                                acc_pay_idx = 6
                                topay_idx = 7
                                paid_idx = 8
                                boxes_idx = 10
                                
                                if t_headers:
                                    lr_no_idx = next((i for i, h in enumerate(t_headers) if "LRNO" in h or "LR NO" in h or "LR_NO" in h), 1)
                                    consignor_idx = next((i for i, h in enumerate(t_headers) if "CONSIGNOR" in h), 2)
                                    consignee_idx = next((i for i, h in enumerate(t_headers) if "CONSIGNEE" in h), 3)
                                    dest_idx = next((i for i, h in enumerate(t_headers) if "DESTINATION" in h), 4)
                                    acc_pay_idx = next((i for i, h in enumerate(t_headers) if "ACCOUNT PAY" in h or "ACCOUNT_PAY" in h), 6)
                                    topay_idx = next((i for i, h in enumerate(t_headers) if "TOPAY" in h or "TO PAY" in h or "TO_PAY" in h), 7)
                                    paid_idx = next((i for i, h in enumerate(t_headers) if "PAID" in h), 8)
                                    boxes_idx = next((i for i, h in enumerate(t_headers) if "BOXES" in h or "BOX" in h), 10)
                                
                                for row in rows[1:]:
                                    cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
                                    if len(cells) > max(lr_no_idx, topay_idx, paid_idx):
                                        lr_no = cells[lr_no_idx].strip()
                                        if lr_no and not lr_no.upper().startswith("TOTAL"):
                                            try:
                                                topay_val = float(cells[topay_idx].replace(",", "")) if cells[topay_idx] else 0.0
                                                paid_val = float(cells[paid_idx].replace(",", "")) if cells[paid_idx] else 0.0
                                                acc_pay_val = float(cells[acc_pay_idx].replace(",", "")) if cells[acc_pay_idx] else 0.0
                                                boxes_val = float(cells[boxes_idx].replace(",", "")) if cells[boxes_idx] else 0.0
                                            except ValueError:
                                                topay_val = 0.0
                                                paid_val = 0.0
                                                acc_pay_val = 0.0
                                                boxes_val = 0.0
                                                
                                            lr_entries.append({
                                                "lr_no": lr_no,
                                                "consignor": cells[consignor_idx] if len(cells) > consignor_idx else "",
                                                "consignee": cells[consignee_idx] if len(cells) > consignee_idx else "",
                                                "destination": cells[dest_idx] if len(cells) > dest_idx else "",
                                                "account_pay": acc_pay_val,
                                                "topay": topay_val,
                                                "paid": paid_val,
                                                "boxes": boxes_val
                                            })
                                
                                gdm_details[gdm_no] = lr_entries
                                print(f"    Scraped GDM {gdm_no} successfully: {len(lr_entries)} LRs found.", flush=True)
                            else:
                                print(f"    No table found on view page for GDM {gdm_no}.", flush=True)
                        except Exception as gdm_err:
                            print(f"    Error scraping GDM {gdm_no}: {gdm_err}", flush=True)
                            
                    # Save GDM details to json file
                    gdm_json_path = os.path.join(DOWNLOAD_DIR, "gdm_details.json")
                    import json
                    with open(gdm_json_path, "w", encoding="utf-8") as json_f:
                        json.dump(gdm_details, json_f, indent=4)
                    print(f"Saved GDM details map to: {gdm_json_path}", flush=True)
                 
                return lr_file_path if mode in ("morning", "daily_evening_report", "afternoon_open_lrs", "reconcile") else None, despatch_file_path if mode != "afternoon_open_lrs" else None, from_date_str, to_date_str
            
        except Exception as e:
            print("Error downloading from ERP:", e)
            screenshot_path = os.path.join(DOWNLOAD_DIR, "error_screenshot.png")
            try:
                page.screenshot(path=screenshot_path, timeout=5000)
                print("Screenshot saved to:", screenshot_path)
            except Exception as ss_err:
                print("Could not save error screenshot:", ss_err)
            raise e
        finally:
            browser.close()

# 6. Apply openpyxl styles
def apply_styles(ws, row_count, col_count, sheet_type="default", enable_filter=False):
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
    
    data_font = Font(name="Arial", size=10)
    data_border = Border(
        bottom=Side(style="thin", color="E2E8F0"),
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0")
    )
    
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = Border(
            top=Side(style="thin", color="000000"),
            bottom=Side(style="medium", color="000000"),
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000")
        )
        
    for row in range(2, row_count + 1):
        row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") if row % 2 == 0 else PatternFill(fill_type=None)
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = data_border
            if row % 2 == 0:
                cell.fill = row_fill
            
            if sheet_type == "summary":
                cell.alignment = Alignment(vertical="center", horizontal="left")
            elif col == 1:
                cell.alignment = Alignment(vertical="center", horizontal="left")
            else:
                cell.alignment = Alignment(vertical="center", horizontal="center")

    if enable_filter and row_count > 1 and col_count > 0:
        import openpyxl.utils
        col_letter = openpyxl.utils.get_column_letter(col_count)
        ws.auto_filter.ref = f"A1:{col_letter}{row_count}"

# Draw Pillow Dashboard Image
def get_pillow_font(size, bold=False):
    font_paths = []
    if bold:
        font_paths = [
            "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
            "/Library/Fonts/Arial Bold.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
            "Arial-Bold.ttf",
            "DejaVuSans-Bold.ttf"
        ]
    else:
        font_paths = [
            "/System/Library/Fonts/Supplemental/Arial.ttf",
            "/Library/Fonts/Arial.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            "Arial.ttf",
            "DejaVuSans.ttf"
        ]
    for p in font_paths:
        if os.path.exists(p):
            try:
                return ImageFont.truetype(p, size)
            except Exception:
                continue
    return ImageFont.load_default()

def generate_pillow_dashboard(overall_stats, date_str):
    im = Image.new("RGB", (800, 180), "#F8FAFC")
    draw = ImageDraw.Draw(im)
    
    title_font = get_pillow_font(16, bold=True)
    card_title_font = get_pillow_font(10, bold=True)
    card_number_font = get_pillow_font(24, bold=True)
    card_detail_font = get_pillow_font(9, bold=False)
    
    draw.text((30, 15), f"DAILY ERP SUMMARY ({date_str}) 📊", fill="#0F172A", font=title_font)
    
    # Card 1: Delivered
    draw.rounded_rectangle([30, 50, 260, 155], radius=10, fill="#F0FDF4", outline="#DCFCE7", width=2)
    draw.text((45, 65), "DELIVERED YESTERDAY", fill="#166534", font=card_title_font)
    draw.text((45, 85), f"{overall_stats['delivered_count']} LRs", fill="#14532D", font=card_number_font)
    draw.text((45, 125), f"Max Delay: {overall_stats['delivered_max_age']} Days ({overall_stats['delivered_max_age_count']} LRs)", fill="#15803D", font=card_detail_font)
    
    # Card 2: Returned
    draw.rounded_rectangle([280, 50, 510, 155], radius=10, fill="#FEF2F2", outline="#FEE2E2", width=2)
    draw.text((295, 65), "RETURNED YESTERDAY", fill="#991B1B", font=card_title_font)
    draw.text((295, 85), f"{overall_stats['returned_count']} LRs", fill="#7F1D1D", font=card_number_font)
    draw.text((295, 125), f"Max Aging: {overall_stats['returned_max_age']} Days ({overall_stats['returned_max_age_count']} LRs)", fill="#B91C1C", font=card_detail_font)
    
    # Card 3: Open
    draw.rounded_rectangle([530, 50, 760, 155], radius=10, fill="#EFF6FF", outline="#DBEAFE", width=2)
    draw.text((545, 65), "OPEN LRS (as of 7 AM)", fill="#1D4ED8", font=card_title_font)
    draw.text((545, 85), f"{overall_stats['open_count']} LRs", fill="#1E3A8A", font=card_number_font)
    draw.text((545, 125), f"Max Aging: {overall_stats['open_max_age']} Days ({overall_stats['open_max_age_count']} LRs)", fill="#1E40AF", font=card_detail_font)
    
    image_path = os.path.join(DOWNLOAD_DIR, "daily_dashboard.png")
    im.save(image_path)
    print(f"Pillow dashboard image generated at: {image_path}")
    return image_path

# Send WhatsApp Business Cloud API Message
def send_whatsapp_message(message_text):
    if not WHATSAPP_TOKEN or not WHATSAPP_PHONE_NUMBER_ID or not RECIPIENT_PHONE_NUMBER:
        print("⚠️ WhatsApp API credentials missing. Skipping WhatsApp notification.")
        return
        
    url = f"https://graph.facebook.com/v18.0/{WHATSAPP_PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {WHATSAPP_TOKEN}",
        "Content-Type": "application/json"
    }
    payload = {
        "messaging_product": "whatsapp",
        "recipient_type": "individual",
        "to": RECIPIENT_PHONE_NUMBER,
        "type": "text",
        "text": {
            "preview_url": False,
            "body": message_text
        }
    }
    try:
        res = requests.post(url, json=payload, headers=headers)
        if res.status_code in (200, 201):
            print("🎉 WhatsApp message sent successfully!")
        else:
            print(f"❌ Failed to send WhatsApp message: Status {res.status_code}, Response: {res.text}")
    except Exception as e:
        print(f"❌ Error sending WhatsApp message: {e}")

# Run Evening Flow
def run_evening_flow(despatch_file, supervisor_map):
    print("Running Evening Flow: Uploading today's dispatches to Supabase...")
    df = load_df(despatch_file)
    if df.empty:
        print("No dispatches found in the downloaded report.")
        return
        
    df.columns = [str(c).strip().upper() for c in df.columns]
    
    lr_col = next((c for c in df.columns if c in ['LR NO', 'LRNO', 'LR_NUMBER', 'LR_NO']), None)
    desp_date_col = next((c for c in df.columns if c in ['DESPATCH DATE', 'DISPATCH DATE', 'DESPATCH_DATE']), None)
    desp_no_col = next((c for c in df.columns if c in ['DESPATCH NO', 'DISPATCH NO', 'DESPATCH_NO']), None)
    driver_col = next((c for c in df.columns if c in ['DELIVERY DRIVER', 'DRIVER', 'DRIVER_NAME']), None)
    sup_col = next((c for c in df.columns if c in ['LD SUPERVISOR', 'SUPERVISOR', 'LD_SUPERVISOR']), None)
    box_col = next((c for c in df.columns if c in ['BOX QTY', 'BOXQTY', 'BOXES', 'QUANTITY']), None)
    dest_col = next((c for c in df.columns if c in ['DESTINATION', 'PLACE', 'AREA']), None)
    consignee_col = next((c for c in df.columns if c in ['CONSIGNEE', 'CONSIGNEE NAME', 'SHIP TO PARTY', 'CONSIGNEE_NAME']), None)
    
    del_type_col = next((c for c in df.columns if c in ['DELIVERY TYPE', 'DELIVERY_TYPE', 'DEL_TYPE'] or ('DELIVERY' in c and 'TYPE' in c)), None)
    desp_branch_col = next((c for c in df.columns if c in ['BRANCH', 'BRANCH NAME', 'BRANCH_NAME', 'DESPATCH_BRANCH', 'DESPATCH BRANCH'] or c == 'BRANCH' or ('BRANCH' in c and 'DESPATCH' in c)), None)
    
    if not lr_col or not desp_date_col:
        print("Error: Required columns (LR NO, DESPATCH DATE) not found in Despatch report.")
        return
        
    ui_times = {}
    import json
    import tempfile
    ui_times_file = os.path.join(tempfile.gettempdir(), "ui_despatch_times.json")
    if os.path.exists(ui_times_file):
        try:
            with open(ui_times_file, "r") as f:
                ui_times = json.load(f)
        except Exception:
            pass
            
    payload = []
    for _, r in df.iterrows():
        lr_val = clean_val(r[lr_col])
        if not lr_val:
            continue
            
        sup_val = clean_val(r[sup_col]) if sup_col else ""
        driver_val = clean_val(r[driver_col]) if driver_col else ""
        desp_no_val = clean_val(r[desp_no_col]) if desp_no_col else ""
        
        norm_sup = normalize_name(sup_val)
        branch_val = supervisor_map.get(norm_sup, "N/A")
        
        # Check for Other Godown Delivery Type
        ui_delivery_type = ""
        ui_despatch_branch = ""
        if desp_no_val and desp_no_val in ui_times:
            val = ui_times[desp_no_val]
            if isinstance(val, dict):
                ui_delivery_type = val.get("delivery_type", "")
                ui_despatch_branch = val.get("despatch_branch", "")
                
        del_type_val = ui_delivery_type if ui_delivery_type else (clean_val(r[del_type_col]) if del_type_col else "")
        desp_branch_val = ui_despatch_branch if ui_despatch_branch else (clean_val(r[desp_branch_col]) if desp_branch_col else "")
        clean_desp_branch = desp_branch_val.strip()
        if clean_desp_branch.upper().startswith("EFF "):
            clean_desp_branch = clean_desp_branch[4:].strip()
            
        if is_other_godown_delivery(del_type_val) and clean_desp_branch:
            resolved_b = branch_val if branch_val else "N/A"
            branch_val = f"LH-{resolved_b} to {clean_desp_branch}"
        
        box_qty_val = 0
        if box_col and pd.notna(r[box_col]):
            try:
                box_qty_val = int(float(r[box_col]))
            except ValueError:
                box_qty_val = 0
                
        dest_val = clean_val(r[dest_col]) if dest_col else ""
        consignee_val = clean_val(r[consignee_col]) if consignee_col else ""
        
        d_parsed = parse_date(clean_val(r[desp_date_col]))
        desp_date_str = d_parsed.strftime("%Y-%m-%d") if d_parsed else clean_val(r[desp_date_col])
        
        payload.append({
            "despatch_date": desp_date_str,
            "despatch_no": desp_no_val,
            "lr_no": lr_val,
            "driver_name": driver_val,
            "supervisor_name": sup_val,
            "branch": branch_val,
            "box_qty": box_qty_val,
            "destination": dest_val,
            "consignee": consignee_val
        })
        
    if not payload:
        print("No valid rows found to upload.")
        return
        
    # Bulk upload using REST API upsert
    print(f"Uploading {len(payload)} rows to Supabase daily_despatch_snapshot...")
    chunk_size = 100
    for i in range(0, len(payload), chunk_size):
        chunk = payload[i:i+chunk_size]
        url = f"{SUPABASE_URL}/rest/v1/daily_despatch_snapshot?on_conflict=despatch_date,despatch_no,lr_no"
        headers = {
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates"
        }
        res = requests.post(url, json=chunk, headers=headers)
        res.raise_for_status()
        
    print("Evening upload completed successfully!")

# Run Morning Flow
def calculate_aging_metrics(lrs, reference_date):
    if not lrs:
        return 0, 0, 0
    agings = [x["aging"] for x in lrs]
    max_age = max(agings)
    max_age_count = sum(1 for x in agings if x == max_age)
    unique_pts = len(set(f"{x['consignee'].lower()}||{x['destination'].lower()}" for x in lrs))
    return max_age, max_age_count, unique_pts

def calculate_aging(start_date, end_date, holidays=None):
    """
    Calculates aging between start_date and end_date (both are datetime.date or datetime.datetime objects or parseable dates),
    excluding Sundays and optionally holidays.
    """
    if not start_date or not end_date:
        return 0
    if isinstance(start_date, str):
        start_date = parse_date(start_date)
    if isinstance(end_date, str):
        end_date = parse_date(end_date)
        
    # Convert to date objects if they are datetime objects
    if isinstance(start_date, datetime):
        start_date = start_date.date()
    if isinstance(end_date, datetime):
        end_date = end_date.date()
        
    if end_date <= start_date:
        return 0
        
    days = 0
    curr = start_date
    while curr < end_date:
        curr += timedelta(days=1)
        # Exclude Sundays (weekday 6 is Sunday)
        if curr.weekday() == 6:
            continue
        # Exclude holidays if provided
        if holidays:
            d_str = curr.strftime("%Y-%m-%d")
            d_str2 = curr.strftime("%d/%m/%Y")
            if d_str in holidays or d_str2 in holidays:
                continue
        days += 1
    return days

def run_daily_evening_report_flow(lr_file, despatch_file, supervisor_map, yesterday_str, today_str, start_time_override=None, end_time_override=None):
    start_time = start_time_override if start_time_override else "19:00:00"
    end_time = end_time_override if end_time_override else "19:00:00"
    print(f"Running Daily Evening Flow: {yesterday_str} {start_time} to {today_str} {end_time}")
    
    # Load holidays
    holidays = fetch_holidays()
    
    # 1. Load Despatch Report
    df_desp = load_df(despatch_file)
    if df_desp.empty:
        print("Error: Despatch report is empty.")
        sys.exit(1)
        
    df_desp.columns = [str(c).strip().upper() for c in df_desp.columns]
    
    # Required columns
    lr_col_desp = next((c for c in df_desp.columns if c in ['LR NO', 'LRNO', 'LR_NUMBER', 'LR_NO']), None)
    sup_col_desp = next((c for c in df_desp.columns if c in ['LD SUPERVISOR', 'SUPERVISOR', 'LD_SUPERVISOR']), None)
    driver_col_desp = next((c for c in df_desp.columns if c in ['DELIVERY DRIVER', 'DRIVER', 'DRIVER_NAME']), None)
    desp_no_col_desp = next((c for c in df_desp.columns if c in ['DESPATCH NO', 'DISPATCH NO', 'DESPATCH_NO']), None)
    date_col_desp = next((c for c in df_desp.columns if c in ['DP DATE', 'DESPATCH DATE', 'DISPATCH DATE']), None)
    time_col_desp = next((c for c in df_desp.columns if c in ['DP TIME', 'DESPATCH TIME', 'DISPATCH TIME']), None)
    del_type_col_desp = next((c for c in df_desp.columns if c in ['DELIVERY TYPE', 'DELIVERY_TYPE', 'DEL_TYPE'] or ('DELIVERY' in c and 'TYPE' in c)), None)
    branch_col_desp = next((c for c in df_desp.columns if c in ['BRANCH', 'BRANCH NAME', 'BRANCH_NAME', 'DESPATCH_BRANCH', 'DESPATCH BRANCH'] or c == 'BRANCH' or ('BRANCH' in c and 'DESPATCH' in c)), None)
    
    # Date time boundaries
    start_dt = datetime.strptime(f"{yesterday_str} {start_time}", "%Y-%m-%d %H:%M:%S")
    end_dt = datetime.strptime(f"{today_str} {end_time}", "%Y-%m-%d %H:%M:%S")
    
    # Filter Despatch Data
    import json
    import tempfile
    ui_times = {}
    ui_times_file = os.path.join(tempfile.gettempdir(), "ui_despatch_times.json")
    if os.path.exists(ui_times_file):
        try:
            with open(ui_times_file, "r") as f:
                ui_times = json.load(f)
        except Exception:
            pass
            

            
    # Extract supervisor and driver from despatch to avoid empty values on some LRs
    desp_meta = {}
    for _, r in df_desp.iterrows():
        dp_no = clean_val(r[desp_no_col_desp]) if desp_no_col_desp else ""
        if dp_no:
            sup = clean_val(r[sup_col_desp]) if sup_col_desp else ""
            drv = clean_val(r[driver_col_desp]) if driver_col_desp else ""
            if dp_no not in desp_meta:
                desp_meta[dp_no] = {"supervisor": sup, "driver": drv}
            else:
                if sup: desp_meta[dp_no]["supervisor"] = sup
                if drv: desp_meta[dp_no]["driver"] = drv

    filtered_despatches = {} # Map from LR No -> Despatch Info
    for _, r in df_desp.iterrows():
        lr = clean_val(r[lr_col_desp])
        if not lr: continue
        
        dp_no = clean_val(r[desp_no_col_desp]) if desp_no_col_desp else ""
        
        # Combine date and time
        d_val = clean_val(r[date_col_desp]) if date_col_desp else ""
        t_val = clean_val(r[time_col_desp]) if time_col_desp else ""
        
        ui_delivery_type = ""
        ui_despatch_branch = ""
        if dp_no and dp_no in ui_times:
            val = ui_times[dp_no]
            if isinstance(val, dict):
                t_val = val.get("time", "")
                ui_delivery_type = val.get("delivery_type", "")
                ui_despatch_branch = val.get("despatch_branch", "")
            else:
                t_val = val
        
        # Try to parse DP Date and DP Time
        dt_obj = parse_date(f"{d_val} {t_val}") if t_val else parse_date(d_val)
        
        if dt_obj:
            if "AM" not in t_val.upper() and "PM" not in t_val.upper():
                if dt_obj.date() == start_dt.date() and dt_obj.hour < 12 and dt_obj.hour >= 1:
                    dt_obj = dt_obj + timedelta(hours=12)
        
        if dt_obj and start_dt <= dt_obj <= end_dt:
            # Extract supervisor, driver, and despatch_no with N/A fallback
            sup_val = desp_meta.get(dp_no, {}).get("supervisor") or (clean_val(r[sup_col_desp]) if sup_col_desp else "")
            drv_val = desp_meta.get(dp_no, {}).get("driver") or (clean_val(r[driver_col_desp]) if driver_col_desp else "")
            
            if not sup_val or pd.isna(sup_val): sup_val = "N/A"
            if not drv_val or pd.isna(drv_val): drv_val = "N/A"
            dp_no_val = dp_no if dp_no else "N/A"
            
            # Valid despatch in time window
            filtered_despatches[lr] = {
                "lr_no": lr,
                "supervisor": sup_val,
                "driver": drv_val,
                "despatch_no": dp_no_val,
                "dp_date": dt_obj,
                "delivery_type": ui_delivery_type if ui_delivery_type else (clean_val(r[del_type_col_desp]) if del_type_col_desp else ""),
                "despatch_branch": ui_despatch_branch if ui_despatch_branch else (clean_val(r[branch_col_desp]) if branch_col_desp else "")
            }
            
    print(f"Filtered {len(filtered_despatches)} LRs dispatched between {start_dt} and {end_dt}")
    
    # 3. Load LR Report
    df_lr = load_df(lr_file)
    if df_lr.empty:
        print("Error: LR raw report is empty.")
        sys.exit(1)
        
    df_lr.columns = [str(c).strip().upper() for c in df_lr.columns]
    
    lr_col = next((c for c in df_lr.columns if c in ['LR NO', 'LRNO', 'LR_NUMBER', 'LR_NO']), None)
    date_col = next((c for c in df_lr.columns if c in ['DATE', 'LR DATE', 'LR_DATE']), None)
    del_time_col = next((c for c in df_lr.columns if c in ['DELIVERY TIME', 'DELIVERY_TIME', 'DELIVERED_DATE']), None)
    consignor_col = next((c for c in df_lr.columns if c in ['CONSIGNOR', 'CONSIGNOR_NAME']), None)
    consignee_col = next((c for c in df_lr.columns if c in ['CONSIGNEE', 'CONSIGNEE_NAME']), None)
    dest_col = next((c for c in df_lr.columns if c in ['DESTINATION', 'PLACE', 'AREA']), None)
    status_col = next((c for c in df_lr.columns if c in ['LR STATUS', 'LRSTATUS', 'STATUS']), None)
    box_col = next((c for c in df_lr.columns if c in ['BOX QTY', 'BOXQTY', 'BOXES', 'QUANTITY']), None)
    
    # Build LR map from LR report
    lr_rows_map = {clean_val(r[lr_col]): r for _, r in df_lr.iterrows() if clean_val(r[lr_col])}
    
    branch_stats = {}
    unmapped_supervisors = set()
    
    # Lists for spreadsheet generation
    despatch_snapshot_rows = []
    open_lrs_rows = []
    
    # Track overall metrics
    overall_stats = {
        "delivered_count": 0, "returned_count": 0, "open_count": 0,
        "delivered_max_age": 0, "delivered_max_age_count": 0,
        "returned_max_age": 0, "returned_max_age_count": 0,
        "open_max_age": 0, "open_max_age_count": 0
    }
    
    # Used for aggregating WhatsApp driver breakdown
    driver_breakdowns = {}
    
    for lr_no, snap in filtered_despatches.items():
        row = lr_rows_map.get(lr_no)
        
        consignor = clean_val(row[consignor_col]) if row is not None else ""
        if consignor.upper().startswith('EFF'):
            continue
            
        consignee = clean_val(row[consignee_col]) if row is not None else ""
        destination = clean_val(row[dest_col]) if row is not None else ""
        box_qty = int(float(row[box_col])) if (row is not None and pd.notna(row[box_col])) else 0
        lr_date = clean_val(row[date_col]) if row is not None else ""
        del_time = clean_val(row[del_time_col]) if row is not None else ""
        
        raw_status = clean_val(row[status_col], "Open") if row is not None else "Open"
        status_map = {
            'Despatched': 'On transit',
            'Open': 'Not Despatched',
            'Delivered': 'Delivery Process completed.',
            'Despatched from Branch': 'Cancelled LR'
        }
        mapped_status = status_map.get(raw_status, "Not Despatched")
        if "cancelled" in raw_status.lower() or mapped_status == 'Cancelled LR':
            mapped_status = 'Cancelled LR'
            
        supervisor = snap.get("supervisor", "")
        driver = snap.get("driver", "")
        despatch_no = snap.get("despatch_no", "")
        dp_date_str = snap.get("dp_date").strftime("%Y-%m-%d %H:%M:%S")
        
        norm_sup = normalize_name(supervisor)
        if supervisor and supervisor != "N/A" and supervisor_map.get(norm_sup):
            branch = supervisor_map.get(norm_sup)
        else:
            branch = resolve_branch_name(destination, supervisor, supervisor_map)
            
        # Check if Delivery Type is Other Godown and Despatch Branch is present
        del_type_val = snap.get("delivery_type", "")
        desp_branch_val = snap.get("despatch_branch", "")
        clean_desp_branch = desp_branch_val.strip()
        if clean_desp_branch.upper().startswith("EFF "):
            clean_desp_branch = clean_desp_branch[4:].strip()
            
        if is_other_godown_delivery(del_type_val) and clean_desp_branch:
            resolved_b = branch if branch else "N/A"
            branch = f"LH-{resolved_b} to {clean_desp_branch}"

        
        if supervisor and supervisor != "N/A" and branch == "N/A":
            unmapped_supervisors.add(supervisor)
            
        lr_date_obj = parse_date(lr_date)
        is_delivered = (mapped_status == 'Delivery Process completed.')
        if is_delivered:
            del_time_obj = parse_date(del_time)
            aging = calculate_aging(lr_date_obj, del_time_obj, holidays)
        else:
            aging = calculate_aging(lr_date_obj, end_dt, holidays)
            
        lr_item = {
            "lr_no": lr_no,
            "consignee": consignee,
            "destination": destination,
            "box_qty": box_qty,
            "lr_date": lr_date,
            "delivery_time": del_time,
            "status": mapped_status,
            "driver": driver,
            "despatch_no": despatch_no,
            "despatch_time": dp_date_str,
            "aging": aging
        }
        
        is_delivered = (mapped_status == 'Delivery Process completed.')
        
        if branch not in branch_stats:
            branch_stats[branch] = {
                "delivered_lrs": [], "returned_lrs": [], "open_lrs": [], "dispatches": {}
            }
            
        b_group = branch_stats[branch]
        
        if is_delivered:
            b_group["delivered_lrs"].append(lr_item)
            overall_stats["delivered_count"] += 1
        elif mapped_status == 'On transit':
            b_group["returned_lrs"].append(lr_item)
            overall_stats["returned_count"] += 1
            open_lrs_rows.append({
                "Branch": branch, "Despatch No": despatch_no, "Despatch Time": dp_date_str,
                "Driver": driver, "LR No": lr_no, "LR Date": lr_date, "Consignee": consignee,
                "Destination": destination, "Current Status": "Returned (On transit)", "Aging (Days)": aging
            })
        else:
            b_group["open_lrs"].append(lr_item)
            overall_stats["open_count"] += 1
            open_lrs_rows.append({
                "Branch": branch, "Despatch No": despatch_no, "Despatch Time": dp_date_str,
                "Driver": driver, "LR No": lr_no, "LR Date": lr_date, "Consignee": consignee,
                "Destination": destination, "Current Status": mapped_status, "Aging (Days)": aging
            })
            
        disp_key = (driver, despatch_no)
        if disp_key not in b_group["dispatches"]:
            b_group["dispatches"][disp_key] = {"delivered_lrs": [], "returned_lrs": [], "open_lrs": []}
            
        disp_group = b_group["dispatches"][disp_key]
        if is_delivered:
            disp_group["delivered_lrs"].append(lr_item)
        elif mapped_status == 'On transit':
            disp_group["returned_lrs"].append(lr_item)
        else:
            disp_group["open_lrs"].append(lr_item)
            
        if driver and despatch_no:
            d_key = f"{branch}|{driver}|{despatch_no}"
            if d_key not in driver_breakdowns:
                driver_breakdowns[d_key] = {"delivered": [], "returned": [], "open": [], "points": set()}
            if is_delivered:
                driver_breakdowns[d_key]["delivered"].append(lr_item)
            elif mapped_status == 'On transit':
                driver_breakdowns[d_key]["returned"].append(lr_item)
            else:
                driver_breakdowns[d_key]["open"].append(lr_item)
            
            # Count unique points for all LRs in the despatch, or just delivered/returned?
            # Let's count for all LRs given to the driver in this despatch.
            driver_breakdowns[d_key]["points"].add(f"{consignee.lower()}||{destination.lower()}")
            
        despatch_snapshot_rows.append({
            "Branch": branch,
            "Despatch No": despatch_no,
            "Despatch Time": dp_date_str,
            "Driver Name": driver,
            "Supervisor Name": supervisor,
            "LR No": lr_no,
            "LR Date": lr_date,
            "Consignee": consignee,
            "Destination": destination,
            "Box Qty": box_qty,
            "Current Status": mapped_status,
            "Delivery Time": del_time if is_delivered else "-",
            "LR Age (Days)": aging
        })
        
    # Process branch summary for Excel and WhatsApp
    wb = openpyxl.Workbook()
    
    # 1. Summary Sheet
    ws_sum = wb.active
    ws_sum.title = "Daily Summary"
    ws_sum.append(["BRANCH", "DELIVERED", "DESPATCHED (RETURNED)", "OPEN"])
    
    whatsapp_msg = f"📊 *DAILY DESPATCH & DELIVERY SUMMARY ({yesterday_str} 8PM - {today_str} 8PM)*\n"
    
    all_delivered_lrs = []
    all_returned_lrs = []
    all_open_lrs = []
    
    for branch, stats in branch_stats.items():
        del_count = len(stats["delivered_lrs"])
        ret_count = len(stats["returned_lrs"])
        opn_count = len(stats["open_lrs"])
        ws_sum.append([branch, del_count, ret_count, opn_count])
        
        all_delivered_lrs.extend(stats["delivered_lrs"])
        all_returned_lrs.extend(stats["returned_lrs"])
        all_open_lrs.extend(stats["open_lrs"])
        
    ws_sum.append(["GRAND TOTAL", overall_stats["delivered_count"], overall_stats["returned_count"], overall_stats["open_count"]])
    apply_styles(ws_sum, ws_sum.max_row, ws_sum.max_column, sheet_type="summary")
    
    # Adjust column widths
    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 15
    ws_sum.column_dimensions["C"].width = 25
    ws_sum.column_dimensions["D"].width = 15
    
    # WhatsApp detailed breakdown per branch
    for branch in sorted(branch_stats.keys()):
        stats = branch_stats[branch]
        whatsapp_msg += f"\n🏢 *{branch}*\n"
        
        # Group by driver within branch
        for (driver, desp_no), dstats in stats["dispatches"].items():
            del_count = len(dstats["delivered_lrs"])
            ret_count = len(dstats["returned_lrs"])
            open_count = len(dstats.get("open_lrs", []))
            total_lrs = del_count + ret_count + open_count
            if total_lrs == 0:
                continue
                
            d_key = f"{branch}|{driver}|{desp_no}"
            pts = len(driver_breakdowns.get(d_key, {}).get("points", set()))
            
            times = [parse_date(x["delivery_time"]) for x in dstats["delivered_lrs"] if x["delivery_time"] and parse_date(x["delivery_time"])]
            t_str = ""
            if times:
                first_t = min(times).strftime("%I:%M %p")
                last_t = max(times).strftime("%I:%M %p")
                t_str = f", 1st: {first_t}, Last: {last_t}"
                
            whatsapp_msg += (f"  🚚 {driver} (Desp: {desp_no})\n"
                             f"     Total LRs: {total_lrs}, Points: {pts}{t_str}\n"
                             f"     Delivered: {del_count}, Returned: {ret_count}\n")
                             
    # Max Aging logic for Dashboard
    overall_stats["delivered_max_age"], overall_stats["delivered_max_age_count"], _ = calculate_aging_metrics(all_delivered_lrs, None)
    overall_stats["returned_max_age"], overall_stats["returned_max_age_count"], _ = calculate_aging_metrics(all_returned_lrs, None)
    overall_stats["open_max_age"], overall_stats["open_max_age_count"], _ = calculate_aging_metrics(all_open_lrs, None)
    
    whatsapp_msg += f"\n*OVERALL STATUS*\n"
    whatsapp_msg += f"✅ Delivered: {overall_stats['delivered_count']}\n"
    whatsapp_msg += f"⚠️ Returned: {overall_stats['returned_count']}\n"
    whatsapp_msg += f"⏳ Open: {overall_stats['open_count']}\n"
    
    # 2. Despatch Snapshot Sheet
    ws_snap = wb.create_sheet("Despatch Snapshot")
    headers = ["Branch", "Despatch No", "Despatch Time", "Driver Name", "Supervisor Name", "LR No", "LR Date", "Consignee", "Destination", "Box Qty", "Current Status", "Delivery Time", "LR Age (Days)"]
    ws_snap.append(headers)
    for r in despatch_snapshot_rows:
        ws_snap.append([r[h] for h in headers])
    apply_styles(ws_snap, ws_snap.max_row, ws_snap.max_column, enable_filter=True)
    
    # 3. Open/Returned LRs Sheet
    ws_open = wb.create_sheet("Open LRs")
    headers_open = ["Branch", "Despatch No", "Despatch Time", "Driver", "LR No", "LR Date", "Consignee", "Destination", "Current Status", "Aging (Days)"]
    ws_open.append(headers_open)
    for r in open_lrs_rows:
        ws_open.append([r[h] for h in headers_open])
    apply_styles(ws_open, ws_open.max_row, ws_open.max_column, enable_filter=True)
    
    # 4. Despatch Summary Sheet
    ws_ds = wb.create_sheet("Despatch Summary")
    headers_ds = ["Branch", "Driver", "Despatch No", "Despatch Time", "Total LRs", "Delivery Points", "Delivered", "Despatched (Returned)", "1st Delivery", "Last Delivery"]
    ws_ds.append(headers_ds)
    for b_d_d, d_info in driver_breakdowns.items():
        pts = len(d_info["points"])
        del_lrs = d_info["delivered"]
        ret_lrs = d_info["returned"]
        open_lrs = d_info.get("open", [])
        total_lrs = len(del_lrs) + len(ret_lrs) + len(open_lrs)
        b, dr, dn = b_d_d.split("|")
        
        times = [parse_date(x["delivery_time"]) for x in del_lrs if x["delivery_time"] and parse_date(x["delivery_time"])]
        f_time = min(times).strftime("%d/%m %I:%M %p") if times else "-"
        l_time = max(times).strftime("%d/%m %I:%M %p") if times else "-"
        
        # Extract despatch time from the items
        desp_time = ""
        if del_lrs:
            desp_time = del_lrs[0].get("despatch_time", "")
        elif ret_lrs:
            desp_time = ret_lrs[0].get("despatch_time", "")
        elif open_lrs:
            desp_time = open_lrs[0].get("despatch_time", "")
            
        ws_ds.append([b, dr, dn, desp_time, total_lrs, pts, len(del_lrs), len(ret_lrs), f_time, l_time])
    apply_styles(ws_ds, ws_ds.max_row, ws_ds.max_column, enable_filter=True)
    
    processed_file_path = os.path.join(DOWNLOAD_DIR, f"Daily_Evening_Report_{today_str}.xlsx")
    wb.save(processed_file_path)
    print(f"Processed Excel report saved to: {processed_file_path}")
    
    dashboard_image_path = generate_pillow_dashboard(overall_stats, today_str)
    
    send_whatsapp_message(whatsapp_msg)
    
    delay_tables_html = generate_delay_tables_html(despatch_snapshot_rows)
    return processed_file_path, dashboard_image_path, unmapped_supervisors, delay_tables_html

def run_morning_flow(lr_file, despatch_file, supervisor_map, yesterday_str):
    print(f"Running Morning Flow: Analyzing deliveries for date {yesterday_str}...")
    
    # Load holidays
    holidays = fetch_holidays()
    
    # 1. Fetch yesterday's snapshot from Supabase REST API
    url = f"{SUPABASE_URL}/rest/v1/daily_despatch_snapshot?despatch_date=eq.{yesterday_str}"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}"
    }
    res = requests.get(url, headers=headers)
    res.raise_for_status()
    snapshot_data = res.json()
    print(f"Fetched {len(snapshot_data)} yesterday dispatches from Supabase.")
    
    # Group yesterday's snapshot dispatches by LR NO
    snapshot_lrs = {item["lr_no"]: item for item in snapshot_data if item.get("lr_no")}
    
    ui_times = {}
    import json
    import tempfile
    ui_times_file = os.path.join(tempfile.gettempdir(), "ui_despatch_times.json")
    if os.path.exists(ui_times_file):
        try:
            with open(ui_times_file, "r") as f:
                ui_times = json.load(f)
        except Exception:
            pass
            
    # 2. Build 30-day dispatch map from the downloaded despatch file as supervisor fallback
    df_desp = load_df(despatch_file)
    desp_30d_map = {}
    if not df_desp.empty:
        df_desp.columns = [str(c).strip().upper() for c in df_desp.columns]
        lr_col_desp = next((c for c in df_desp.columns if c in ['LR NO', 'LRNO', 'LR_NUMBER', 'LR_NO']), None)
        sup_col_desp = next((c for c in df_desp.columns if c in ['LD SUPERVISOR', 'SUPERVISOR', 'LD_SUPERVISOR']), None)
        driver_col_desp = next((c for c in df_desp.columns if c in ['DELIVERY DRIVER', 'DRIVER', 'DRIVER_NAME']), None)
        desp_no_col_desp = next((c for c in df_desp.columns if c in ['DESPATCH NO', 'DISPATCH NO', 'DESPATCH_NO']), None)
        dest_col_desp = next((c for c in df_desp.columns if c in ['DESTINATION', 'PLACE', 'AREA']), None)
        del_type_col_desp = next((c for c in df_desp.columns if c in ['DELIVERY TYPE', 'DELIVERY_TYPE', 'DEL_TYPE'] or ('DELIVERY' in c and 'TYPE' in c)), None)
        branch_col_desp = next((c for c in df_desp.columns if c in ['BRANCH', 'BRANCH NAME', 'BRANCH_NAME', 'DESPATCH_BRANCH', 'DESPATCH BRANCH'] or c == 'BRANCH' or ('BRANCH' in c and 'DESPATCH' in c)), None)
        
        for _, r in df_desp.iterrows():
            lr = clean_val(r[lr_col_desp])
            if lr:
                sup_val = clean_val(r[sup_col_desp]) if sup_col_desp else ""
                norm_sup = normalize_name(sup_val)
                if sup_val and sup_val != "N/A" and supervisor_map.get(norm_sup):
                    branch_val = supervisor_map.get(norm_sup)
                else:
                    dest_val = clean_val(r[dest_col_desp]) if dest_col_desp else ""
                    branch_val = resolve_branch_name(dest_val, sup_val, supervisor_map)
                
                # Check for Other Godown
                dp_no_val = clean_val(r[desp_no_col_desp]) if desp_no_col_desp else ""
                ui_delivery_type = ""
                ui_despatch_branch = ""
                if dp_no_val and dp_no_val in ui_times:
                    val = ui_times[dp_no_val]
                    if isinstance(val, dict):
                        ui_delivery_type = val.get("delivery_type", "")
                        ui_despatch_branch = val.get("despatch_branch", "")
                        
                del_type_val = ui_delivery_type if ui_delivery_type else (clean_val(r[del_type_col_desp]) if del_type_col_desp else "")
                desp_branch_val = ui_despatch_branch if ui_despatch_branch else (clean_val(r[branch_col_desp]) if branch_col_desp else "")
                clean_desp_branch = desp_branch_val.strip()
                if clean_desp_branch.upper().startswith("EFF "):
                    clean_desp_branch = clean_desp_branch[4:].strip()
                
                if is_other_godown_delivery(del_type_val) and clean_desp_branch:
                    resolved_b = branch_val if branch_val else "N/A"
                    branch_val = f"LH-{resolved_b} to {clean_desp_branch}"
                
                desp_30d_map[lr] = {
                    "supervisor": sup_val,
                    "driver": clean_val(r[driver_col_desp]) if driver_col_desp else "",
                    "despatch_no": clean_val(r[desp_no_col_desp]) if desp_no_col_desp else "",
                    "branch": branch_val
                }
                
    # 3. Load LR Report
    df_lr = load_df(lr_file)
    if df_lr.empty:
        print("Error: LR raw report is empty.")
        sys.exit(1)
        
    df_lr.columns = [str(c).strip().upper() for c in df_lr.columns]
    
    lr_col = next((c for c in df_lr.columns if c in ['LR NO', 'LRNO', 'LR_NUMBER', 'LR_NO']), None)
    date_col = next((c for c in df_lr.columns if c in ['DATE', 'LR DATE', 'LR_DATE']), None)
    del_time_col = next((c for c in df_lr.columns if c in ['DELIVERY TIME', 'DELIVERY_TIME', 'DELIVERED_DATE']), None)
    consignor_col = next((c for c in df_lr.columns if c in ['CONSIGNOR', 'CONSIGNOR_NAME']), None)
    consignee_col = next((c for c in df_lr.columns if c in ['CONSIGNEE', 'CONSIGNEE_NAME']), None)
    dest_col = next((c for c in df_lr.columns if c in ['DESTINATION', 'PLACE', 'AREA']), None)
    status_col = next((c for c in df_lr.columns if c in ['LR STATUS', 'LRSTATUS', 'STATUS']), None)
    box_col = next((c for c in df_lr.columns if c in ['BOX QTY', 'BOXQTY', 'BOXES', 'QUANTITY']), None)
    
    # Build LR map from LR report
    lr_rows_map = {clean_val(r[lr_col]): r for _, r in df_lr.iterrows() if clean_val(r[lr_col])}
    
    yesterday_dt = datetime.strptime(yesterday_str, "%Y-%m-%d")
    yesterday_7am = datetime(yesterday_dt.year, yesterday_dt.month, yesterday_dt.day, 7, 0, 0)
    
    branch_stats = {}
    unmapped_supervisors = set()
    
    # Lists for spreadsheet generation
    despatch_snapshot_rows = []
    open_lrs_rows = []
    
    # Aggregate statistics
    for lr_no, snap in snapshot_lrs.items():
        # Get matching row in downloaded LR report
        row = lr_rows_map.get(lr_no)
        
        consignor = clean_val(row[consignor_col]) if row is not None else snap.get("consignor", "")
        if consignor.upper().startswith('EFF'):
            continue
            
        consignee = clean_val(row[consignee_col]) if row is not None else snap.get("consignee", "")
        destination = clean_val(row[dest_col]) if row is not None else snap.get("destination", "")
        box_qty = int(row[box_col]) if (row is not None and box_col and pd.notna(row[box_col]) and str(row[box_col]).strip().isdigit()) else int(snap.get("box_qty", 0) or 0)
        lr_date = clean_val(row[date_col]) if row is not None else snap.get("despatch_date", yesterday_str)
        del_time = clean_val(row[del_time_col]) if row is not None else ""
        
        raw_status = clean_val(row[status_col], "Open") if row is not None else "Open"
        status_map = {
            'Despatched': 'On transit',
            'Open': 'Not Despatched',
            'Delivered': 'Delivery Process completed.',
            'Despatched from Branch': 'Cancelled LR'
        }
        mapped_status = status_map.get(raw_status, "Not Despatched")
        if "cancelled" in raw_status.lower() or mapped_status == 'Cancelled LR':
            mapped_status = 'Cancelled LR'
            
        lr_date_dt = parse_date(lr_date)
        del_time_dt = parse_date(del_time)
        
        branch = snap.get("branch", "N/A")
        if not branch or pd.isna(branch):
            branch = "N/A"
            
        driver = snap.get("driver_name", "")
        if not driver or pd.isna(driver):
            driver = "N/A"
            
        despatch_no = snap.get("despatch_no", "")
        if not despatch_no or pd.isna(despatch_no):
            despatch_no = "N/A"
            
        supervisor = snap.get("supervisor_name", "")
        if not supervisor or pd.isna(supervisor):
            supervisor = "N/A"
        
        if supervisor and supervisor != "N/A" and branch == "N/A":
            unmapped_supervisors.add(supervisor)
            
        lr_item = {
            "lr_no": lr_no,
            "consignee": consignee,
            "destination": destination,
            "box_qty": box_qty,
            "lr_date": lr_date,
            "delivery_time": del_time,
            "status": mapped_status,
            "driver": driver,
            "despatch_no": despatch_no,
            "despatch_time": snap.get("despatch_date", ""),
            "aging": 0
        }
        
        is_delivered = (mapped_status == 'Delivery Process completed.')
        
        if is_delivered:
            lr_item["aging"] = calculate_aging(lr_date_dt, del_time_dt, holidays)
        else:
            lr_item["aging"] = calculate_aging(lr_date_dt, yesterday_dt, holidays)
            
        # Accumulate Branch statistics
        if branch not in branch_stats:
            branch_stats[branch] = {
                "delivered_lrs": [], "returned_lrs": [], "open_lrs": [], "dispatches": {}
            }
            
        b_group = branch_stats[branch]
        if is_delivered:
            b_group["delivered_lrs"].append(lr_item)
        else:
            b_group["returned_lrs"].append(lr_item)
            
        # Accumulate Dispatch/Driver statistics
        disp_key = (driver, despatch_no)
        if disp_key not in b_group["dispatches"]:
            b_group["dispatches"][disp_key] = {"delivered_lrs": [], "returned_lrs": []}
            
        disp_group = b_group["dispatches"][disp_key]
        if is_delivered:
            disp_group["delivered_lrs"].append(lr_item)
        else:
            disp_group["returned_lrs"].append(lr_item)
            
        # Snapshot sheet row
        despatch_snapshot_rows.append({
            "Branch": branch,
            "Despatch No": despatch_no,
            "Despatch Date": snap.get("despatch_date"),
            "Driver Name": driver,
            "Supervisor Name": supervisor,
            "LR No": lr_no,
            "LR Date": lr_date,
            "Consignee": consignee,
            "Destination": destination,
            "Box Qty": box_qty,
            "Current Status": mapped_status,
            "Delivery Time": del_time if is_delivered else "-",
            "LR Age (Days)": lr_item["aging"]
        })

    # Classify Open LRs as of yesterday 7:00 AM
    for lr_no, r in lr_rows_map.items():
        consignor = clean_val(r[consignor_col])
        if consignor.upper().startswith('EFF'):
            continue
            
        lr_date = clean_val(r[date_col])
        lr_date_dt = parse_date(lr_date)
        if not lr_date_dt or lr_date_dt.date() > yesterday_dt.date():
            continue
            
        raw_status = clean_val(r[status_col], "Open")
        status_map = {
            'Despatched': 'On transit',
            'Open': 'Not Despatched',
            'Delivered': 'Delivery Process completed.',
            'Despatched from Branch': 'Cancelled LR'
        }
        mapped_status = status_map.get(raw_status, "Not Despatched")
        if "cancelled" in raw_status.lower() or mapped_status == 'Cancelled LR':
            continue  # Exclude cancelled LRs
            
        del_time = clean_val(r[del_time_col])
        del_time_dt = parse_date(del_time)
        
        # Determine if undelivered as of yesterday 7:00 AM
        is_open_at_7am = False
        if mapped_status != 'Delivery Process completed.':
            is_open_at_7am = True
        elif del_time_dt and del_time_dt > yesterday_7am:
            is_open_at_7am = True
            
        if is_open_at_7am:
            # Resolve branch using 30-day despatch map or fallback
            branch = "N/A"
            dest_val = clean_val(r[dest_col])
            consignee_val = clean_val(r[consignee_col])
            box_qty_val = int(r[box_col]) if (box_col and pd.notna(r[box_col]) and str(r[box_col]).strip().isdigit()) else 0
            
            # 1. Try yesterday snapshot
            if lr_no in snapshot_lrs:
                branch = snapshot_lrs[lr_no].get("branch", "N/A")
            # 2. Try 30-day despatch report
            elif lr_no in desp_30d_map:
                branch = desp_30d_map[lr_no].get("branch", "N/A")
            # 3. Guess based on area
            else:
                branch = resolve_branch_name(dest_val, "", supervisor_map)
                
            open_age = calculate_aging(lr_date_dt, yesterday_dt, holidays)
            
            open_item = {
                "lr_no": lr_no,
                "consignee": consignee_val,
                "destination": dest_val,
                "box_qty": box_qty_val,
                "lr_date": lr_date,
                "status": mapped_status,
                "aging": open_age
            }
            
            if branch not in branch_stats:
                branch_stats[branch] = {
                    "delivered_lrs": [], "returned_lrs": [], "open_lrs": [], "dispatches": {}
                }
            branch_stats[branch]["open_lrs"].append(open_item)
            
            open_lrs_rows.append({
                "Branch": branch,
                "LR No": lr_no,
                "LR Date": lr_date,
                "Consignor": consignor,
                "Consignee": consignee_val,
                "Destination": dest_val,
                "Box Qty": box_qty_val,
                "Current Status": mapped_status,
                "LR Age (Days)": open_age
            })

    # Calculations for Overall Summary Card
    overall_delivered = []
    overall_returned = []
    overall_open = []
    for b, stats in branch_stats.items():
        overall_delivered.extend(stats["delivered_lrs"])
        overall_returned.extend(stats["returned_lrs"])
        overall_open.extend(stats["open_lrs"])
        
    overall_del_max_age, overall_del_max_age_count, _ = calculate_aging_metrics(overall_delivered, yesterday_dt)
    overall_ret_max_age, overall_ret_max_age_count, _ = calculate_aging_metrics(overall_returned, yesterday_dt)
    overall_open_max_age, overall_open_max_age_count, _ = calculate_aging_metrics(overall_open, yesterday_dt)
    
    overall_stats = {
        "delivered_count": len(overall_delivered),
        "delivered_max_age": overall_del_max_age,
        "delivered_max_age_count": overall_del_max_age_count,
        "returned_count": len(overall_returned),
        "returned_max_age": overall_ret_max_age,
        "returned_max_age_count": overall_ret_max_age_count,
        "open_count": len(overall_open),
        "open_max_age": overall_open_max_age,
        "open_max_age_count": overall_open_max_age_count
    }
    
    # 4. Construct WhatsApp Message text
    date_display = yesterday_dt.strftime("%d/%m/%Y")
    wa_msg = f"DAILY ERP SUMMARY ({date_display}) 📊\n\n"
    wa_msg += f"🟢 DELIVERED YESTERDAY: {overall_stats['delivered_count']} LRs ⏱️ Max Delay: {overall_stats['delivered_max_age']} days ({overall_stats['delivered_max_age_count']} LRs)\n"
    wa_msg += f"🔴 RETURNED YESTERDAY: {overall_stats['returned_count']} LRs ⏳ Max Aging: {overall_stats['returned_max_age']} days ({overall_stats['returned_max_age_count']} LRs)\n"
    wa_msg += f"📦 OPEN/DESPATCH (as of 7 AM): {overall_stats['open_count']} LRs ⏳ Max Aging: {overall_stats['open_max_age']} days ({overall_stats['open_max_age_count']} LRs)\n\n"
    
    # Branch and Driver Wise detail
    branch_summary_excel_rows = []
    despatch_summary_rows = []
    
    for b in sorted(branch_stats.keys()):
        if b == "N/A" and not branch_stats[b]["delivered_lrs"] and not branch_stats[b]["returned_lrs"] and not branch_stats[b]["open_lrs"]:
            continue
            
        b_group = branch_stats[b]
        del_max_age, del_max_age_count, del_pts = calculate_aging_metrics(b_group["delivered_lrs"], yesterday_dt)
        ret_max_age, ret_max_age_count, ret_pts = calculate_aging_metrics(b_group["returned_lrs"], yesterday_dt)
        open_max_age, open_max_age_count, _ = calculate_aging_metrics(b_group["open_lrs"], yesterday_dt)
        
        wa_msg += f"{b} BRANCH 🏢\n"
        wa_msg += f"🟢 Delivered: {len(b_group['delivered_lrs'])} LRs, (max Aging {del_max_age}Day {del_max_age_count}nos),Total Delivery {del_pts} Nos.\n"
        wa_msg += f"🔴 Returned: {len(b_group['returned_lrs'])} LRs,(max Aging {ret_max_age} Day {ret_max_age_count}nos),Total Delivery {ret_pts} Nos.\n"
        
        branch_summary_excel_rows.append({
            "Branch": b,
            "Delivered Count": len(b_group["delivered_lrs"]),
            "Delivered Max Aging (Days)": del_max_age,
            "Delivered Max Aging Count": del_max_age_count,
            "Delivered Unique Points": del_pts,
            "Returned Count": len(b_group["returned_lrs"]),
            "Returned Max Aging (Days)": ret_max_age,
            "Returned Max Aging Count": ret_max_age_count,
            "Returned Unique Points": ret_pts,
            "Open Count": len(b_group["open_lrs"]),
            "Open Max Aging (Days)": open_max_age,
            "Open Max Aging Count": open_max_age_count
        })
        
        # Driver/Despatch details
        drv_idx = 1
        for (driver, desp_no), disp in sorted(b_group["dispatches"].items()):
            all_lrs_in_desp = disp["delivered_lrs"] + disp["returned_lrs"] + disp.get("open_lrs", [])
            total_lrs = len(all_lrs_in_desp)
            total_pts = len(set(f"{x['consignee'].lower()}||{x['destination'].lower()}" for x in all_lrs_in_desp))
            
            # Parse delivery times (from delivered only)
            del_times_dt = [parse_date(x["delivery_time"]) for x in disp["delivered_lrs"] if parse_date(x["delivery_time"])]
            if del_times_dt:
                min_time = min(del_times_dt).strftime("%d/%m %I:%M %p")
                max_time = max(del_times_dt).strftime("%d/%m %I:%M %p")
            else:
                min_time = "-"
                max_time = "-"
                
            delivered_count = len(disp["delivered_lrs"])
            returned_count = len(disp["returned_lrs"])
            open_count = len(disp.get("open_lrs", []))
                
            wa_msg += f"{drv_idx}) Driver: {driver if driver else 'N/A'} | Despatch: {desp_no}\n"
            wa_msg += f"   📦 Total LRs: {total_lrs} | Points: {total_pts}\n"
            wa_msg += f"   🟢 Delivered: {delivered_count} | 🔴 Despatched (Returned): {returned_count}"
            if open_count > 0:
                wa_msg += f" | ⏳ Open: {open_count}"
            wa_msg += "\n"
            wa_msg += f"   ⏱️ 1st Delivery: {min_time} | Last Delivery: {max_time}\n"
            drv_idx += 1
            
            desp_time = all_lrs_in_desp[0].get("despatch_time", "") if all_lrs_in_desp else ""
            
            despatch_summary_rows.append({
                "Branch": b,
                "Driver Name": driver,
                "Despatch No": desp_no,
                "Despatch Time": desp_time,
                "Total LRs": total_lrs,
                "Total Delivery Points": total_pts,
                "Delivered Count": delivered_count,
                "Despatched (Returned) Count": returned_count,
                "1st Delivery Time": min_time,
                "Last Delivery Time": max_time
            })
            
        wa_msg += "\n"
        
    print("Generated WhatsApp Message:\n", wa_msg)
    
    # Send WhatsApp Business Message
    send_whatsapp_message(wa_msg)
    
    # 5. Generate Excel file
    processed_file = os.path.join(DOWNLOAD_DIR, "Interactive_Delivery_Report.xlsx")
    writer = pd.ExcelWriter(processed_file, engine='openpyxl')
    
    # Create sheets
    # Sheet 1: Daily Summary
    df_branch_summary = pd.DataFrame(branch_summary_excel_rows)
    df_branch_summary.to_excel(writer, sheet_name="1. Daily Summary", index=False)
    
    # Sheet 2: Despatch Snapshot
    df_despatch_snap = pd.DataFrame(despatch_snapshot_rows)
    if df_despatch_snap.empty:
        df_despatch_snap = pd.DataFrame(columns=["Branch", "Despatch No", "Despatch Date", "Driver Name", "Supervisor Name", "LR No", "LR Date", "Consignee", "Destination", "Box Qty", "Current Status", "Delivery Time", "LR Age (Days)"])
    df_despatch_snap.to_excel(writer, sheet_name="2. Despatch Snapshot", index=False)
    
    # Sheet 3: Open LRs
    df_open_lrs = pd.DataFrame(open_lrs_rows)
    if df_open_lrs.empty:
        df_open_lrs = pd.DataFrame(columns=["Branch", "LR No", "LR Date", "Consignor", "Consignee", "Destination", "Box Qty", "Current Status", "LR Age (Days)"])
    df_open_lrs.to_excel(writer, sheet_name="3. Open LRs", index=False)
    
    # Sheet 4: Despatch Summary
    df_despatch_summary = pd.DataFrame(despatch_summary_rows)
    if df_despatch_summary.empty:
        df_despatch_summary = pd.DataFrame(columns=["Branch", "Driver Name", "Despatch No", "Despatch Time", "Total LRs", "Total Delivery Points", "Delivered Count", "Despatched (Returned) Count", "1st Delivery Time", "Last Delivery Time"])
    df_despatch_summary.to_excel(writer, sheet_name="4. Despatch Summary", index=False)
    
    writer.close()
    
    # Format and Style Workbook
    wb = openpyxl.load_workbook(processed_file)
    apply_styles(wb["1. Daily Summary"], len(df_branch_summary) + 1, len(df_branch_summary.columns), enable_filter=True)
    apply_styles(wb["2. Despatch Snapshot"], len(df_despatch_snap) + 1, len(df_despatch_snap.columns), enable_filter=True)
    apply_styles(wb["3. Open LRs"], len(df_open_lrs) + 1, len(df_open_lrs.columns), enable_filter=True)
    apply_styles(wb["4. Despatch Summary"], len(df_despatch_summary) + 1, len(df_despatch_summary.columns), enable_filter=True)
    wb.save(processed_file)
    
    dashboard_image_path = generate_pillow_dashboard(overall_stats, date_display)
    delay_tables_html = generate_delay_tables_html(despatch_snapshot_rows)
    return processed_file, dashboard_image_path, unmapped_supervisors, delay_tables_html

def generate_delay_tables_html(despatch_snapshot_rows):
    delivered_buckets = {}
    open_buckets = {}
    
    for r in despatch_snapshot_rows:
        branch = r.get("Branch", "N/A")
        if not branch or pd.isna(branch):
            branch = "N/A"
        branch_str = str(branch).strip()
        if branch_str.startswith("LH-"):
            parts = branch_str[3:].split(" to ")
            if parts:
                branch_str = parts[0].strip()
        branch_str = branch_str.upper()
        if not branch_str:
            branch_str = "N/A"
            
        status = r.get("Current Status", "Open")
        is_delivered = (status == "Delivery Process completed.")
        
        age = r.get("LR Age (Days)", 0)
        try:
            age = int(float(age))
        except (ValueError, TypeError):
            age = 0
        if age < 0:
            age = 0
            
        bucket_key = age if age <= 15 else 16
        
        if is_delivered:
            if branch_str not in delivered_buckets:
                delivered_buckets[branch_str] = {}
            delivered_buckets[branch_str][bucket_key] = delivered_buckets[branch_str].get(bucket_key, 0) + 1
        else:
            if age >= 1:
                if branch_str not in open_buckets:
                    open_buckets[branch_str] = {}
                open_buckets[branch_str][bucket_key] = open_buckets[branch_str].get(bucket_key, 0) + 1

    headers_delivered = ["Branch", "SAME DAY", "NEXT DAY", "2nd Day", "3rd Day", "4th Day", "5th Day", "6th Day", "7th Day", "8th Day", "9th Day", "10th Day", "11th Day", "12th Day", "13th Day", "14th Day", "15th Day", "16th+ Day"]
    headers_open = ["Branch", "1st Day", "2nd Day", "3rd Day", "4th Day", "5th Day", "6th Day", "7th Day", "8th Day", "9th Day", "10th Day", "11th Day", "12th Day", "13th Day", "14th Day", "15th Day", "16th+ Day"]
    
    def build_html_table(title, headers, data_dict, is_open=False):
        sorted_branches = sorted(data_dict.keys())
        
        html = f"""
        <h3 style="color: #4C1D95; font-family: Arial, sans-serif; margin-top: 25px; margin-bottom: 10px; font-size: 16px;">{title}</h3>
        <table style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; font-size: 11px; margin-bottom: 20px; border: 1px solid #E2E8F0;">
          <thead>
            <tr style="background-color: #522A7A; color: #FFFFFF;">
              <th style="padding: 8px 10px; border: 1px solid #E2E8F0; text-align: left; font-weight: bold;">Branch</th>
        """
        for h in headers[1:]:
            html += f'<th style="padding: 8px 10px; border: 1px solid #E2E8F0; text-align: center; font-weight: bold;">{h}</th>'
        html += "</tr></thead><tbody>"
        
        col_totals = {h: 0 for h in headers[1:]}
        
        for idx, br in enumerate(sorted_branches):
            bg_color = "#F8FAFC" if idx % 2 == 0 else "#FFFFFF"
            html += f'<tr style="background-color: {bg_color};">'
            html += f'<td style="padding: 8px 10px; border: 1px solid #E2E8F0; font-weight: bold; color: #0F172A;">{br}</td>'
            
            for h in headers[1:]:
                if h == "SAME DAY":
                    val = data_dict[br].get(0, 0)
                elif h == "NEXT DAY":
                    val = data_dict[br].get(1, 0)
                elif h == "16th+ Day":
                    val = data_dict[br].get(16, 0)
                else:
                    day_num = int(h.split()[0][:-2])
                    val = data_dict[br].get(day_num, 0)
                    
                val_str = str(val) if val > 0 else ""
                html += f'<td style="padding: 8px 10px; border: 1px solid #E2E8F0; text-align: center; font-weight: {"bold" if val > 0 else "normal"}; color: #0F172A;">{val_str}</td>'
                col_totals[h] += val
                
            html += "</tr>"
            
        html += '<tr style="background-color: #CBD5E1; font-weight: bold; color: #0F172A;">'
        html += '<td style="padding: 8px 10px; border: 1px solid #E2E8F0; text-align: left;">TOTAL</td>'
        for h in headers[1:]:
            val = col_totals[h]
            val_str = str(val) if val > 0 else ""
            html += f'<td style="padding: 8px 10px; border: 1px solid #E2E8F0; text-align: center;">{val_str}</td>'
        html += "</tr></tbody></table>"
        return html

    delivered_html = build_html_table("Branch-wise Delivery Delay Analysis (Buckets) 📦", headers_delivered, delivered_buckets) if delivered_buckets else ""
    open_html = build_html_table("Branch-wise Open LR Aging Analysis (1st Day onwards) 🕒", headers_open, open_buckets, is_open=True) if open_buckets else ""
    
    return delivered_html + open_html

def check_if_delayed(mode="daily_evening_report"):
    """
    Check if this execution is delayed.
    It returns True if:
    1. The current IST time is past 8:30 PM (for daily_evening_report) or past 6:30 AM (for morning).
    2. Or there are failed runs in the GitHub workflow history for today.
    """
    import subprocess
    import json
    
    ist_tz = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(ist_tz)
    
    # Time-based delay detection
    if mode == "daily_evening_report":
        if now_ist.hour > 20 or (now_ist.hour == 20 and now_ist.minute >= 30):
            return True
    elif mode == "morning":
        if now_ist.hour > 6 or (now_ist.hour == 6 and now_ist.minute >= 30):
            return True
            
    # Check GitHub run history for any failures today
    gh_token = os.getenv("GH_TOKEN") or os.getenv("GITHUB_TOKEN")
    if not gh_token:
        # If we are not in GitHub Actions or have no token, rely on time-based only
        return False
        
    try:
        cmd = ["gh", "run", "list", "--workflow", "daily_report.yml", "--json", "conclusion,createdAt,status,databaseId"]
        result = subprocess.run(cmd, capture_output=True, text=True, check=True)
        runs = json.loads(result.stdout)
        
        current_run_id = os.getenv("GITHUB_RUN_ID")
        if current_run_id:
            try:
                current_run_id = int(current_run_id)
            except ValueError:
                current_run_id = None
            
        today_str = now_ist.strftime("%Y-%m-%d")
        for run in runs:
            # Skip current running run
            if current_run_id and run.get("databaseId") == current_run_id:
                continue
                
            if run.get("status") == "completed" and run.get("conclusion") == "failure":
                created_at_utc_str = run.get("createdAt")
                if created_at_utc_str.endswith("Z"):
                    created_at_utc_str = created_at_utc_str[:-1] + "+00:00"
                created_at_utc = datetime.fromisoformat(created_at_utc_str)
                created_at_ist = created_at_utc.astimezone(ist_tz)
                if created_at_ist.strftime("%Y-%m-%d") == today_str:
                    return True
    except Exception as e:
        print(f"Error checking GitHub runs for failure detection: {e}")
        
    return False

# Email function
def email_report(processed_file_path, raw_lr_path, raw_despatch_path, dashboard_image_path, from_date=None, to_date=None, unmapped_supervisors=None, delay_tables_html="", is_delayed=False):
    print("Sending daily report email...")
    if not SENDER_EMAIL or not SENDER_PASSWORD or not RECEIVER_EMAIL:
        print("⚠️ Email credentials (SENDER_EMAIL, SENDER_PASSWORD, RECEIVER_EMAIL) are missing. Skipping email sending.")
        return
    
    # Format the dates for display in the email
    today_str = datetime.now().strftime("%d-%m-%Y")
    msg = MIMEMultipart('related')
    msg['From'] = SENDER_EMAIL
    msg['To'] = RECEIVER_EMAIL
    msg['Subject'] = f"Daily ERP Dispatch & Delivery Performance Report - {today_str} 📊"
    
    msg_alt = MIMEMultipart('alternative')
    msg.attach(msg_alt)
    
    unmapped_str = ""
    if unmapped_supervisors:
        unmapped_str = "<br>⚠️ <b>Unmapped Supervisors found today</b> (Please add them to the Supervisor Mapping table in the dashboard to assign them to branches):<br>"
        for s in sorted(unmapped_supervisors):
            unmapped_str += f"- {s}<br>"
            
    delayed_banner = ""
    if is_delayed:
        delayed_banner = """
        <div style="background-color: #FEF3C7; border-left: 4px solid #D97706; padding: 10px 15px; margin-bottom: 20px; font-family: Arial, sans-serif; font-size: 12px; color: #92400E; border-radius: 4px;">
          <strong>⚠️ Delayed Report Notice:</strong> This report was delayed due to temporary connection or infrastructure issues with the ERP server/runner earlier today. It has been successfully processed and sent now.
        </div>
        """
            
    html_body = f"""
    <html>
      <body style="font-family: Arial, sans-serif; color: #1E293B;">
        {delayed_banner}
        <p>Dear User,</p>
        <p>Please find the daily ERP Dispatch & Delivery performance dashboard summary for <b>{today_str}</b> below:</p>
        <img src="cid:dashboard_image"><br>
        <p>Report Period: <b>{from_date}</b> to <b>{to_date}</b> 📅</p>
        {delay_tables_html}
        {unmapped_str}
        <p><b>Included sheets in Interactive_Delivery_Report.xlsx:</b><br>
        1. Daily Summary (Overall & Branch summaries)<br>
        2. Despatch Snapshot (Yesterday's LRs and their final morning status)<br>
        3. Open LRs (List of undelivered LRs as of yesterday morning 7:00 AM)<br>
        4. Despatch Summary (Driver-wise delivery points, times, and returns)</p>
        <p>Also attached are the raw downloaded ERP reports for your reference.</p>
        <p>Best Regards,<br>
        <b>ERP Daily Automation Engine</b> ⚡</p>
      </body>
    </html>
    """
    msg_alt.attach(MIMEText(html_body, 'html'))
    
    # Inline image attachment
    with open(dashboard_image_path, 'rb') as f:
        img_part = MIMEBase('image', 'png')
        img_part.set_payload(f.read())
        encode_base64(img_part)
        img_part.add_header('Content-ID', '<dashboard_image>')
        img_part.add_header('Content-Disposition', 'inline', filename='daily_dashboard.png')
        msg.attach(img_part)
        
    # File attachments
    def get_extension(file_path):
        if not os.path.exists(file_path):
            return ".xlsx"
        try:
            with open(file_path, "rb") as f:
                head = f.read(4)
            if head == b"PK\x03\x04":
                return ".xlsx"
            elif head == b"\xd0\xcf\x11\xe0":
                return ".xls"
            else:
                return ".csv"
        except Exception:
            return ".xlsx"

    lr_ext = get_extension(raw_lr_path) if raw_lr_path else ".xlsx"
    desp_ext = get_extension(raw_despatch_path)
    
    files_to_attach = [
        (processed_file_path, f"Interactive_Delivery_Report_{today_str}.xlsx"),
        (raw_despatch_path, f"raw_despatch_data_{today_str}{desp_ext}")
    ]
    if raw_lr_path:
        files_to_attach.append((raw_lr_path, f"raw_lr_data_{today_str}{lr_ext}"))
        
    for file_path, attachment_name in files_to_attach:
        if file_path and os.path.exists(file_path):
            with open(file_path, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
                encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename= {attachment_name}",
                )
                msg.attach(part)
                print(f"Attached: {attachment_name}")
                
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        recipient_list = [r.strip() for r in RECEIVER_EMAIL.split(",") if r.strip()]
        server.send_message(msg, to_addrs=recipient_list)
        server.quit()
        print("🎉 Daily report email sent successfully!")
    except Exception as e:
        print("❌ Error sending email:", e)
        raise e

# 9. Main orchestrator
def main():
    parser = argparse.ArgumentParser(description="ERP Dispatch & Delivery Performance Report")
    parser.add_argument("--mode", choices=["evening", "morning", "daily_evening_report", "afternoon_open_lrs", "reconcile"], required=True, help="Run mode")
    parser.add_argument("--from-date", help="Override from date (YYYY-MM-DD)")
    parser.add_argument("--to-date", help="Override to date (YYYY-MM-DD)")
    parser.add_argument("--from-time", help="Override from time (HH:MM:SS)")
    parser.add_argument("--to-time", help="Override to time (HH:MM:SS)")
    args = parser.parse_args()
    
    print(f"[{datetime.now()}] Starting daily report automation runner in mode: {args.mode}")

    # Calculate target date for checking Sunday/holiday skipping
    ist_tz = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(ist_tz)
    if args.mode in ("daily_evening_report", "evening", "afternoon_open_lrs"):
        if now_ist.hour < 12:
            target_date = now_ist - timedelta(days=1)
        else:
            target_date = now_ist
    elif args.mode == "morning":
        target_date = now_ist - timedelta(days=1)
    else:
        target_date = now_ist

    # Fetch holidays list
    holidays_list = fetch_holidays()

    # Auto-skip Sunday & Holidays if no date overrides are provided
    if not args.from_date and not args.to_date:
        target_date_str = target_date.strftime("%Y-%m-%d")
        if target_date.weekday() == 6:
            print(f"Target date {target_date_str} is Sunday. Skipping run gracefully.")
            sys.exit(0)
        if target_date_str in holidays_list:
            print(f"Target date {target_date_str} is a marked holiday ({target_date_str}). Skipping run gracefully.")
            sys.exit(0)

    # Fetch mappings
    supervisor_map = fetch_supervisor_mappings()
    
    if args.mode == "evening":
        # Download despatch report
        lr_file, despatch_file, from_date, to_date = download_erp_reports(mode="evening", from_override=args.from_date, to_override=args.to_date)
        run_evening_flow(despatch_file, supervisor_map)
        print("Evening flow execution completed successfully.")
        
    elif args.mode == "daily_evening_report":
        # Download both reports
        lr_file, despatch_file, from_date, to_date = download_erp_reports(mode="daily_evening_report", from_override=args.from_date, to_override=args.to_date)
        
        ist_tz = timezone(timedelta(hours=5, minutes=30))
        now_ist = datetime.now(ist_tz)
        
        # If running before 12:00 PM (noon) IST, treat it as a delayed run for the previous calendar day.
        if now_ist.hour < 12:
            target_date = now_ist - timedelta(days=1)
        else:
            target_date = now_ist
            
        today_str = target_date.strftime("%Y-%m-%d")
        # On Monday evenings (or when the target date is Monday), default starting date is Saturday (2 days ago)
        # Resolve yesterday_str by looking back to find the last working day (non-Sunday, non-holiday)
        lookback_date = target_date - timedelta(days=1)
        while lookback_date.weekday() == 6 or lookback_date.strftime("%Y-%m-%d") in holidays_list:
            lookback_date -= timedelta(days=1)
        yesterday_str = lookback_date.strftime("%Y-%m-%d")
        
        if args.from_date:
            yesterday_str = args.from_date
        if args.to_date:
            today_str = args.to_date
            
        start_time_override = getattr(args, 'from_time', None)
        end_time_override = getattr(args, 'to_time', None)
            
        processed_file, dashboard_image_path, unmapped_supervisors, delay_tables_html = run_daily_evening_report_flow(
            lr_file, despatch_file, supervisor_map, yesterday_str, today_str, start_time_override, end_time_override
        )
        
        # Email report
        is_delayed = check_if_delayed("daily_evening_report")
        email_report(processed_file, lr_file, despatch_file, dashboard_image_path, from_date, to_date, unmapped_supervisors, delay_tables_html, is_delayed=is_delayed)
        print("Daily Evening report flow execution completed successfully.")
        
    elif args.mode == "morning":
        # Download both reports
        lr_file, despatch_file, from_date, to_date = download_erp_reports(mode="morning", from_override=args.from_date, to_override=args.to_date)
        
        # Yesterday's date in IST
        ist_tz = timezone(timedelta(hours=5, minutes=30))
        yesterday_str = (datetime.now(ist_tz) - timedelta(days=1)).strftime("%Y-%m-%d")
        if args.from_date:
            yesterday_str = args.from_date  # Use overridden date as yesterday's date
            
        processed_file, dashboard_image_path, unmapped_supervisors, delay_tables_html = run_morning_flow(lr_file, despatch_file, supervisor_map, yesterday_str)
        
        # Email report
        is_delayed = check_if_delayed("morning")
        email_report(processed_file, lr_file, despatch_file, dashboard_image_path, from_date, to_date, unmapped_supervisors, delay_tables_html, is_delayed=is_delayed)
        print("Morning flow execution completed successfully.")

    elif args.mode == "afternoon_open_lrs":
        # We need to import the afternoon flow script
        from run_afternoon_open_lrs import run_afternoon_open_lrs_flow
        
        # Download ONLY the LR reports for 30 days
        # We use from_override to go back 30 days
        ist_tz = timezone(timedelta(hours=5, minutes=30))
        now_ist = datetime.now(ist_tz)
        from_30_days = (now_ist - timedelta(days=30)).strftime("%Y-%m-%d")
        
        lr_file, despatch_file, from_date, to_date = download_erp_reports(
            mode="afternoon_open_lrs", 
            from_override=args.from_date if args.from_date else from_30_days, 
            to_override=args.to_date
        )
        
        if lr_file:
            run_afternoon_open_lrs_flow(lr_file)
            print("Afternoon Open LRs flow execution completed successfully.")
        else:
            print("Failed to download LR file. Aborting.")
            
    elif args.mode == "reconcile":
        # Download reports needed for reconciliation (raw LR, raw Despatch, bill clear reports, and GDM details)
        lr_file, despatch_file, from_date, to_date = download_erp_reports(mode="reconcile", from_override=args.from_date, to_override=args.to_date)
        print("Reconciliation downloads completed successfully. Commencing discrepancy analysis...", flush=True)
        
        # Invoke freight calculation engine
        import sys
        sys.path.append("/Users/anwar/Antigravity-Related/EFF PARCEL FREIGHT WORKING")
        from freight_calculator import process_freight_data
        
        # Target Google Sheet Title
        sheet_title = "Topay & Paid Parcel Billing"
        creds_path = "/Users/anwar/Antigravity-Related/ERP nxt Data collection/Invoice_Extractor_Tool/credentials.json"
        
        # Load local excel or fetch from Google Sheet dynamically
        rates_excel = "/Users/anwar/Antigravity-Related/EFF PARCEL FREIGHT WORKING/All Consignors - RATES Combined.xlsx"
        
        try:
            print("Running freight calculation engine...", flush=True)
            result_df, summary_stats, df_whole = process_freight_data(lr_file, rates_excel)
            print(f"Calculations complete: {summary_stats}", flush=True)
            
            # Extract records older than 52 days for archiving (only on automated daily runs without date overrides)
            if not args.from_date:
                result_df['parsed_dt'] = pd.to_datetime(result_df['DATE'], format='%d/%m/%Y', errors='coerce')
                cutoff_date = pd.Timestamp.now() - pd.Timedelta(days=52)
                df_archive = result_df[pd.notna(result_df['parsed_dt']) & (result_df['parsed_dt'] < cutoff_date)].copy().drop(columns=['parsed_dt'])
                result_df = result_df[result_df['parsed_dt'].isna() | (result_df['parsed_dt'] >= cutoff_date)].copy().drop(columns=['parsed_dt'])

            # Filter payment buckets before formatting floats to string
            ptype_col = 'PAYMENT TYPE' if 'PAYMENT TYPE' in result_df.columns else ''
            topay_mask = pd.to_numeric(result_df['To Pay'], errors='coerce') > 0
            paid_mask = pd.to_numeric(result_df['Paid'], errors='coerce') > 0
            if ptype_col:
                topay_mask = topay_mask | result_df[ptype_col].astype(str).str.upper().str.contains('TO PAY|TOPAY')
                paid_mask = paid_mask | result_df[ptype_col].astype(str).str.upper().str.contains('PAID')

            df_topay = result_df[topay_mask].copy()
            df_paid = result_df[paid_mask].copy()

            # Format dataframe for google sheets transfer (replace nan values, format headers)
            result_df = result_df.fillna("")
            num_cols = ['WEIGHT', 'Existing ERP Total', 'Calculated Total Freight', 'Account Pay', 'To Pay', 'Paid', 
                        'Calculated Stationary Charge', 'Calculated Unloading Charge', 'Grand Total with UL', 'Amount Difference']
            for col in result_df.columns:
                if col in num_cols:
                    result_df[col] = pd.to_numeric(result_df[col], errors='coerce').apply(lambda x: f"{x:.2f}" if pd.notna(x) else "")
                else:
                    result_df[col] = result_df[col].astype(str)

            for df_sub in [df_topay, df_paid]:
                if not df_sub.empty:
                    df_sub.fillna("", inplace=True)
                    for col in df_sub.columns:
                        if col in num_cols:
                            df_sub[col] = pd.to_numeric(df_sub[col], errors='coerce').apply(lambda x: f"{x:.2f}" if pd.notna(x) else "")
                        else:
                            df_sub[col] = df_sub[col].astype(str)

            def robust_read_df(filepath):
                if not filepath or not os.path.exists(filepath):
                    return pd.DataFrame()
                try:
                    return pd.read_excel(filepath)
                except Exception:
                    try:
                        dfs = pd.read_html(filepath)
                        if dfs: return dfs[0]
                    except Exception:
                        try: return pd.read_csv(filepath)
                        except Exception: pass
                return pd.DataFrame()

            # Prepare raw datasets for Despatch Data and LR Data tabs
            df_despatch_raw = robust_read_df(despatch_file).fillna("")
            if not df_despatch_raw.empty and not args.from_date:
                for col in df_despatch_raw.columns:
                    if 'DATE' in str(col).upper():
                        df_despatch_raw['parsed_dt'] = pd.to_datetime(df_despatch_raw[col], errors='coerce')
                        df_despatch_raw = df_despatch_raw[df_despatch_raw['parsed_dt'].isna() | (df_despatch_raw['parsed_dt'] >= cutoff_date)].drop(columns=['parsed_dt'])
                        break
            if not df_despatch_raw.empty:
                for col in df_despatch_raw.columns:
                    df_despatch_raw[col] = df_despatch_raw[col].astype(str)

            df_lr_raw = robust_read_df(lr_file).fillna("")
            if not df_lr_raw.empty and not args.from_date:
                for col in df_lr_raw.columns:
                    if 'DATE' in str(col).upper():
                        df_lr_raw['parsed_dt'] = pd.to_datetime(df_lr_raw[col], errors='coerce')
                        df_lr_raw = df_lr_raw[df_lr_raw['parsed_dt'].isna() | (df_lr_raw['parsed_dt'] >= cutoff_date)].drop(columns=['parsed_dt'])
                        break
            if not df_lr_raw.empty:
                for col in df_lr_raw.columns:
                    df_lr_raw[col] = df_lr_raw[col].astype(str)

            target_tabs = [
                ("Reconciled Audit", result_df),
                ("All Data", result_df),
                ("Despatch Data", df_despatch_raw),
                ("LR Data", df_lr_raw),
                ("Topay", df_topay),
                ("Paid", df_paid)
            ]

            scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
            creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
            client = gspread.authorize(creds)
            sh = client.open(sheet_title)

            for tab_name, dataset_df in target_tabs:
                if dataset_df.empty and tab_name in ("Topay", "Paid"):
                    print(f"Skipping empty dataset for tab '{tab_name}'...", flush=True)
                    continue

                data_to_sync = [dataset_df.columns.tolist()] + dataset_df.values.tolist()
                print(f"Syncing {len(data_to_sync)-1} rows to tab '{tab_name}' in '{sheet_title}'...", flush=True)

                try:
                    ws = sh.worksheet(tab_name)
                    ws.clear()
                    if ws.row_count < len(data_to_sync) + 100:
                        ws.resize(rows=len(data_to_sync) + 100, cols=max(30, len(dataset_df.columns)))
                except gspread.exceptions.WorksheetNotFound:
                    ws = sh.add_worksheet(title=tab_name, rows=str(max(1000, len(data_to_sync)+100)), cols=str(max(30, len(dataset_df.columns))))

                chunk_size = 5000
                for i in range(0, len(data_to_sync), chunk_size):
                    chunk = data_to_sync[i:i+chunk_size]
                    start_row = i + 1
                    ws.update(range_name=f"A{start_row}", values=chunk)
                print(f"  ✅ Tab '{tab_name}' synced ({len(data_to_sync)-1} rows).", flush=True)

            print("🎉 Success! Reconciled Audit, All Data, Topay, and Paid tabs synced to Google Sheets.", flush=True)

            # Ensure master rate sheet tabs remain hidden
            try:
                rate_sheet_names = pd.ExcelFile(rates_excel).sheet_names
                for sname in rate_sheet_names:
                    try:
                        rate_ws = sh.worksheet(sname)
                        sh.batch_update({'requests': [{'updateSheetProperties': {'properties': {'sheetId': rate_ws.id, 'hidden': True}, 'fields': 'hidden'}}]})
                    except Exception:
                        pass
                print("🔒 Master rate sheets visibility updated (hidden).", flush=True)
            except Exception as hide_err:
                print(f"Note: Could not hide rate tabs: {hide_err}", flush=True)
            
            # If archive records exist, append them safely to 'Parcel Billing - Archive 2026'
            if 'df_archive' in locals() and not df_archive.empty:
                archive_title = "Parcel Billing - Archive 2026"
                try:
                    try:
                        ash = client.open(archive_title)
                    except gspread.exceptions.SpreadsheetNotFound:
                        ash = client.create(archive_title)
                    try:
                        aws = ash.worksheet("Archive")
                    except gspread.exceptions.WorksheetNotFound:
                        aws = ash.sheet1
                        aws.update_title("Archive")
                        aws.update(range_name="A1", values=[df_archive.columns.tolist()])
                    archive_data = df_archive.fillna("").values.tolist()
                    aws.append_rows(archive_data, value_input_option='USER_ENTERED')
                    print(f"📦 Archived {len(archive_data)} historical rows to '{archive_title}'.", flush=True)
                except Exception as arch_err:
                    print(f"Note: Archiving error: {arch_err}", flush=True)
            
        except Exception as audit_err:
            print(f"❌ Error during reconciliation/sync flow: {audit_err}", flush=True)
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    main()
