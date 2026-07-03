import pandas as pd
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from playwright.sync_api import sync_playwright

from download_and_email import (
    fetch_supervisor_mappings, fetch_holidays, clean_val, parse_date, SENDER_EMAIL, SENDER_PASSWORD,
    RECEIVER_EMAIL, ERP_USERNAME, ERP_PASSWORD, load_df
)

def encode_base64(part):
    encoders.encode_base64(part)

def apply_styles(ws, max_row, max_col, enable_filter=True):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = alignment
        
        column_letter = cell.column_letter
        max_length = 0
        for r in range(1, max_row + 1):
            c = ws.cell(row=r, column=col)
            c.border = border
            c.alignment = alignment
            try:
                if len(str(c.value)) > max_length:
                    max_length = len(str(c.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
    if enable_filter and max_row > 1:
        ws.auto_filter.ref = f"A1:{ws.cell(row=max_row, column=max_col).column_letter}{max_row}"

def verify_open_lrs_in_erp(open_lrs):
    print(f"Double checking {len(open_lrs)} Open LRs on ERP UI...")
    verified_open_lrs = []
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()
        
        try:
            # Login
            page.goto("https://eff.aadhocc.in/eff_2021/login")
            page.wait_for_load_state("load")
            if page.locator("#login_user_id").count() > 0 or "login" in page.url.lower():
                if page.locator("#login_user_id").count() > 0:
                    page.fill("#login_user_id", ERP_USERNAME)
                else:
                    page.fill("input[type='text']", ERP_USERNAME)
                if page.locator("#login_password").count() > 0:
                    page.fill("#login_password", ERP_PASSWORD)
                else:
                    page.fill("input[type='password']", ERP_PASSWORD)
                page.click("button[type='submit'], input[type='submit'], button:has-text('Login')")
                page.wait_for_load_state("networkidle", timeout=15000)
                
            # Go to LR list
            page.goto("https://eff.aadhocc.in/eff_2021/list_lr")
            page.wait_for_load_state("networkidle")
            
            # Change pagination to 100 to increase chances of finding it quickly
            try:
                page.select_option("select[name='example_length']", "100")
                page.wait_for_timeout(1000)
            except:
                pass
                
            # Search each LR
            search_input = page.locator("input[type='search']")
            
            for lr in open_lrs:
                lr_no = lr["LR No"]
                search_input.fill("")
                search_input.fill(lr_no)
                page.wait_for_timeout(1500) # Wait for table to filter
                
                # Get the row
                row = page.locator(f"tr:has-text('{lr_no}')")
                if row.count() > 0:
                    html_content = row.first.inner_html().lower()
                    if "cancelled" in html_content or "despatched from branch" in html_content or "delivered" in html_content or "delivery process completed" in html_content or "on transit" in html_content:
                        print(f"LR {lr_no} is actually CANCELLED/DESPATCHED/DELIVERED on UI. Skipping.")
                        continue
                
                print(f"LR {lr_no} is verified as OPEN.")
                verified_open_lrs.append(lr)
                
        except Exception as e:
            print("Error verifying LRs on UI:", e)
            # If error, return all of them to be safe
            return open_lrs
        finally:
            browser.close()
            
    return verified_open_lrs


def calculate_delay(start_dt, end_dt, holidays):
    if not start_dt or not end_dt: return 0
    from datetime import timedelta
    start_dt = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    end_dt = end_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    delay = (end_dt - start_dt).days
    curr = start_dt
    while curr < end_dt:
        d_str = curr.strftime("%Y-%m-%d")
        d_str2 = curr.strftime("%d/%m/%Y")
        if curr.weekday() == 6 or d_str in holidays or d_str2 in holidays:
            delay -= 1
        curr += timedelta(days=1)
    return max(0, delay)

def run_afternoon_open_lrs_flow(lr_file_path):
    print("Running Afternoon Open LRs Flow...")
    holidays = fetch_holidays()
    
    # Read LR Data
    try:
        df_lr = load_df(lr_file_path)
        lr_records = df_lr.to_dict('records')
    except Exception as e:
        print(f"Failed to read LR file: {e}")
        return None
        
    lr_no_col = None
    status_col = None
    date_col = None
    consignor_col = None
    consignee_col = None
    dest_col = None
    qty_col = None
    
    if lr_records:
        for k in lr_records[0].keys():
            k_lower = str(k).strip().lower()
            if "lr no" in k_lower or "lrno" in k_lower: lr_no_col = k
            elif "status" in k_lower: status_col = k
            elif "lr date" in k_lower or "date" in k_lower: date_col = k
            elif "consignor" in k_lower: consignor_col = k
            elif "consignee" in k_lower: consignee_col = k
            elif "destination" in k_lower: dest_col = k
            elif "qty" in k_lower or "box" in k_lower: qty_col = k
            
    if not lr_no_col or not status_col:
        print("Could not find LR No or Status column.")
        return None
        
    initial_open_lrs = []
    
    # Find Open LRs
    for r in lr_records:
        status = clean_val(r.get(status_col, ""))
        if status not in ["delivered", "cancelled", "despatched from branch", "despatch", "delivery process completed.", "delivery process completed", "on transit"] and status != "":
            lr_no = clean_val(r.get(lr_no_col, ""))
            if not lr_no: continue
            
            lr_date_val = clean_val(r.get(date_col, ""))
            dt_obj = parse_date(lr_date_val)
            age = calculate_delay(dt_obj, datetime.now(), holidays) if dt_obj else 0
            
            initial_open_lrs.append({
                "LR No": lr_no,
                "LR Date": lr_date_val,
                "Consignor": clean_val(r.get(consignor_col, "")),
                "Consignee": clean_val(r.get(consignee_col, "")),
                "Destination": clean_val(r.get(dest_col, "")),
                "Box Qty": clean_val(r.get(qty_col, "")),
                "Current Status": status,
                "LR Age (Days)": age
            })
            
    print(f"Found {len(initial_open_lrs)} Open LRs from Excel. Verifying in UI...")
    
    verified_open_lrs = verify_open_lrs_in_erp(initial_open_lrs)
    
    # Generate Excel Report
    today_str = datetime.now().strftime("%d-%m-%Y")
    report_file = os.path.join(os.path.dirname(lr_file_path), f"Afternoon_Open_LRs_Report_{today_str}.xlsx")
    
    # Group by Branch (Destination)
    branch_stats = {}
    for lr in verified_open_lrs:
        b = lr["Destination"] if lr["Destination"] else "N/A"
        if b not in branch_stats:
            branch_stats[b] = []
        branch_stats[b].append(lr)
        
    branch_summary = []
    for b, lrs in branch_stats.items():
        max_age = max([x["LR Age (Days)"] for x in lrs]) if lrs else 0
        branch_summary.append({
            "Branch": b,
            "Total Open LRs": len(lrs),
            "Max Age (Days)": max_age
        })
        
    branch_summary.sort(key=lambda x: x["Total Open LRs"], reverse=True)
    
    # Create Excel
    writer = pd.ExcelWriter(report_file, engine='openpyxl')
    
    df_summary = pd.DataFrame(branch_summary)
    if df_summary.empty:
        df_summary = pd.DataFrame(columns=["Branch", "Total Open LRs", "Max Age (Days)"])
    df_summary.to_excel(writer, sheet_name="1. Branch Summary", index=False)
    
    df_details = pd.DataFrame(verified_open_lrs)
    if df_details.empty:
        df_details = pd.DataFrame(columns=["LR No", "LR Date", "Consignor", "Consignee", "Destination", "Box Qty", "Current Status", "LR Age (Days)"])
    df_details.to_excel(writer, sheet_name="2. Open LRs Details", index=False)
    
    writer.close()
    
    # Style
    wb = load_workbook(report_file)
    apply_styles(wb["1. Branch Summary"], len(df_summary)+1, len(df_summary.columns))
    apply_styles(wb["2. Open LRs Details"], len(df_details)+1, len(df_details.columns))
    wb.save(report_file)
    
    # Email it
    print("Sending Afternoon Open LRs Email...")
    if not SENDER_EMAIL or not RECEIVER_EMAIL:
        print("Missing email credentials. Skipping.")
        return report_file
        
    msg = MIMEMultipart('alternative')
    msg['From'] = SENDER_EMAIL
    msg['To'] = RECEIVER_EMAIL
    msg['Subject'] = f"Afternoon Open LRs Report (Double Checked) - {today_str} 📦"
    
    html_body = f"""
    <html>
      <body style="font-family: Arial, sans-serif; color: #1E293B;">
        <p>Dear User,</p>
        <p>Please find attached the <b>Afternoon Open LRs Report</b> for <b>{today_str} 1:00 PM</b>.</p>
        <p><i>Note: The LRs in this report have been double-checked against the live ERP UI to exclude those that are "Cancelled" or "Despatched from branch".</i></p>
        <p>Total Open LRs verified: <b>{len(verified_open_lrs)}</b></p>
        <br>
        <p><b>Included sheets in the attachment:</b><br>
        1. Branch Summary (Total Open LRs and Max Aging per Branch)<br>
        2. Open LRs Details (Full list of pending LRs)</p>
        <p>Best Regards,<br>
        <b>ERP Daily Automation Engine</b> ⚡</p>
      </body>
    </html>
    """
    msg.attach(MIMEText(html_body, 'html'))
    
    with open(report_file, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename=Afternoon_Open_LRs_Report_{today_str}.xlsx",
        )
        msg.attach(part)
        
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.send_message(msg)
        server.quit()
        print("🎉 Afternoon Open LRs email sent successfully!")
    except Exception as e:
        print("❌ Error sending email:", e)
        
    return report_file
